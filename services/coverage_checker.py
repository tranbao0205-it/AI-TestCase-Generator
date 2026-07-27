"""

Phạm vi CHÍNH của bản này: kiểm tra xem có CHỨC NĂNG nào bị AI bỏ sót
so với những gì thực sự tồn tại (trong ảnh đã OCR, hoặc trong mô tả
text người dùng liệt kê) hay không — và tổng hợp lại thành 1
COVERAGE REPORT (số chức năng áp dụng/đã có/thiếu + % coverage Ở MỨC
CHỨC NĂNG) để hiển thị cho người dùng.
KHÔNG thuộc phạm vi bản 1 (để dành cho bản 2/3 sau này, khi có Rule
Engine/RAG domain cung cấp business rule chuẩn để đối chiếu):
- Số lượng TC trong từng chức năng (vẫn do _validate_testcase_count đảm nhiệm).
- Loại kỹ thuật kiểm thử (Positive/Negative/Boundary/Security...).
- Tính đúng nghiệp vụ của TC theo Business Rule.
- % coverage Ở MỨC LOẠI KỊCH BẢN/TC (coverage_report bản 1 CHỈ tính %
Ở MỨC CHỨC NĂNG — có/thiếu chức năng, không phải có/thiếu TC).
Được tách ra từ ai_service.py (logic gốc: _detect_missing_modules /
_detect_missing_targeted_modules) để dễ maintain/test độc lập và làm
nền cho Rule Engine domain gắn vào sau.
"""

import re
import difflib

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

_STANDARD_CHECK_LABELS = [
    'Thêm mới', 'Quay lại', 'Cập nhật', 'Xóa', 'Phân trang', 'Tìm kiếm', 'Tìm',
]

_SEARCH_BUTTON_PATTERN = re.compile(
    r'^(tìm|nút tìm|button tìm)(\s*(theo|by)\s+.+)?$',
    re.IGNORECASE,
)
_SEARCH_INPUT_PATTERN = re.compile(
    r'^(tìm kiếm|ô tìm kiếm|search)(\s*(theo|by)\s+.+)?$',
    re.IGNORECASE,
)
SEARCH_GENERIC_NAMES = {'tìm kiếm', 'ô tìm kiếm', 'search'}
SEARCH_BUTTON_NAMES = {'tìm', 'nút tìm', 'button tìm'}
_POPUP_ACTION_ALIASES = {
    'sinh mã': {'sinh mã', 'tạo mã', 'generate code', 'generate-code'},
    'hủy bỏ': {'hủy bỏ', 'huỷ bỏ', 'hủy', 'huỷ', 'cancel'},
    'thêm mới và tiếp tục': {
        'thêm mới và tiếp tục', 'lưu và tiếp tục', 'thêm và tiếp tục',
        'save and continue',
    },
    'đóng popup': {'đóng popup', 'đóng', 'close', 'nút x', 'x'},
}
_OUTCOME_SUFFIXES = (' không thành công', ' thành công')
def _strip_outcome_suffix(name: str) -> str:
    n = (name or '').strip().lower()
    for suf in _OUTCOME_SUFFIXES:
        if n.endswith(suf):
            return n[: -len(suf)].strip()
    return n
def _normalize_action_name(name: str) -> str:
    """Chuẩn hóa tên action OCR/chức năng về nhãn dùng chung cho coverage.
    Bóc hậu tố "thành công"/"không thành công" (do Success/Failure Grouping
    ở ai_service.py gắn vào tên module cuối cùng) TRƯỚC khi so khớp alias,
    để "Hủy bỏ thành công" vẫn nhận diện đúng là action "hủy bỏ"."""
    n = re.sub(r'\s+', ' ', (name or '').strip().lower())
    if not n:
        return ''
    n = _strip_outcome_suffix(n)
    n = n.strip('+-×✕✖✓✔←→ ')
    for canonical, aliases in _POPUP_ACTION_ALIASES.items():
        if n in aliases:
            return canonical
    if is_search_button_module(n):
        return 'tìm'
    if is_search_input_module(n):
        return 'tìm kiếm'
    if n in {'xoá', 'delete', 'remove'}:
        return 'xóa'
    if n in {'sửa', 'edit', 'update', 'chỉnh sửa'}:
        return 'cập nhật'
    return n
def _extract_dynamic_action_checks(scanned: str, modules: dict) -> list[dict]:
    """Tạo check động cho các action button thực sự xuất hiện trong popup/form."""
    if not scanned or not isinstance(modules, dict):
        return []
    standard_keys = {
        'thêm mới', 'quay lại', 'cập nhật', 'xóa', 'phân trang',
        'tìm kiếm', 'tìm',
    }
    module_keys = {_normalize_action_name(m) for m in modules.keys()}
    seen: set[str] = set()
    checks: list[dict] = []
    for raw_line in scanned.splitlines():
        line = raw_line.strip()
        if not line.startswith('-') or '|' not in line:
            continue
        parts = [part.strip() for part in line.lstrip('- ').split('|')]
        if len(parts) < 2:
            continue
        name, kind = parts[0], parts[1].lower()
        if not any(token in kind for token in ('button', 'icon-button', 'menu-item')):
            continue

        normalized = _normalize_action_name(name)
        if not normalized or normalized in standard_keys or normalized in seen:
            continue
        if normalized in {'menu item', 'button', 'icon button'}:
            continue

        seen.add(normalized)
        label = {
            'sinh mã': 'Sinh mã',
            'hủy bỏ': 'Hủy bỏ',
            'thêm mới và tiếp tục': 'Thêm mới và tiếp tục',
            'đóng popup': 'Đóng popup',
        }.get(normalized, name.strip())
        checks.append({
            'label': label,
            'applicable': True,
            'covered': normalized in module_keys,
        })

    return checks


def is_search_button_module(name: str) -> bool:
    """True nếu `name` là NÚT "Tìm" (button) — vd "Tìm", "Nút Tìm",
    "Button Tìm", hoặc biến thể bị đặt tên nhầm kèm field ("Tìm theo mã").
    Tự động bỏ qua hậu tố "thành công"/"không thành công" nếu có."""
    return bool(_SEARCH_BUTTON_PATTERN.match(_strip_outcome_suffix(name)))


def is_search_input_module(name: str) -> bool:
    """True nếu `name` là Ô INPUT tìm kiếm — vd "Tìm kiếm", "Ô tìm kiếm",
    "Search" (chung), hoặc "Tìm kiếm theo [X]"/"Search by [X]" (field cụ
    thể). KHÔNG khớp nút "Tìm" (xem is_search_button_module). Tự động bỏ
    qua hậu tố "thành công"/"không thành công" nếu có."""
    return bool(_SEARCH_INPUT_PATTERN.match(_strip_outcome_suffix(name)))


def is_search_module(name: str) -> bool:
    """
    True nếu `name` thuộc BẤT KỲ nhóm nào trong 2 canonical group Search
    (search_input HOẶC search_button). Dùng cho các chỗ chỉ cần biết
    "đây có phải chức năng search nói chung không", KHÔNG dùng để so khớp
    coverage giữa input và button — xem is_search_input_module /
    is_search_button_module / canonical_module_key cho việc đó.
    """
    n = name.strip().lower()
    return is_search_input_module(n) or is_search_button_module(n)


def canonical_module_key(name: str) -> str:
    """
    Trả về khoá canonical dùng để SO KHỚP chức năng — chức năng nào cùng khoá
    canonical được coi là CÙNG 1 chức năng, dù tên hiển thị khác nhau.

    Trả về "search_input" cho ô tìm kiếm, "search_button" cho nút Tìm —
    2 khoá KHÁC NHAU vì đây là 2 UI element khác nhau (xem giải thích ở
    khối canonical group phía trên). Các chức năng khác trả về tên đã lower.

    Các nhóm chức năng không thuộc Search dùng logic canonical riêng trong
    ai_service.py (_MODULE_CANONICAL_GROUPS) — có thể mở rộng hàm này về
    sau nếu cần 1 nguồn canonical chung cho toàn bộ chức năng.
    """
    n = name.strip().lower()
    if is_search_button_module(n):
        return 'search_button'
    if is_search_input_module(n):
        return 'search_input'
    return n


def _names_match_loosely(a: str, b: str) -> bool:
    """
    So khớp lỏng (substring 2 chiều) giữa 2 tên đã lower — dùng ở NHIỀU chỗ
    trong file này để "coi 2 tên là cùng 1 chức năng" khi không thể so
    exact (vd tên bị AI đặt hơi khác giữa các lần gọi).

    GUARD BẮT BUỘC: KHÔNG được coi khớp theo kiểu substring thô như
    `"thêm mới" in "thêm mới và tiếp tục"` — đây là 2 chức năng/button ĐỘC
    LẬP (nút lưu chính vs nút lưu-và-giữ-popup-mở), tuyệt đối không được
    kết luận cùng 1 chức năng chỉ vì 1 tên là substring của tên kia. Áp
    dụng tổng quát: nếu "tiếp tục" xuất hiện ở ĐÚNG 1 trong 2 tên (không
    phải cả 2, không phải không có), coi là KHÔNG khớp — bất kể phần còn
    lại có là substring của nhau hay không.
    """
    a = (a or '').strip()
    b = (b or '').strip()
    if not a or not b:
        return False
    if a == b:
        return True
    if ('tiếp tục' in a) != ('tiếp tục' in b):
        return False
    return a in b or b in a


def _evaluate_standard_checks(scanned: str, modules: dict) -> list[dict]:
    """
    Helper DÙNG CHUNG cho detect_missing_modules() và build_coverage_report()
    — tránh viết lặp 2 lần cùng 1 logic nhận diện (bài học đã ghi nhận:
    logic trùng lặp giữa 2 hàm rất dễ bị lệch nhau khi sửa sau này).

    Trả về list các dict {"label": str, "applicable": bool, "covered": bool}
    cho các check chuẩn và action động trong popup/form (Thêm mới/Quay lại/Cập nhật/Xóa/Phân trang/Tìm).
    "applicable" = chức năng này CÓ xuất hiện trong ảnh (scanned) không.
    "covered" = ĐÃ có chức năng tương ứng trong kết quả AI sinh ra chưa.
    """
    if not scanned or not isinstance(modules, dict):
        return [
            {"label": label, "applicable": False, "covered": False}
            for label in _STANDARD_CHECK_LABELS
        ]

    scanned_lower = scanned.lower()
    mod_names_lower = [m.strip().lower() for m in modules.keys()]
    existing_blob = ' '.join(mod_names_lower)

    def _has_any(*keywords: str) -> bool:
        return any(kw in existing_blob for kw in keywords)

    checks = []

    def _has_them_moi() -> bool:
        return any('thêm' in m and 'tiếp tục' not in m for m in mod_names_lower)
    applicable = 'thêm mới' in scanned_lower
    checks.append({
        "label": "Thêm mới",
        "applicable": applicable,
        "covered": _has_them_moi() if applicable else False,
    })

    applicable = 'quay lại' in scanned_lower
    checks.append({
        "label": "Quay lại",
        "applicable": applicable,
        "covered": _has_any('quay lại') if applicable else False,
    })

    applicable = bool('cập nhật' in scanned_lower or re.search(r'-\s*sửa\b', scanned_lower))
    checks.append({
        "label": "Cập nhật",
        "applicable": applicable,
        "covered": _has_any('cập nhật', 'sửa') if applicable else False,
    })

    applicable = bool('xóa' in scanned_lower or 'xoá' in scanned_lower)
    checks.append({
        "label": "Xóa",
        "applicable": applicable,
        "covered": _has_any('xóa', 'xoá') if applicable else False,
    })

    applicable = 'phân trang' in scanned_lower
    checks.append({
        "label": "Phân trang",
        "applicable": applicable,
        "covered": _has_any('trang') if applicable else False,
    })
    applicable = bool(
        re.search(r'\btìm kiếm\b', scanned_lower)
        or re.search(r'\bsearch\b', scanned_lower)
    )
    checks.append({
        "label": "Tìm kiếm",
        "applicable": applicable,
        "covered": (
            any(is_search_input_module(n) for n in mod_names_lower)
            if applicable else False
        ),
    })
    applicable = bool(
        re.search(r'-\s*tìm\s*\|', scanned_lower)
        or re.search(r'\bnút tìm\b', scanned_lower)
    )
    checks.append({
        "label": "Tìm",
        "applicable": applicable,
        "covered": (
            any(is_search_button_module(n) for n in mod_names_lower)
            if applicable else False
        ),
    })

    checks.extend(_extract_dynamic_action_checks(scanned, modules))
    return checks


def detect_missing_modules(scanned: str, modules: dict) -> list[str]:
    """
    Dùng khi CÓ ẢNH (có kết quả OCR trong `scanned`).

    Phát hiện chức năng UI xuất hiện trong ảnh (scanned) nhưng KHÔNG có
    chức năng tương ứng nào trong kết quả AI đã sinh.

    LƯU Ý: "Tìm kiếm theo [X]" (ô lọc), "Tìm kiếm"/"Search" (ô lọc chung),
    và "Tìm" (nút) đều thuộc CÙNG 1 canonical group Search (xem
    is_search_module() / canonical_module_key() ở đầu file) — chức năng nào
    thuộc group này cũng được coi là "đã có chức năng search", KHÔNG đòi
    hỏi khớp đúng tên/synonym cụ thể. Việc AI có tách chức năng search
    thành 1 hay nhiều chức năng (vd chức năng riêng cho nút "Tìm" và chức năng
    riêng cho ô lọc) là quyết định ở tầng generate/normalize (ai_service.py)
    — Coverage Checker chỉ quan tâm "đã có chức năng search hay chưa", không
    phạt/đòi thêm chức năng nếu thiếu 1 trong các synonym.

    Args:
        scanned: Kết quả OCR (mỗi dòng element bắt đầu bằng "-").
        modules: Dict {tên_module: [test_case, ...]} AI đã sinh ra.

    Returns:
        Danh sách tên chức năng (theo cách gọi chuẩn hoá) bị thiếu.
    """
    checks = _evaluate_standard_checks(scanned, modules)
    return [c["label"] for c in checks if c["applicable"] and not c["covered"]]


def detect_missing_targeted_modules(description: str, modules: dict) -> list[str]:
    """
    Dùng khi KHÔNG CÓ ẢNH (mô tả text thuần) — vì khi đó `scanned` luôn
    rỗng nên detect_missing_modules không detect được gì.

    So khớp từng cụm chức năng user liệt kê trong description (tách
    theo dấu phẩy/xuống dòng, bỏ phần role đứng trước dấu ":") với tên
    chức năng đã sinh ra, bằng substring 2 chiều + fuzzy ratio (difflib).
    Cụm nào không khớp chức năng nào → coi là bị AI bỏ sót.

    Args:
        description: Mô tả text người dùng nhập (liệt kê chức năng).
        modules: Dict {tên_module: [test_case, ...]} AI đã sinh ra.

    Returns:
        Danh sách cụm mô tả chức năng bị thiếu (nguyên văn từ description).
    """
    _matched, missing = _match_targeted_phrases(description, modules)
    return missing


def _match_targeted_phrases(description: str, modules: dict) -> tuple[list[str], list[str]]:
    """
    Helper DÙNG CHUNG cho detect_missing_targeted_modules() và
    build_coverage_report() — tách phrase hợp lệ từ description rồi phân
    loại matched/missing, tránh lặp logic parse 2 lần.

    Returns:
        (matched_phrases, missing_phrases) — cả 2 đều là cụm nguyên văn
        từ description (đã strip), KHÔNG gồm các cụm bị lọc bởi stop_phrases
        hoặc quá ngắn (<3 ký tự).
    """
    if not description or not isinstance(modules, dict):
        return [], []

    valid_modules = {
        m: tcs for m, tcs in modules.items()
        if isinstance(tcs, list) and tcs
    }
    module_names_lower = [m.strip().lower() for m in valid_modules.keys()]
    desc = description
    if ':' in desc:
        desc = desc.rsplit(':', 1)[-1]

    phrases = [p.strip() for p in re.split(r'[,\n;]', desc) if p.strip()]
    stop_phrases = {
        'admin', 'employee', 'user', 'customer', 'hr', 'khách hàng',
        'nhân viên', 'quản trị viên', 'người dùng',
    }

    matched: list[str] = []
    missing: list[str] = []
    for phrase in phrases:
        p_lower = phrase.lower().strip()
        if len(p_lower) < 3 or p_lower in stop_phrases:
            continue
        if not module_names_lower:
            missing.append(phrase)
            continue
        if is_search_input_module(p_lower) and any(is_search_input_module(m) for m in module_names_lower):
            matched.append(phrase)
            continue
        if is_search_button_module(p_lower) and any(is_search_button_module(m) for m in module_names_lower):
            matched.append(phrase)
            continue

        is_matched = any(
            ('tiếp tục' in p_lower) == ('tiếp tục' in m_lower)
            and (
                p_lower in m_lower or m_lower in p_lower
                or difflib.SequenceMatcher(None, p_lower, m_lower).ratio() >= 0.55
            )
            for m_lower in module_names_lower
        )
        (matched if is_matched else missing).append(phrase)

    return matched, missing


def build_coverage_report(
    modules: dict,
    scanned: str | None = None,
    description: str | None = None,
    has_image: bool = True,
) -> dict:
    """
    Tổng hợp Coverage Report Ở MỨC CHỨC NĂNG (KHÔNG phải % TC/loại kịch bản
    — xem giới hạn phạm vi ở docstring đầu file).

    Dùng ĐÚNG 1 trong 2 nguồn theo has_image (giống cách _enforce_min_coverage
    trong ai_service.py chọn nhánh):
    - has_image=True  → dùng `scanned` (kết quả OCR), check các chức năng chuẩn và action button trong popup/form
      chuẩn (Thêm mới/Quay lại/Cập nhật/Xóa/Phân trang/Tìm).
    - has_image=False → dùng `description` (text thuần), check từng cụm
      chức năng user liệt kê.

    Args:
        modules: Dict {tên_module: [test_case, ...]} AI đã sinh ra.
        scanned: Kết quả OCR — bắt buộc nếu has_image=True.
        description: Mô tả text người dùng — bắt buộc nếu has_image=False.
        has_image: Chọn nguồn đối chiếu (xem trên).

    Returns:
        {
            "total_expected": int,      
            "total_covered": int,       
            "total_missing": int,
            "coverage_percent": float, 
            "covered_items": list[str],
            "missing_items": list[str],
            "source": "image" | "text",
        }
    """
    modules = modules if isinstance(modules, dict) else {}

    if has_image:
        checks = _evaluate_standard_checks(scanned or '', modules)
        applicable_checks = [c for c in checks if c["applicable"]]
        covered_items = [c["label"] for c in applicable_checks if c["covered"]]
        missing_items = [c["label"] for c in applicable_checks if not c["covered"]]
        source = "image"
    else:
        covered_items, missing_items = _match_targeted_phrases(description or '', modules)
        source = "text"

    total_expected = len(covered_items) + len(missing_items)
    total_covered = len(covered_items)
    coverage_percent = (
        round(total_covered / total_expected * 100, 1) if total_expected else 100.0
    )

    return {
        "total_expected": total_expected,
        "total_covered": total_covered,
        "total_missing": len(missing_items),
        "coverage_percent": coverage_percent,
        "covered_items": covered_items,
        "missing_items": missing_items,
        "source": source,
    }
_SAVE_ACTION_KEYWORDS = ('thêm mới', 'cập nhật', 'sửa')
def _text_blob_for_module(tcs) -> str:
    """Gộp toàn bộ nội dung text (scenario/description/title/expected_result/
    then/steps/test_data) của 1 danh sách TC thành 1 chuỗi lowercase duy nhất
    để so khớp field/business-rule/workflow theo từ khóa."""
    if not isinstance(tcs, list):
        return ''
    parts = []
    for tc in tcs:
        if not isinstance(tc, dict):
            continue
        for key in ('scenario', 'description', 'title', 'expected_result', 'then', 'steps', 'test_data'):
            value = tc.get(key)
            if value:
                parts.append(str(value))
    return ' '.join(parts).lower()
def _find_action_module_tcs(modules: dict, action_keywords: tuple) -> list:
    """Trả về TC list của chức năng ĐẦU TIÊN khớp 1 trong action_keywords (substring,
    lower) — dùng để lấy đúng chức năng lưu (Thêm mới/Cập nhật) làm nơi đối chiếu
    field/business-rule/workflow coverage."""
    if not isinstance(modules, dict):
        return []
    for name, tcs in modules.items():
        low = name.strip().lower()
        for kw in action_keywords:
            if kw in low and not ('tiếp tục' in low and 'tiếp tục' not in kw):
                if isinstance(tcs, list):
                    return tcs
                break
    return []


def _name_of(item) -> str:
    """Lấy tên hiển thị từ 1 phần tử field/button — chấp nhận cả dict
    {"name": ...} lẫn string thô, để chịu lỗi nếu AI trả JSON hơi khác cấu
    trúc schema mong đợi."""
    if isinstance(item, dict):
        return str(item.get('name') or '').strip()
    return str(item or '').strip()


def detect_missing_fields(form_structure: dict, modules: dict) -> list[str]:
    """
    Field Coverage: mỗi field khai báo trong form_structure['fields'] phải
    được nhắc tới (theo tên) trong nội dung TC của chức năng hành động LƯU
    tương ứng (Thêm mới/Cập nhật) — field KHÔNG bao giờ có chức năng riêng nên
    KHÔNG so theo tên chức năng.
    Trả về danh sách tên field bị thiếu.
    """
    if not isinstance(form_structure, dict):
        return []
    fields = form_structure.get('fields') or []
    if not fields:
        return []
    blob = _text_blob_for_module(_find_action_module_tcs(modules, _SAVE_ACTION_KEYWORDS))
    missing = []
    for f in fields:
        name = _name_of(f)
        if not name:
            continue
        if name.lower() not in blob:
            missing.append(name)
    return missing


def detect_missing_buttons(form_structure: dict, modules: dict) -> list[str]:
    """
    Button Coverage: mỗi button khai báo trong form_structure['buttons']
    phải có chức năng riêng tương ứng (so khớp theo canonical action name, dùng
    chung _normalize_action_name với _POPUP_ACTION_ALIASES ở trên để không
    lệch tiêu chí với _extract_dynamic_action_checks).
    Trả về danh sách tên button bị thiếu.
    """
    if not isinstance(form_structure, dict):
        return []
    buttons = form_structure.get('buttons') or []
    if not buttons:
        return []
    module_keys = {_normalize_action_name(m) for m in modules.keys()} if isinstance(modules, dict) else set()
    missing = []
    for b in buttons:
        name = _name_of(b)
        if not name:
            continue
        key = _normalize_action_name(name)
        if not key:
            continue
        covered = key in module_keys or any(_names_match_loosely(key, mk) for mk in module_keys if mk)
        if not covered:
            missing.append(name)
    return missing


def detect_missing_business_rules(form_structure: dict, modules: dict) -> list[str]:
    """
    Business Rule Coverage: mỗi rule khai báo trong
    form_structure['business_rules'] phải được phản ánh (đa số từ khóa nội
    dung xuất hiện) trong TC của chức năng hành động LƯU tương ứng.
    So khớp bằng tỉ lệ từ khóa có nghĩa (>=3 ký tự) xuất hiện trong blob —
    KHÔNG cần khớp nguyên văn (rule do AI diễn giải lại, câu chữ có thể khác
    TC thật).
    """
    if not isinstance(form_structure, dict):
        return []
    rules = form_structure.get('business_rules') or []
    if not rules:
        return []
    blob = _text_blob_for_module(_find_action_module_tcs(modules, _SAVE_ACTION_KEYWORDS))
    missing = []
    for rule in rules:
        rule_text = str(rule).strip()
        if not rule_text:
            continue
        keywords = [w for w in re.split(r'[\s,.\-–]+', rule_text.lower()) if len(w) >= 3]
        if not keywords:
            continue
        hit_ratio = sum(1 for kw in keywords if kw in blob) / len(keywords)
        if hit_ratio < 0.3:
            missing.append(rule_text)
    return missing


def detect_missing_workflow_steps(form_structure: dict, modules: dict) -> list[str]:
    """
    Workflow Coverage: mỗi bước trong form_structure['workflow'] phải được
    phản ánh trong TC của chức năng hành động LƯU tương ứng (thứ tự thao tác
    end-to-end của TC "thành công").
    """
    if not isinstance(form_structure, dict):
        return []
    steps = form_structure.get('workflow') or []
    if not steps:
        return []
    blob = _text_blob_for_module(_find_action_module_tcs(modules, _SAVE_ACTION_KEYWORDS))
    missing = []
    for step in steps:
        step_text = str(step).strip()
        if not step_text:
            continue
        keywords = [w for w in re.split(r'[\s,.\-–]+', step_text.lower()) if len(w) >= 3]
        if not keywords:
            continue
        hit_ratio = sum(1 for kw in keywords if kw in blob) / len(keywords)
        if hit_ratio < 0.25:
            missing.append(step_text)
    return missing


def build_form_structure_coverage(form_structure: dict, modules: dict) -> dict:
    """
    Tổng hợp Coverage theo 4 CHIỀU MỚI dựa trên Form Structure đã phân tích
    (form_name/fields/buttons/business_rules/workflow) — BỔ SUNG cho
    build_coverage_report() (vốn CHỈ kiểm tra Ở MỨC CHỨC NĂNG).

    Trả về {} nếu form_structure rỗng/không hợp lệ (màn hình không phải
    Form/Modal, hoặc bước Form Understanding bị tắt/lỗi).
    """
    if not isinstance(form_structure, dict) or not form_structure:
        return {}
    modules = modules if isinstance(modules, dict) else {}

    missing_fields = detect_missing_fields(form_structure, modules)
    missing_buttons = detect_missing_buttons(form_structure, modules)
    missing_business_rules = detect_missing_business_rules(form_structure, modules)
    missing_workflow = detect_missing_workflow_steps(form_structure, modules)

    def _pct(total: int, missing_count: int) -> float:
        if total <= 0:
            return 100.0
        return round((total - missing_count) / total * 100, 1)

    n_fields = len(form_structure.get('fields') or [])
    n_buttons = len(form_structure.get('buttons') or [])
    n_rules = len(form_structure.get('business_rules') or [])
    n_workflow = len(form_structure.get('workflow') or [])

    return {
        'form_name': form_structure.get('form_name', ''),
        'field_coverage_percent': _pct(n_fields, len(missing_fields)),
        'missing_fields': missing_fields,
        'button_coverage_percent': _pct(n_buttons, len(missing_buttons)),
        'missing_buttons': missing_buttons,
        'business_rule_coverage_percent': _pct(n_rules, len(missing_business_rules)),
        'missing_business_rules': missing_business_rules,
        'workflow_coverage_percent': _pct(n_workflow, len(missing_workflow)),
        'missing_workflow': missing_workflow,
    }


def write_coverage_report_sheet(workbook, report: dict, sheet_name: str = "Coverage Report"):
    """
    Ghi 1 sheet "Coverage Report" vào workbook openpyxl ĐÃ MỞ SẴN (dùng để
    nối vào ĐÚNG bước xuất file .xlsx hiện có của project — hàm này KHÔNG
    tự tạo/lưu file, chỉ thêm sheet vào workbook rồi trả lại worksheet vừa
    tạo; nơi gọi vẫn tự `wb.save(...)` như cũ).
    Args:
        workbook: đối tượng openpyxl.Workbook đang mở (vd wb = openpyxl.Workbook()
            hoặc workbook đang ghi các sheet TC khác).
        report: dict trả về từ build_coverage_report().
        sheet_name: tên sheet, mặc định "Coverage Report".
    Returns:
        Worksheet vừa tạo, hoặc None nếu thiếu openpyxl / report rỗng.
    """
    if not _OPENPYXL_AVAILABLE:
        print(
            "[CoverageReport] openpyxl chưa cài (pip install openpyxl) — "
            "bỏ qua bước ghi sheet 'Coverage Report'."
        )
        return None
    if not isinstance(report, dict):
        return None

    ws = workbook.create_sheet(title=sheet_name)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    label_font = Font(bold=True)
    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    ok_font = Font(color="006100")
    bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bad_font = Font(color="9C0006")

    ws["A1"] = "COVERAGE REPORT"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    source_label = "Ảnh (OCR)" if report.get("source") == "image" else "Mô tả text"
    summary_rows = [
        ("Nguồn đối chiếu", source_label),
        ("Tổng số chức năng áp dụng", report.get("total_expected", 0)),
        ("Đã có chức năng", report.get("total_covered", 0)),
        ("Còn thiếu", report.get("total_missing", 0)),
        ("Coverage (%)", f'{report.get("coverage_percent", 0)}%'),
    ]
    row = 3
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Chức năng").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="Trạng thái").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    header_row = row
    row += 1

    for item in report.get("covered_items", []):
        ws.cell(row=row, column=1, value=item)
        status_cell = ws.cell(row=row, column=2, value="Đã có")
        status_cell.fill = ok_fill
        status_cell.font = ok_font
        row += 1

    for item in report.get("missing_items", []):
        ws.cell(row=row, column=1, value=item)
        status_cell = ws.cell(row=row, column=2, value="Thiếu")
        status_cell.fill = bad_fill
        status_cell.font = bad_font
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    for r in (header_row,):
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")

    return ws
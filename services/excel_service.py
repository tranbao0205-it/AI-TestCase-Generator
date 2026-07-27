"""
Excel Service - Creates professionally formatted .xlsx test case files.
Uses openpyxl to generate multi-sheet workbooks with styling.
"""

import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def _ensure_string(value) -> str:
    """Convert values to text safely so Vietnamese is preserved in Excel cells."""
    if value is None:
        return ''
    if isinstance(value, str):
        return value
    return str(value)


def _slugify(text: str) -> str:
    """Convert Vietnamese text to ASCII-safe filename (no diacritics)."""
    import unicodedata
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    return unicodedata.normalize('NFC', text)


class ExcelService:
    COLOR_TITLE_BG    = '1E3A8A'   
    COLOR_HEADER_BG   = '2563EB'   
    COLOR_SUBHEAD_BG  = '3B82F6'   
    COLOR_ALT_ROW     = 'EFF6FF'   
    COLOR_WHITE       = 'FFFFFF'
    COLOR_FONT_LIGHT  = 'FFFFFF'
    COLOR_HIGH        = 'DC2626'   
    COLOR_MEDIUM      = 'D97706'   
    COLOR_LOW         = '16A34A'   
    COLOR_PASSED      = '16A34A'
    COLOR_FAILED      = 'DC2626'
    COLOR_NOT_RUN     = '6B7280'
    def __init__(self):
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.output_folder = os.path.join(base_dir, 'outputs')
        os.makedirs(self.output_folder, exist_ok=True)
    @staticmethod
    def _is_valid_tc(tc) -> bool:
        """
        Một TC được coi là hợp lệ nếu có ít nhất 1 trong 3 field cốt lõi
        (Tên Test Case / Mô tả / Kết quả mong đợi) có dữ liệu thật.
        Dùng để lọc bỏ các dict rỗng/lỗi trước khi đếm & ghi, tránh tình
        trạng sheet "Tổng hợp" báo số TC khác với số dòng thực tế trong
        từng sheet module (STT có số nhưng không có nội dung).
        """
        if not isinstance(tc, dict):
            return False
        title = (tc.get('feature') or tc.get('title') or '').strip()
        desc = (tc.get('scenario') or tc.get('description') or '').strip()
        expected = (tc.get('expected_result') or '').strip()
        return bool(title or desc or expected)

    def create_excel(self, test_cases_data: dict, project_name: str = 'Project') -> str:
        """Build an Excel workbook and return the saved filename."""
        wb = Workbook()
        wb.remove(wb.active) 

        if not isinstance(test_cases_data, dict):
            raise ValueError('Dữ liệu test case phải là object JSON')

        raw_modules = test_cases_data.get('modules', {})
        if not isinstance(raw_modules, dict):
            raw_modules = {}
        modules = {
            module_name: [tc for tc in tcs if self._is_valid_tc(tc)]
            for module_name, tcs in raw_modules.items()
            if isinstance(module_name, str) and not module_name.startswith('_')
            if isinstance(tcs, list)
        }
        description = test_cases_data.get('description', '')
        now = datetime.now()
        total_tc = sum(len(v) for v in modules.values())
        self._create_general_info_sheet(wb, project_name, description, total_tc, now)
        self._create_summary_sheet(wb, modules)
        for module_name, test_cases in modules.items():
            self._create_module_sheet(wb, module_name, test_cases)
        project_name = re.sub(r'===.*?===', '', project_name, flags=re.DOTALL).strip()
        project_name = re.sub(r'^(HUONG DAN|HƯỚNG DẪN|Phan tich|Phân tích toàn bộ|IMAGE_GUIDE).*', '', project_name, flags=re.IGNORECASE | re.DOTALL).strip()
        project_name = project_name[:60] or 'My Project'
        safe_name = _slugify(project_name)
        safe_name = re.sub(r'[\/*?:"<>|]', '_', safe_name).strip().replace(' ', '_')
        safe_name = re.sub(r'_+', '_', safe_name).strip('_') 
        safe_name = safe_name or 'Project'
        date_str  = now.strftime('%Y%m%d_%H%M%S')
        filename  = f'{safe_name}_{date_str}.xlsx'
        filepath  = os.path.join(self.output_folder, filename)
        wb.save(filepath)
        return filename
    def _thin_border(self, color: str = 'CBD5E1') -> Border:
        side = Side(style='thin', color=color)
        return Border(left=side, right=side, top=side, bottom=side)
    def _medium_border(self, color: str = '1E40AF') -> Border:
        side = Side(style='medium', color=color)
        return Border(left=side, right=side, top=side, bottom=side)
    def _fill(self, hex_color: str) -> PatternFill:
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
    def _header_font(self, size: int = 11, bold: bool = True,
                     color: str = 'FFFFFF') -> Font:
        return Font(name='Calibri', size=size, bold=bold, color=color)
    def _data_font(self, size: int = 10, bold: bool = False,
                   color: str = '1E293B') -> Font:
        return Font(name='Calibri', size=size, bold=bold, color=color)
    def _center_align(self, wrap: bool = True) -> Alignment:
        return Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    def _left_align(self, wrap: bool = True) -> Alignment:
        return Alignment(horizontal='left', vertical='top', wrap_text=wrap)
    def _create_general_info_sheet(self, wb: Workbook, project_name: str,
                                   description: str, total_tc: int,
                                   now: datetime):
        ws = wb.create_sheet('Thong tin chung')
        ws.sheet_view.showGridLines = False
        ws.merge_cells('A1:C1')
        ws['A1'].value     = 'THÔNG TIN CHUNG DỰ ÁN'
        ws['A1'].fill      = self._fill('7C3AED')
        ws['A1'].font      = self._header_font(size=16, color=self.COLOR_FONT_LIGHT)
        ws['A1'].alignment = self._center_align()
        ws.row_dimensions[1].height = 45

        rows = [
            ('Mã dự án',       f'PRJ-{now.strftime("%Y%m%d")}'),
            ('Tên dự án',      _ensure_string(project_name)),
            ('Mô tả dự án',    _ensure_string(description) if description and description.strip() else _ensure_string(project_name)),
            ('Người lập',      'AI Test Case Generator'),
            ('Ngày đánh giá',  now.strftime('%d/%m/%Y')),
            ('Tổng test case', str(total_tc)),
            ('Passed',         '0'),
            ('Failed',         '0'),
            ('Not Run',        str(total_tc)),
            ('Ghi chú',        ''),
        ]

        for i, (label, value) in enumerate(rows, start=2):
            lc = ws.cell(row=i, column=1, value=_ensure_string(label))
            lc.fill      = self._fill(self.COLOR_HEADER_BG)
            lc.font      = self._header_font(size=11)
            lc.alignment = Alignment(horizontal='left', vertical='center',
                                     indent=1, wrap_text=True)
            lc.border    = self._thin_border()
            ws.merge_cells(f'B{i}:C{i}')
            vc = ws.cell(row=i, column=2, value=_ensure_string(value))
            vc.font      = self._data_font(size=11)
            vc.alignment = Alignment(horizontal='left', vertical='center',
                                     indent=1, wrap_text=True)
            vc.border    = self._thin_border()
            ws.row_dimensions[i].height = 28

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20

    def _create_summary_sheet(self, wb: Workbook, modules: dict):
        ws = wb.create_sheet('Tong hop')
        ws.sheet_view.showGridLines = False
        ws.merge_cells('A1:H1')
        ws['A1'].value     = 'BẢNG TỔNG HỢP TEST CASE THEO MODULE'
        ws['A1'].fill      = self._fill('7C3AED')
        ws['A1'].font      = self._header_font(size=15, color=self.COLOR_FONT_LIGHT)
        ws['A1'].alignment = self._center_align()
        ws.row_dimensions[1].height = 40
        headers = ['STT', 'Tên Sheet / Module', 'Tổng TC', 'Passed',
                   'Failed', 'Not Run', 'Tỷ lệ HT', 'Ghi chú']
        col_widths = [7, 35, 12, 12, 12, 12, 14, 25]

        for col, (hdr, w) in enumerate(zip(headers, col_widths), start=1):
            c = ws.cell(row=2, column=col, value=hdr)
            c.fill      = self._fill(self.COLOR_HEADER_BG)
            c.font      = self._header_font()
            c.alignment = self._center_align()
            c.border    = self._thin_border()
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[2].height = 30
        total_all = 0
        for i, (module_name, test_cases) in enumerate(modules.items(), start=1):
            row = i + 2
            count = len(test_cases)
            total_all += count
            bg = self.COLOR_ALT_ROW if i % 2 == 0 else self.COLOR_WHITE
            values = [i, module_name, count, 0, 0, count, '0%', '']
            for col, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=_ensure_string(val))
                c.fill      = self._fill(bg)
                c.font      = self._data_font()
                c.alignment = self._center_align() if col != 2 else \
                               Alignment(horizontal='left', vertical='center',
                                         indent=1, wrap_text=True)
                c.border    = self._thin_border()
            ws.row_dimensions[row].height = 22
        tr = len(modules) + 3
        ws.merge_cells(f'A{tr}:B{tr}')
        tc = ws.cell(row=tr, column=1, value='TỔNG CỘNG')
        tc.fill      = self._fill(self.COLOR_SUBHEAD_BG)
        tc.font      = self._header_font()
        tc.alignment = self._center_align()
        tc.border    = self._thin_border()
        for col in range(3, 9):
            c = ws.cell(row=tr, column=col,
                        value=total_all if col == 3 else '')
            c.fill      = self._fill(self.COLOR_SUBHEAD_BG)
            c.font      = self._header_font()
            c.alignment = self._center_align()
            c.border    = self._thin_border()
        ws.row_dimensions[tr].height = 26
    def _create_module_sheet(self, wb: Workbook, module_name: str,
                             test_cases: list):
        safe_name = re.sub(r'[\\/*?:\[\]]', '_', module_name)[:31]
        ws = wb.create_sheet(safe_name)
        ws.sheet_view.showGridLines = False
        num_cols = 14
        ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
        ws['A1'].value     = _ensure_string(f'TEST CASES – {module_name.upper()}')
        ws['A1'].fill      = self._fill(self.COLOR_TITLE_BG)
        ws['A1'].font      = self._header_font(size=14)
        ws['A1'].alignment = self._center_align()
        ws.row_dimensions[1].height = 38
        ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
        first_feature = ''
        if test_cases and isinstance(test_cases[0], dict):
            first_feature = test_cases[0].get('feature') or test_cases[0].get('scenario') or ''
        sub_info = f'Màn hình: {module_name}  |  Chức năng: {first_feature}' if first_feature else f'Màn hình: {module_name}'
        ws['A2'].value     = sub_info
        ws['A2'].fill      = self._fill('1E40AF')
        ws['A2'].font      = Font(name='Calibri', size=10, italic=True, color='DBEAFE')
        ws['A2'].alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
        ws.row_dimensions[2].height = 20
        columns = [
            ('STT',                    6),
            ('Mã TC',                 16),
            ('Tên Test Case',         32),
            ('Mô tả',                 40),
            ('Điều kiện tiên quyết',  32),
            ('Các bước thực hiện',    50),
            ('Dữ liệu đầu vào',       30),
            ('Kết quả mong đợi',      45),
            ('Kết quả thực tế',       30),
            ('Trạng thái',            13),
            ('Mức độ ưu tiên',        14),
            ('Loại test',             22),
            ('Ghi chú',               22),
            ('Người test',            16),
        ]

        for col, (hdr, w) in enumerate(columns, start=1):
            c = ws.cell(row=3, column=col, value=hdr)
            c.fill      = self._fill(self.COLOR_SUBHEAD_BG)
            c.font      = self._header_font(size=10)
            c.alignment = self._center_align()
            c.border    = self._thin_border()
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[3].height = 32
        for i, tc in enumerate(test_cases, start=1):
            row = i + 3
            bg  = self.COLOR_ALT_ROW if i % 2 == 0 else self.COLOR_WHITE
            steps_val = _ensure_string(tc.get('steps', ''))
            if not steps_val.strip():
                parts = []
                if tc.get('given'):
                    parts.append(f"Given: {tc['given']}")
                if tc.get('when'):
                    parts.append(f"When: {tc['when']}")
                if tc.get('then'):
                    parts.append(f"Then: {tc['then']}")
                steps_val = '\n'.join(parts)
            title_val = (
                tc.get('feature') or
                tc.get('title') or
                tc.get('scenario') or ''
            )
            desc_val = tc.get('scenario') or tc.get('description') or ''

            row_data = [
                i,                                                  # STT
                tc.get('id', ''),                                   # Mã TC
                title_val,                                          # Tên Test Case
                desc_val,                                           # Mô tả
                tc.get('precondition', ''),                         # Điều kiện tiên quyết
                steps_val,                                          # Các bước thực hiện
                tc.get('test_data', ''),                            # Dữ liệu đầu vào
                tc.get('expected_result', ''),                      # Kết quả mong đợi
                tc.get('actual_result', ''),                        # Kết quả thực tế
                tc.get('status', 'Chưa chạy'),                      # Trạng thái
                tc.get('priority', 'Trung bình'),                   # Mức độ ưu tiên
                tc.get('test_type', 'Kiểm thử chức năng'),          # Loại test
                tc.get('note', ''),                                 # Ghi chú
                tc.get('tester', ''),                               # Người test
            ]
            CENTER_COLS = {1, 2, 10, 11, 12}
            for col, val in enumerate(row_data, start=1):
                c = ws.cell(row=row, column=col, value=_ensure_string(val) if val else '')
                c.fill   = self._fill(bg)
                c.border = self._thin_border()

                if col in CENTER_COLS:
                    c.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                else:
                    c.alignment = self._left_align()
                if col == 11:
                    pmap = {
                        'Cao': self.COLOR_HIGH,
                        'Trung bình': self.COLOR_MEDIUM,
                        'Thấp': self.COLOR_LOW,
                    }
                    c.font = Font(name='Calibri', size=10, bold=True,
                                  color=pmap.get(str(val), self.COLOR_MEDIUM))
                elif col == 10:
                    smap = {
                        'Passed': self.COLOR_PASSED,   'Đạt': self.COLOR_PASSED,
                        'Failed': self.COLOR_FAILED,   'Không đạt': self.COLOR_FAILED,
                        'Not Run': self.COLOR_NOT_RUN, 'Chưa chạy': self.COLOR_NOT_RUN,
                        'Bị chặn': self.COLOR_NOT_RUN,
                    }
                    c.font = Font(name='Calibri', size=10, bold=True,
                                  color=smap.get(str(val), self.COLOR_NOT_RUN))
                else:
                    c.font = self._data_font()
            ws.row_dimensions[row].height = 70
        if test_cases:
            dv = DataValidation(
                type='list',
                formula1='"Đạt,Không đạt,Chưa chạy,Bị chặn"',
                allow_blank=True,
                showDropDown=False,
            )
            ws.add_data_validation(dv)
            last_row = len(test_cases) + 3
            dv.add(f'J4:J{last_row}')
        ws.freeze_panes = 'A4'
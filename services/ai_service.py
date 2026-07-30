"""
AI Service - Connects to OpenAI API to generate structured test cases.
Implements retry logic and JSON validation.
"""

import os
import re
import time
import json
import hashlib
import difflib
from pathlib import Path
try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:  
    _OPENPYXL_AVAILABLE = False

try:
    from json_repair import repair_json as _repair_json
    _JSON_REPAIR_AVAILABLE = True
except ImportError:
    _JSON_REPAIR_AVAILABLE = False
from openai import OpenAI
from dotenv import load_dotenv

try:
    from services.vision_service import VisionService
    from services.workflow_service import WorkflowService
except Exception:
    from vision_service import VisionService
    from workflow_service import WorkflowService

try:
    from services.rag_service import RAGService
    print("[RAG] Import RAGService từ 'services.rag_service' thành công.")
except Exception as _exc_services_pkg:
    try:
        from rag_service import RAGService
        print("[RAG] Import RAGService từ 'rag_service' (cùng cấp) thành công.")
    except Exception as _exc_flat:
        print(
            "[RAG] KHÔNG import được RAGService — RAG sẽ bị TẮT hoàn toàn.\n"
            f"  - Thử 'services.rag_service': {_exc_services_pkg}\n"
            f"  - Thử 'rag_service' (cùng cấp): {_exc_flat}"
        )
        RAGService = None

try:
    from services.coverage_checker import (
        detect_missing_modules as _cc_detect_missing_modules,
        detect_missing_targeted_modules as _cc_detect_missing_targeted_modules,
        build_coverage_report as _cc_build_coverage_report,
        build_form_structure_coverage as _cc_build_form_structure_coverage,
        is_search_module as _cc_is_search_module,
        canonical_module_key as _cc_canonical_module_key,
        SEARCH_GENERIC_NAMES as _cc_search_generic_names,
        SEARCH_BUTTON_NAMES as _cc_search_button_names,
    )
    print("[CoverageChecker] Import từ 'services.coverage_checker' thành công.")
except Exception as _exc_cc_services_pkg:
    try:
        from coverage_checker import (
            detect_missing_modules as _cc_detect_missing_modules,
            detect_missing_targeted_modules as _cc_detect_missing_targeted_modules,
            build_coverage_report as _cc_build_coverage_report,
            build_form_structure_coverage as _cc_build_form_structure_coverage,
            is_search_module as _cc_is_search_module,
            canonical_module_key as _cc_canonical_module_key,
            SEARCH_GENERIC_NAMES as _cc_search_generic_names,
            SEARCH_BUTTON_NAMES as _cc_search_button_names,
        )
        print("[CoverageChecker] Import từ 'coverage_checker' (cùng cấp) thành công.")
    except Exception as _exc_cc_flat:
        print(
            "[CoverageChecker] KHÔNG import được coverage_checker — sẽ dùng logic "
            "dự phòng nội bộ (fallback trả về rỗng, không detect chức năng thiếu, "
            "không có coverage report).\n"
            f"  - Thử 'services.coverage_checker': {_exc_cc_services_pkg}\n"
            f"  - Thử 'coverage_checker' (cùng cấp): {_exc_cc_flat}"
        )
        _cc_detect_missing_modules = None
        _cc_detect_missing_targeted_modules = None
        _cc_build_coverage_report = None
        _cc_build_form_structure_coverage = None
        _cc_is_search_module = None
        _cc_canonical_module_key = None
        _cc_search_generic_names = None
        _cc_search_button_names = None
try:
    from services.scenario_rule_engine import (
        replace_generated_cases_with_template as _sre_replace_generated_cases_with_template,
        normalize_function_name as _sre_normalize_function_name,
        build_testcases_for_module as _sre_build_testcases_for_module,
        DEFAULT_ENFORCED_CANONICALS as _SRE_DEFAULT_ENFORCED_CANONICALS,
    )
    print("[ScenarioRuleEngine] Import từ 'services.scenario_rule_engine' thành công.")
except Exception as _exc_sre_services_pkg:
    try:
        from scenario_rule_engine import (
            replace_generated_cases_with_template as _sre_replace_generated_cases_with_template,
            normalize_function_name as _sre_normalize_function_name,
            build_testcases_for_module as _sre_build_testcases_for_module,
            DEFAULT_ENFORCED_CANONICALS as _SRE_DEFAULT_ENFORCED_CANONICALS,
        )
        print("[ScenarioRuleEngine] Import từ 'scenario_rule_engine' (cùng cấp) thành công.")
    except Exception as _exc_sre_flat:
        print(
            "[ScenarioRuleEngine] KHÔNG import được scenario_rule_engine — fixed "
            "template tập trung (Quay lại, Đăng nhập, Xem danh sách/chi tiết, "
            "Xuất file, In, Lưu, Hủy, Đóng popup) sẽ KHÔNG được enforce ở bước "
            "cuối; pipeline cũ (nếu còn) vẫn chạy như trước.\n"
            f"  - Thử 'services.scenario_rule_engine': {_exc_sre_services_pkg}\n"
            f"  - Thử 'scenario_rule_engine' (cùng cấp): {_exc_sre_flat}"
        )
        _sre_replace_generated_cases_with_template = None
        _sre_normalize_function_name = None
        _sre_build_testcases_for_module = None
        _SRE_DEFAULT_ENFORCED_CANONICALS = frozenset()

if _cc_is_search_module is None:
    _FALLBACK_SEARCH_PATTERN = re.compile(
        r'^(tìm|nút tìm|button tìm|tìm kiếm|ô tìm kiếm|search)'
        r'(\s*(theo|by)\s+.+)?$',
        re.IGNORECASE,
    )
    _cc_search_generic_names = {'tìm kiếm', 'ô tìm kiếm', 'search'}
    _cc_search_button_names = {'tìm', 'nút tìm', 'button tìm'}

    def _cc_is_search_module(name: str) -> bool:
        return bool(_FALLBACK_SEARCH_PATTERN.match(name.strip().lower()))

    def _cc_canonical_module_key(name: str) -> str:
        n = name.strip().lower()
        return 'search' if _cc_is_search_module(n) else n

try:
    from services.rule_engine import (
        select_domain_rules as _re_select_domain_rules,
        build_crud_rule_prompt as _re_build_crud_rule_prompt,
        detect_crud_action as _re_detect_crud_action,
        detect_crud_context as _re_detect_crud_context,
    )
    print("[RuleEngine] Import từ 'services.rule_engine' thành công.")
except Exception as _exc_re_services_pkg:
    try:
        from rule_engine import (
            select_domain_rules as _re_select_domain_rules,
            build_crud_rule_prompt as _re_build_crud_rule_prompt,
            detect_crud_action as _re_detect_crud_action,
            detect_crud_context as _re_detect_crud_context,
        )
        print("[RuleEngine] Import từ 'rule_engine' (cùng cấp) thành công.")
    except Exception as _exc_re_flat:
        print(
            "[RuleEngine] KHÔNG import được rule_engine — Domain/CRUD Rule Engine sẽ bị "
            "TẮT hoàn toàn (fallback trả về rỗng, không ảnh hưởng pipeline).\n"
            f"  - Thử 'services.rule_engine': {_exc_re_services_pkg}\n"
            f"  - Thử 'rule_engine' (cùng cấp): {_exc_re_flat}"
        )
        _re_select_domain_rules = None
        _re_build_crud_rule_prompt = None
        _re_detect_crud_action = None
        _re_detect_crud_context = None
load_dotenv()
def load_web2519_examples(excel_path: str | None = None, max_examples: int = 18) -> str:
    """
    Đọc file Excel WEB2519 và trích xuất các cặp (Tình huống kiểm định, Kết quả mong đợi)
    làm few-shot examples cho system prompt.

    Args:
        excel_path: Đường dẫn tới file Excel WEB2519. Nếu None, tự tìm trong thư mục hiện tại.
        max_examples: Số ví dụ tối đa trích xuất (trải đều qua các sheet). Giảm từ 40 → 18
            để few-shot block không phình to (ưu tiên ví dụ ngắn gọn, đủ dạy văn phong,
            thay vì lấy nguyên cả đoạn dài nhiều dòng).

    Returns:
        Chuỗi few-shot examples để nhúng vào system prompt, hoặc chuỗi rỗng nếu không đọc được.
    """
    if not _OPENPYXL_AVAILABLE:
        return ""
    if excel_path is None:
        search_dirs = [Path("."), Path(__file__).parent]
        for d in search_dirs:
            matches = sorted(d.glob("WEB2519*.xlsx"))
            if matches:
                excel_path = str(matches[0])
                break

    if not excel_path or not Path(excel_path).exists():
        return ""

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    except Exception:
        return ""
    skip_sheets = {"testcase lan 1", "thongtinchung", "tonghop", "test", "dashboard"}
    target_sheets = [s for s in wb.sheetnames if s.lower() not in skip_sheets]
    per_sheet = max(1, max_examples // max(len(target_sheets), 1))

    all_examples = [] 

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        header_row = None
        col_chucnang = col_tinhhong = col_ketqua = None
        def get_col(r, idx, max_chars: int = 220):
            if idx is not None and idx < len(r) and r[idx] is not None:
                text = str(r[idx]).strip()
                if len(text) > max_chars:
                    text = text[:max_chars].rstrip() + "..."
                return text
            return ""

        for row in ws.iter_rows(values_only=True, max_col=12):
            cells = [str(c).strip().lower() if c else "" for c in row]
            if header_row is None:
                if any("tình huống" in c for c in cells) and any("kết quả mong" in c for c in cells):
                    header_row = cells
                    col_chucnang = next((i for i, c in enumerate(cells) if "chức năng" in c), 2)
                    col_tinhhong = next((i for i, c in enumerate(cells) if "tình huống" in c), 3)
                    col_ketqua = next((i for i, c in enumerate(cells) if "kết quả mong" in c), 4)
                continue

            if header_row is None:
                continue

            raw = list(row)
            tinh_huong = get_col(raw, col_tinhhong)
            ket_qua = get_col(raw, col_ketqua)

            if not tinh_huong or not ket_qua:
                continue
            if len(tinh_huong) < 5 or len(ket_qua) < 5:
                continue

            chuc_nang = get_col(raw, col_chucnang)
            all_examples.append((chuc_nang, tinh_huong, ket_qua))
    seen_pairs: set = set()
    unique_examples = []
    for ex in all_examples:
        key = (ex[1][:60], ex[2][:60])
        if key not in seen_pairs:
            seen_pairs.add(key)
            unique_examples.append(ex)
    ACTION_GROUPS = {
        "đăng nhập thành công": [],
        "đăng nhập không thành công": [],
        "đăng ký thành công": [],
        "đăng ký không thành công": [],
        "đăng xuất": [],
        "quên mật khẩu": [],
        "đổi mật khẩu": [],
        "upload file": [],
        "tải xuống": [],
        "import file": [],
        "xuất file": [],
        "khóa kích hoạt": [],
        "tìm kiếm": [],
        "phân trang": [],
        "thêm mới thành công": [],
        "thêm mới không thành công": [],
        "sửa thành công": [],
        "sửa không thành công": [],
        "xóa thành công": [],
        "xóa không thành công": [],
        "xem danh sách": [],
        "xem chi tiết": [],
        "khác": [],
    }
    for ex in unique_examples:
        th_lower = ex[1].lower()
        cf_lower = ex[0].lower() if ex[0] else ""
        combined = cf_lower + " " + th_lower
        if "đăng nhập" in combined and ("không thành công" in combined or "không hợp lệ" in th_lower or "sai" in th_lower):
            ACTION_GROUPS["đăng nhập không thành công"].append(ex)
        elif "đăng nhập" in combined:
            ACTION_GROUPS["đăng nhập thành công"].append(ex)
        elif "đăng xuất" in combined:
            ACTION_GROUPS["đăng xuất"].append(ex)
        elif "quên mật khẩu" in combined or "quên mk" in combined:
            ACTION_GROUPS["quên mật khẩu"].append(ex)
        elif "đổi mật khẩu" in combined or "thay đổi mật khẩu" in combined:
            ACTION_GROUPS["đổi mật khẩu"].append(ex)
        elif "tải lên" in combined or "upload" in combined or "tải ảnh" in combined:
            ACTION_GROUPS["upload file"].append(ex)
        elif "tải xuống" in combined or "download" in combined or "tải tài liệu" in combined or "tải biểu mẫu" in combined or "tải file" in combined:
            ACTION_GROUPS["tải xuống"].append(ex)
        elif "import" in combined or "nhập dữ liệu" in combined or "nhập file" in combined:
            ACTION_GROUPS["import file"].append(ex)
        elif "xuất excel" in combined or "xuất word" in combined or "xuất file" in combined or "export" in combined:
            ACTION_GROUPS["xuất file"].append(ex)
        elif "khóa tài khoản" in combined or "kích hoạt" in combined or "vô hiệu hóa" in combined:
            ACTION_GROUPS["khóa kích hoạt"].append(ex)
        elif "đăng ký" in combined and ("không thành công" in combined or "không hợp lệ" in th_lower or "bắt buộc" in th_lower):
            ACTION_GROUPS["đăng ký không thành công"].append(ex)
        elif "đăng ký" in combined:
            ACTION_GROUPS["đăng ký thành công"].append(ex)
        elif "tìm kiếm" in combined:
            ACTION_GROUPS["tìm kiếm"].append(ex)
        elif "phân trang" in combined or "chuyển sang trang" in th_lower or "chuyển về trang" in th_lower:
            ACTION_GROUPS["phân trang"].append(ex)
        elif "thêm" in combined and (
            "không thành công" in combined or "không hợp lệ" in th_lower
            or "bắt buộc" in th_lower or "sai" in th_lower or "trống" in th_lower
        ):
            ACTION_GROUPS["thêm mới không thành công"].append(ex)
        elif "thêm" in combined:
            ACTION_GROUPS["thêm mới thành công"].append(ex)
        elif ("sửa" in combined or "cập nhật" in combined) and ("không thành công" in combined or "không hợp lệ" in th_lower):
            ACTION_GROUPS["sửa không thành công"].append(ex)
        elif "sửa" in combined or "cập nhật" in combined:
            ACTION_GROUPS["sửa thành công"].append(ex)
        elif "xóa" in combined and ("không thành công" in combined or "đã có thông tin" in th_lower):
            ACTION_GROUPS["xóa không thành công"].append(ex)
        elif "xóa" in combined:
            ACTION_GROUPS["xóa thành công"].append(ex)
        elif "xem danh sách" in combined or "quan sát danh sách" in th_lower:
            ACTION_GROUPS["xem danh sách"].append(ex)
        elif "xem chi tiết" in combined or "quan sát thông tin" in th_lower:
            ACTION_GROUPS["xem chi tiết"].append(ex)
        else:
            ACTION_GROUPS["khác"].append(ex)
    PRIORITY_GROUPS = {"đăng nhập thành công", "đăng nhập không thành công",
                        "đăng ký thành công", "đăng ký không thành công", "đăng xuất"}
    per_group = max(1, max_examples // len(ACTION_GROUPS))
    examples = []
    for group_name, group_list in ACTION_GROUPS.items():
        group_list_sorted = sorted(group_list, key=lambda ex: len(ex[1]) + len(ex[2]))
        take = max(per_group, 2) if group_name in PRIORITY_GROUPS else per_group
        examples.extend(group_list_sorted[:take])
    if len(examples) < max_examples:
        for ex in unique_examples:
            if ex not in examples:
                examples.append(ex)
            if len(examples) >= max_examples:
                break
    examples = examples[:max_examples]

    if not examples:
        return ""
    lines = ["=== VÍ DỤ THỰC TẾ TỪ DỰ ÁN WEB2519 (học cách viết Tình huống & Kết quả) ==="]
    lines.append("Áp dụng ĐÚNG phong cách viết như các ví dụ dưới đây:\n")
    for i, (chuc_nang, tinh_huong, ket_qua) in enumerate(examples[:max_examples], 1):
        block = f"Ví dụ {i}:"
        if chuc_nang:
            block += f"\n  Chức năng: {chuc_nang}"
        block += f"\n  Tình huống kiểm định: {tinh_huong}"
        block += f"\n  Kết quả mong đợi: {ket_qua}"
        lines.append(block)

    return "\n".join(lines)
_WEB2519_EXAMPLES: str = load_web2519_examples(max_examples=18)
def build_system_prompt(base_prompt: str, excel_path: str | None = None) -> str:
    """
    Ghép system prompt cơ sở với ví dụ WEB2519.
    Nếu đã cache examples thì dùng cache, không thì load từ excel_path.
    """
    global _WEB2519_EXAMPLES
    if excel_path:
        examples = load_web2519_examples(excel_path)
    else:
        examples = _WEB2519_EXAMPLES
    if examples:
        return base_prompt + "\n\n" + examples
    return base_prompt
_SCHEMA = '{"project_name":"...","description":"...","modules":{"Tên chức năng":[{"id":"TC_001","chức năng":"...","feature":"...","scenario":"...","title":"...","description":"...","given":"...","when":"...","then":"...","precondition":"...","steps":"1. ...\n2. ...\n3. ...","test_data":"...","expected_result":"...","priority":"Cao","test_type":"Kiểm thử chức năng","actual_result":"","status":"Chưa chạy","note":""}]}}'
_BASE_RULES = """- JSON object duy nhất. Tiếng Việt. modules là object {}.
- id tăng liên tục TC_001→TC_002, KHÔNG reset. status="Chưa chạy". actual_result="". note="".
- priority: Cao|Trung bình|Thấp. test_type chỉ được là một trong: Kiểm thử chức năng|Kiểm thử giao diện|Kiểm thử xác thực|Kiểm thử bảo mật|Kiểm thử phân quyền|Kiểm thử âm|Kiểm thử dương|Kiểm thử biên|Kiểm thử tích hợp
Nút "Tìm" (button riêng biệt, KHÔNG phải ô input tìm kiếm) PHẢI đặt tên chức năng CHÍNH XÁC là "Tìm" — TUYỆT ĐỐI KHÔNG ghép thêm mô tả/placeholder của ô input "Tìm kiếm theo [X]" đứng gần đó vào tên (vd KHÔNG đặt "Tìm theo mã", "Tìm theo mã hoặc kho", "Tìm theo [X]" cho chức năng của NÚT Tìm — dù ô input tìm kiếm gần đó tên gì, chức năng của nút Tìm luôn CHỈ là "Tìm").
Các nút hành động UI chung của form/popup — "Hủy"/"Huỷ"/"Hủy bỏ"/"Cancel", "Đóng popup", "Sinh mã", "Lưu", "Quay lại" — KHI người dùng yêu cầu/đặt tên CHÍNH XÁC đúng như vậy (không kèm thêm từ nào khác), PHẢI giữ NGUYÊN tên module ĐÚNG CHÍNH XÁC "Hủy"/"Đóng popup"/"Sinh mã"/"Lưu"/"Quay lại" — TUYỆT ĐỐI KHÔNG tự suy diễn hay ghép thêm tên đối tượng nghiệp vụ của dự án/màn hình hiện tại vào (vd yêu cầu chỉ nói "hủy" trong dự án "Quản lý lớp học" → module PHẢI là "Hủy", TUYỆT ĐỐI KHÔNG đặt "Hủy lớp học"; tương tự KHÔNG tự đặt "Hủy đơn hàng", "Hủy lịch hẹn", "Hủy tài khoản"... trừ khi người dùng đã tự gõ rõ ràng đầy đủ cụm đó chính là tên chức năng họ muốn). Nút Hủy/Cancel của 1 form/popup là hành động ĐÓNG FORM và KHÔNG LƯU thay đổi — KHÔNG phải hành động hủy/xóa đối tượng nghiệp vụ (lớp học/đơn hàng/lịch hẹn...), nên KHÔNG được sinh scenario dạng "hủy lớp học có sinh viên", "xóa lớp học", "dữ liệu tham chiếu của lớp học" cho module "Hủy" — chỉ mô tả hành vi đóng form/popup và có/không giữ dữ liệu."""
_SCENARIO_EXPECTED = """=== MẪU scenario => expected_result (chuẩn WEB2519) ===
Scenario: mô tả đầy đủ thao tác + dữ liệu/trường cụ thể liên quan, KHÔNG viết "Kiểm tra X hiển thị đúng".
Expected: nêu rõ thông báo UI + trạng thái data sau thao tác, KHÔNG viết "X hiển thị đúng".
- Tìm kiếm cơ bản: "Tìm kiếm theo bộ lọc: nhập từ khóa / theo [tên bộ lọc cụ thể từ ảnh]" => "Danh sách hiển thị đúng với thông tin tìm kiếm đã nhập\n  Trường hợp không có dữ liệu xuất hiện thông báo phù hợp"
- Tìm kiếm nâng cao: "Tìm kiếm theo bộ lọc nâng cao: [liệt kê tên bộ lọc cụ thể từ ảnh]\n  Kết hợp các điều kiện" => (expected giống Tìm kiếm cơ bản)
- Thêm mới thành công: "Nhập dữ liệu hợp lệ, lưu bằng nút Thêm mới\n  * Thông tin gồm: [liệt kê tên trường từ ảnh, đánh dấu (*) bắt buộc]" => "Xuất hiện thông báo \"Thêm thành công\". Thông tin được lưu với giá trị hợp lệ và được hiển thị tại danh sách\n  - Trường hợp nhập dữ liệu không phù hợp, không lưu thông tin và có thông báo lỗi phù hợp"
  NẾU màn hình có nhiều LOẠI danh mục con khác nhau (vd dropdown "Loại danh mục" với các lựa chọn: bằng cấp, chứng chỉ, chuyên ngành, danh mục dùng chung...) và MỖI loại có bộ trường bắt buộc RIÊNG, PHẢI liệt kê tách riêng từng loại trong scenario (KHÔNG gộp chung 1 dòng trường):
    "- Đối với danh mục [loại 1]: [tên trường (*)], [tên trường (*)]...\n  - Đối với danh mục [loại 2]: [tên trường (*)]...\n  ..." (liệt kê ĐỦ tất cả loại con thấy được trong ảnh/dữ liệu đầu vào, mỗi loại 1 dòng riêng)
- Thêm mới:"- Nhập liệu hợp lệ, lưu rồi thêm tiếp bằng nút \"Thêm mới và tiếp tục\"" => "Xuất hiện thông báo \"Thêm thành công\". Các thông tin được lưu với giá trị hợp lệ và được hiển thị tại danh sách\n  - Trường hợp nhập dữ liệu không phù hợp, không lưu thông tin và có thông báo lỗi phù hợp"
- Thêm không thành công: "- Không nhập trường bắt buộc\n  - Giá trị nhập/chọn không phù hợp" => "Xuất hiện thông báo lỗi phù hợp, không lưu thông tin"
- Sửa thành công: "Tại danh sách, chọn [đối tượng] để sửa\n  - Giá trị bắt buộc được nhập, phù hợp" => "Thông tin được lưu hợp lệ và hiển thị đúng tại danh sách"
- Sửa không thành công: "- Không nhập trường bắt buộc\n  - Giá trị nhập/chọn không phù hợp" => "Xuất hiện thông báo lỗi phù hợp, không lưu thông tin"
- Cập nhật thành công (icon bút vàng): "Nhấn icon Cập nhật tại dòng [đối tượng], sửa dữ liệu hợp lệ rồi lưu" => "Thông tin được lưu hợp lệ và hiển thị tại danh sách"
- Cập nhật không thành công (icon bút vàng): "Nhấn icon Cập nhật tại dòng [đối tượng]\n  - Không nhập trường bắt buộc / giá trị không phù hợp" => "Xuất hiện thông báo lỗi phù hợp, không lưu thông tin"
- Xóa thành công: "Chọn dòng [đối tượng] cần xóa, thực hiện xóa" => "Hệ thống thông báo xóa thành công, không còn hiển thị trong danh sách sau khi xác nhận"
- Xóa không thành công: "Chọn dòng [đối tượng] đang có dữ liệu tham chiếu, thực hiện xóa" => "Xóa không thành công, xuất hiện thông báo phù hợp"
- Xem chi tiết: "Quan sát thông tin [đối tượng] gồm: [Trường 1]/[Trường 2]/[Trường 3]" => "Hiển thị đúng định dạng, khớp dữ liệu đã cập nhật\n  - Người dùng đủ quyền có thể thêm/sửa/xóa"
- Đăng nhập thành công: "Nhập Email/Tên đăng nhập và Mật khẩu hợp lệ, nhấn Đăng nhập" => "Xuất hiện thông báo \"Đăng nhập thành công\", chuyển đến trang chính theo quyền tài khoản"
- Đăng nhập không thành công (sai MK / không tồn tại / bỏ trống): "Nhập sai Mật khẩu, hoặc tài khoản chưa đăng ký, hoặc để trống 1 trong 2 trường, nhấn Đăng nhập" => "Xuất hiện thông báo lỗi phù hợp (vd \"Email hoặc mật khẩu không đúng\"), vẫn ở lại trang đăng nhập"
- Đăng ký thành công: "Nhập đầy đủ thông tin hợp lệ vào trường bắt buộc\n  * Thông tin gồm: [liệt kê tên trường có (*)]\n  rồi nhấn Đăng ký" => "Xuất hiện thông báo \"Đăng ký thành công\", có thể đăng nhập bằng tài khoản vừa tạo"
- Đăng ký không thành công (bỏ trống / sai định dạng / trùng tài khoản): "Bỏ trống trường bắt buộc, hoặc Email/SĐT sai định dạng, hoặc tài khoản đã tồn tại, nhấn Đăng ký" => "Xuất hiện thông báo lỗi phù hợp từng trường hợp, không lưu tài khoản"
- Đăng xuất: "Nhấn nút/biểu tượng Đăng xuất tại menu tài khoản" => "Kết thúc phiên, chuyển về trang đăng nhập, không truy cập lại trang cần đăng nhập"
- Quên mật khẩu thành công: "Nhập Email/SĐT đã đăng ký vào ô xác thực, nhấn Gửi yêu cầu" => "Xuất hiện thông báo \"Yêu cầu đã được gửi\", liên kết/OTP gửi đến Email/SĐT"
- Quên mật khẩu không thành công: "Nhập Email/SĐT chưa đăng ký, hoặc bỏ trống ô xác thực, nhấn Gửi yêu cầu" => "Xuất hiện thông báo lỗi phù hợp, không gửi yêu cầu"
- Đổi mật khẩu thành công: "Nhập đúng MK hiện tại\n  * MK mới và Xác nhận MK mới khớp, đạt định dạng\n  rồi nhấn Lưu" => "Xuất hiện thông báo \"Đổi mật khẩu thành công\", lần sau phải dùng MK mới"
- Đổi mật khẩu không thành công: "Sai MK hiện tại, hoặc MK mới/Xác nhận không khớp, hoặc sai định dạng, hoặc bỏ trống trường bắt buộc" => "Xuất hiện thông báo lỗi phù hợp, MK giữ nguyên"
- Upload file/ảnh thành công: "Chọn file đúng định dạng và dung lượng cho phép, nhấn Tải lên" => "Xuất hiện thông báo \"Tải lên thành công\", file lưu và hiển thị đúng vị trí"
- Upload file/ảnh không thành công: "Chọn file sai định dạng / vượt dung lượng / không chọn file mà vẫn nhấn Tải lên" => "Xuất hiện thông báo lỗi phù hợp, không lưu file"
- Xuất file (Excel/PDF/Word): "Tại danh sách đang hiển thị (đã lọc nếu có), nhấn nút Xuất tương ứng\n  Nhiều biểu mẫu: chọn biểu mẫu xuất [liệt kê từ ảnh]" => "File tải về đúng định dạng, giá trị khớp với danh sách đang hiển thị"
- Khóa tài khoản: "Tại danh sách, chọn dòng tài khoản cần khóa, nhấn Khóa và xác nhận" => "Xuất hiện thông báo \"Khóa thành công\", trạng thái đổi Đã khóa, không đăng nhập được"
- Kích hoạt lại tài khoản: "Chọn dòng tài khoản đã khóa, nhấn Kích hoạt và xác nhận" => "Xuất hiện thông báo \"Kích hoạt thành công\", trạng thái đổi Hoạt động, đăng nhập lại được"
- Phân trang (BẮT BUỘC tách thành 4 TC riêng trong CÙNG 1 chức năng "Phân trang", KHÔNG gộp 4 hành động vào 1 TC):
  TC1: "Chuyển sang trang kế tiếp" => "Hiển thị dữ liệu trang tiếp theo"
  TC2: "Chuyển về trang trước" => "Hiển thị dữ liệu trang trước đó"
  TC3: "Chuyển sang trang cuối" => "Hiển thị dữ liệu trang cuối"
  TC4: "Chuyển sang trang đầu" => "Hiển thị dữ liệu trang đầu"
- Quay lại khi chưa thay đổi dữ liệu: "Nhấn nút/biểu tượng Quay lại khi chưa nhập/sửa gì" => "Hệ thống quay về màn hình danh sách trước đó, không hiển thị cảnh báo"
- Quay lại khi đã thay đổi dữ liệu: "Đang nhập/sửa dữ liệu, nhấn nút/biểu tượng Quay lại" => "Hiển thị cảnh báo dữ liệu chưa lưu; xác nhận bỏ thay đổi thì quay lại màn hình trước, hủy cảnh báo thì giữ nguyên màn hình và dữ liệu hiện tại"
- Quay lại từ màn hình: "Từ màn hình chi tiết/sửa, nhấn Quay lại" => "Chuyển đúng về màn hình danh sách vừa click trước đó"
- Quay lại khi không có màn hình trước: "Truy cập trực tiếp (vd link/refresh), nhấn Quay lại" => "Hệ thống quay về trang mặc định hoặc không hành động\""""

_CHECKLIST_UI = """=== CHECKLIST THEO LOẠI UI ===
Thêm mới (KHÔNG ép cứng số lượng TC, KHÔNG sinh TC "hủy thao tác"; tách testcase theo NHÓM NGHIỆP VỤ liên quan dựa trên field thực tế của form, mỗi TC CHỈ mô tả 1 nhóm lỗi liên quan, TUYỆT ĐỐI KHÔNG nhồi toàn bộ Required/Length/Whitespace/Validation/Boundary/Duplicate/XSS/SQL Injection vào 1 TC/1 ô Mô tả): (1) thành công — nhập đầy đủ dữ liệu hợp lệ→lưu OK | (2) NẾU form có trường định danh được phép bỏ trống để hệ thống tự sinh (vd "Mã ...") → thêm 1 TC riêng: để trống, hệ thống tự sinh giá trị hợp lệ, không trùng, lưu thành công | (3) NẾU có trường bắt buộc → 1 TC nhóm Required/Whitespace: bỏ trống hoặc chỉ nhập khoảng trắng→lỗi bắt buộc, không lưu | (4) NẾU có trường liên quan → 1 TC nhóm Duplicate/Format/Length: trùng dữ liệu, sai định dạng hoặc vượt độ dài→lỗi phù hợp, không lưu | (5) 1 TC nhóm Security: nhập XSS hoặc SQL Injection→hệ thống không thực thi mã, không lưu dữ liệu nguy hiểm
Cập nhật (ưu tiên 7–9 TC theo NHÓM NGHIỆP VỤ, không tách mỗi field thành 1 TC): mở đúng bản ghi | cập nhật thành công | không thay đổi dữ liệu | Required/Whitespace | Format | Duplicate/Length/Boundary | field phụ thuộc | Security | Permission/Conflict/System error. Chỉ sinh nhóm có bằng chứng từ UI; được gộp biến thể cùng bản chất.
Xóa (ưu tiên 5–6 TC): mở xác nhận đúng đối tượng | xác nhận xóa thành công | hủy/đóng popup | dữ liệu tham chiếu không được xóa | không quyền/bản ghi không tồn tại | lỗi hệ thống/chống double-submit.
Tìm kiếm (ô lọc, ĐÚNG 2 TC, KHÔNG tách lẻ): (1) thành công — hợp lệ→khớp | (2) không thành công — GỘP không khớp/trống/khoảng trắng/biên max/XSS-SQLi vào 1 TC DUY NHẤT→rỗng hoặc thông báo phù hợp
Tìm (nút, ĐÚNG 2 TC, KHÔNG tách lẻ): (1) thành công — hợp lệ→kết quả | (2) không thành công — GỘP trống/ký tự đặc biệt/double-click vào 1 TC DUY NHẤT→toàn bộ hoặc thông báo
Dropdown: mở→đủ option | chọn→cập nhật đúng | mặc định→đúng | rỗng option→thông báo
Checkbox/Radio: chọn/bỏ chọn→đúng dấu | chọn tất→toàn bộ | bỏ 1 sau chọn tất→cập nhật tổng | radio khác→cũ tự bỏ
Datepicker: hợp lệ→đúng | không hợp lệ→lỗi | trống bắt buộc→lỗi | bắt đầu>kết thúc→lỗi
Tab/Menu: click→đúng nội dung+highlight
Badge/Trạng thái: mỗi trạng thái→đúng màu/nhãn | đổi dữ liệu→badge tự đổi
Quay lại (ĐÚNG 4 TC): chưa đổi dữ liệu→về màn hình trước, không cảnh báo | đã đổi dữ liệu→hiện cảnh báo dữ liệu chưa lưu, xác nhận thì quay lại/hủy thì giữ nguyên màn hình và dữ liệu | từ màn hình chi tiết/sửa→đúng về màn hình vừa click | không có màn hình trước→trang mặc định/không hành động
Phân trang: Next/Prev/First/Last→đúng trang | số trang→đúng | Prev disabled trang 1 | Next disabled trang cuối
Số dòng/trang: chọn N→tối đa N | đổi số→reset trang 1
Sort cột: tăng/giảm dần→đúng | giá trị null→đúng vị trí
Export: click→đúng định dạng+đủ data | rỗng→xử lý hợp lý
Import: hợp lệ→nhập OK | sai định dạng/lỗi→báo lỗi, không nhập
Modal: mở→đúng nội dung | xác nhận→đúng hành động | Hủy/X/click ngoài→đóng, data không đổi"""

_ICON_ACTIONS = """=== ICON CỘT THAO TÁC (mỗi icon = 1 chức năng riêng) ===
KHÔNG tạo chức năng "Thao tác" chung. CHỈ tạo chức năng cho icon THỰC SỰ xuất hiện trong ảnh/dữ liệu đầu vào — TUYỆT ĐỐI KHÔNG tự suy đoán hay thêm icon "thường thấy ở các bảng tương tự" nếu không có bằng chứng rõ ràng trong input.
Bút vàng → "Cập nhật": icon ngoài danh sách phải mở đúng form và dữ liệu bản ghi; nếu ảnh/workflow có form Cập nhật chi tiết thì chức năng Cập nhật ưu tiên 7–9 TC theo nhóm nghiệp vụ (success/no-change/required/format/duplicate-boundary/dependency/security/permission-system), không tách riêng từng field.
Thùng đỏ → "Xóa": ưu tiên 5–6 TC gọn theo nhóm: mở xác nhận đúng bản ghi | xác nhận xóa thành công | hủy/đóng | ràng buộc tham chiếu | không quyền/không tồn tại | lỗi hệ thống/double-submit.
Mắt xanh lá → "Xem chi tiết": click→mở đủ thông tin | khớp danh sách | đóng→về danh sách | không quyền→lỗi
Excel xanh lá → "Xuất file Excel": click→tải .xlsx đúng+đủ data dòng | không quyền→lỗi
Word xanh dương → "Xuất file Word": click→tải .docx đúng+đủ data dòng | không quyền→lỗi"""

_FIELD_RULE_TABLE: list[dict] = [
    {
        "key": "email",
        "type_name": "Email",
        "patterns": ["email", "e-mail"],
        "hint": (
            '  * Validation định dạng: nhập sai định dạng (thiếu "@", thiếu domain, '
            'vd "abc.com", "abc@") => "Xuất hiện thông báo lỗi định dạng email, không lưu thông tin"\n'
            '  * Trùng lặp (nếu Email dùng làm định danh tài khoản/đăng nhập): nhập Email đã tồn tại '
            '=> "Xuất hiện thông báo Email đã được sử dụng, không lưu thông tin"'
        ),
    },
    {
        "key": "phone",
        "type_name": "Số điện thoại",
        "patterns": ["số điện thoại", "sđt", "sdt", "phone"],
        "hint": (
            '  * Validation định dạng: nhập chữ cái/ký tự đặc biệt, hoặc không đúng 10 số, hoặc không '
            'bắt đầu bằng số 0 => "Xuất hiện thông báo lỗi định dạng số điện thoại, không lưu thông tin"\n'
            '  * Boundary độ dài: nhập ít hơn hoặc nhiều hơn 10 số => "Xuất hiện thông báo lỗi, không lưu thông tin"'
        ),
    },
    
    {
        "key": "score",
        "type_name": "Điểm",
        "patterns": ["điểm"],
        "hint": (
            '  * Validation định dạng: nhập chữ cái hoặc ký tự đặc biệt '
            '=> "Hệ thống không chấp nhận dữ liệu không đúng định dạng, hiển thị thông báo lỗi phù hợp và không thực hiện lưu/cập nhật dữ liệu."\n'
            '  * Boundary: nhập giá trị nhỏ hơn giá trị tối thiểu hoặc lớn hơn giá trị tối đa theo quy định của hệ thống '
            '=> "Hệ thống từ chối dữ liệu, hiển thị thông báo điểm ngoài phạm vi cho phép và không thực hiện lưu/cập nhật dữ liệu."\n'
            '  * Boundary: nhập giá trị tại giới hạn nhỏ nhất và lớn nhất được phép '
            '=> "Hệ thống chấp nhận dữ liệu hợp lệ và thực hiện lưu/cập nhật thành công."'
        ),
    },
    {
        "key": "password",
        "type_name": "Mật khẩu",
        "patterns": ["mật khẩu", "password"],
        "hint": (
            '  * Boundary độ dài tối thiểu: nhập Mật khẩu ngắn hơn độ dài quy định (vd < 6 ký tự) '
            '=> "Xuất hiện thông báo lỗi độ dài mật khẩu, không lưu thông tin"\n'
            '  * Validation khớp xác nhận (nếu có trường Xác nhận mật khẩu): nhập 2 giá trị không khớp '
            '=> "Xuất hiện thông báo mật khẩu xác nhận không khớp, không lưu thông tin"\n'
            '  * Bảo mật: ký tự Mật khẩu PHẢI hiển thị dạng ẩn (●/*) trên giao diện '
            '=> "Giá trị mật khẩu không hiển thị dạng chữ thường trên UI"'
        ),
    },
    {
        "key": "numeric",
        "type_name": "Số/Tiền tệ",
        "patterns": ["số lượng", "đơn giá", "thành tiền", "tồn kho", "số tiền", "giá", "khối lượng"],
        "hint": (
            '  * Validation định dạng: nhập chữ cái/ký tự đặc biệt vào trường số '
            '=> "Xuất hiện thông báo lỗi định dạng, không lưu thông tin"\n'
            '  * Boundary: nhập giá trị 0, số âm, số thập phân (nếu không cho phép), hoặc vượt giá trị '
            'tối đa cho phép => "Xuất hiện thông báo lỗi phù hợp theo từng trường hợp biên, không lưu thông tin"'
        ),
    },
    {
        "key": "cccd",
        "type_name": "CMND/CCCD",
        "patterns": ["cmnd", "cccd", "căn cước"],
        "hint": (
            '  * Validation định dạng: nhập không đúng số ký tự quy định (9 hoặc 12 số) hoặc chứa chữ cái '
            '=> "Xuất hiện thông báo lỗi định dạng, không lưu thông tin"\n'
            '  * Trùng lặp: nhập số CMND/CCCD đã tồn tại trong hệ thống '
            '=> "Xuất hiện thông báo đã tồn tại, không lưu thông tin"'
        ),
    },
    {
        "key": "code",
        "type_name": "Mã (định danh)",
        "patterns": ["mã"],
        "hint": (
            '  * Trùng lặp: nhập giá trị Mã đã tồn tại trong hệ thống '
            '=> "Xuất hiện thông báo Mã đã tồn tại, không lưu thông tin"\n'
            '  * Validation ký tự: nhập khoảng trắng đầu/cuối hoặc ký tự đặc biệt không cho phép '
            '=> "Xuất hiện thông báo lỗi định dạng Mã, không lưu thông tin"'
        ),
    },
]

SYSTEM_PROMPT_TARGETED = f"""QA Engineer. Sinh testcase CHỈ cho chức năng được yêu cầu. Chỉ JSON, không markdown.

=== SCHEMA ===
{_SCHEMA}

=== QUY TẮC ===
{_BASE_RULES}
- Số module = đúng số chức năng được yêu cầu. Ảnh chỉ dùng để lấy tên trường/test_data.
KHÔNG tạo module từ ảnh. Chức năng Search: SỐ module = SỐ UI element THẬT SỰ tồn tại (ô input search cụ thể VÀ/HOẶC nút "Tìm" riêng biệt) — TUYỆT ĐỐI KHÔNG tự thêm module "Tìm kiếm" tên chung chung nếu nó KHÔNG tương ứng với 1 UI element khác biệt thật sự (vd chỉ có 1 ô "Tìm kiếm theo mã hoặc tên chu kỳ" + 1 nút "Tìm" → CHỈ 2 module, KHÔNG được sinh thêm "Tìm kiếm" là module thứ 3).
- Số lượng TC mỗi module: MẶC ĐỊNH tối thiểu 4 TC. NGOẠI LỆ — "Tìm kiếm"/"Tìm kiếm theo [X]", "Tìm" (nút) LUÔN ĐÚNG 2 TC; riêng "Cập nhật" áp dụng rule CRUD 7–9 TC khi có form chi tiết (thành công / không thành công, gộp mọi biến thể lỗi vào 1 TC DUY NHẤT, KHÔNG tạo thêm TC thứ 3). Riêng "Thêm mới": KHÔNG ép số lượng cố định — tách theo NHÓM NGHIỆP VỤ hợp lý dựa trên field thực tế của form (thành công / tự sinh mã nếu có / thiếu trường bắt buộc / trùng-sai định dạng-vượt độ dài / XSS-SQL Injection), mỗi TC chỉ mô tả 1 nhóm lỗi liên quan.

{_SCENARIO_EXPECTED}

{_CHECKLIST_UI}"""

SYSTEM_PROMPT_FULL = f"""QA Engineer. Sinh testcase cho TẤT CẢ chức năng trong ảnh. Chỉ JSON, không markdown.

=== SCHEMA ===
{_SCHEMA}

=== QUY TẮC ===
{_BASE_RULES}
- Với FORM/MODAL/POPUP: TRƯỚC KHI tạo module, PHẢI xác định cấu trúc Form (tên Form, Fields, Buttons, Business Rules, Workflow — xem block "FORM STRUCTURE ĐÃ PHÂN TÍCH" nếu có trong prompt). FIELD (input/textarea/dropdown/datepicker/checkbox/radio) TUYỆT ĐỐI KHÔNG được tạo module riêng theo tên field — toàn bộ test case Validation/Boundary/Business Rule/XSS/SQL Injection/Required/Length/Duplicate/Whitespace của từng field PHẢI nằm bên trong module hành động LƯU tương ứng (vd "Thêm mới"/"Cập nhật"), KHÔNG tách field thành module con riêng. CHỈ button, icon-button, nút X đóng popup và nút footer (mỗi nút có hành vi riêng) mới là 1 module riêng.\n- Số lượng TC mỗi module: MẶC ĐỊNH tối thiểu 4 TC. NGOẠI LỆ — các module sau LUÔN ĐÚNG 2 TC (thành công / không thành công, gộp mọi biến thể lỗi vào 1 TC "không thành công" DUY NHẤT, KHÔNG tạo thêm TC thứ 3): "Tìm kiếm"/"Tìm kiếm theo [X]", "Tìm" (nút). Riêng "Cập nhật" áp dụng rule CRUD gọn 7–9 TC khi có form chi tiết. Riêng "Thêm mới": KHÔNG ép số lượng TC cố định — TÁCH TESTCASE THEO NHÓM NGHIỆP VỤ hợp lý dựa trên field thực tế của form (mỗi TC chỉ mô tả 1 nhóm lỗi liên quan, KHÔNG nhồi Required/Length/Whitespace/Validation/Boundary/Duplicate/XSS/SQL Injection vào 1 TC): (1) thành công, (2) tự sinh mã — nếu có trường định danh cho phép bỏ trống, (3) thiếu trường bắt buộc/khoảng trắng, (4) trùng/sai định dạng/vượt độ dài, (5) XSS/SQL Injection. Ngoài 2 ngoại lệ trên (nhóm ≥4 TC mặc định và nhóm ĐÚNG 2 TC cố định) và ngoại lệ "Thêm mới" co giãn theo nhóm nghiệp vụ, không còn trường hợp nào khác — mọi module khác (Xóa, Xem chi tiết, Xuất Excel/Word...) áp dụng đúng số TC theo checklist/scenario mẫu bên dưới. Mỗi button/icon-button = 1 module riêng.
 KHÔNG tạo TC cho cột hiển thị tĩnh (STT, Mã, Tên, Ngày tạo, Người tạo...) hay nhãn thông tin ("trên tổng số X dòng").
 KHÔNG tạo module "Thao tác" chung — mỗi icon trong cột Thao tác phải là module RIÊNG (Cập nhật / Xóa / Xem chi tiết...).
 Chức năng Search: SỐ module = SỐ UI element THẬT SỰ có trên ảnh (ô input tìm kiếm cụ thể theo field VÀ/HOẶC nút "Tìm" riêng biệt) — mỗi UI element khác nhau là 1 module riêng, nhưng TUYỆT ĐỐI KHÔNG tự sinh thêm module "Tìm kiếm" tên chung chung nếu nó không tương ứng với 1 UI element khác biệt thật sự trên ảnh (vd ảnh chỉ có ô "Tìm kiếm theo mã hoặc tên chu kỳ" + nút "Tìm" → CHỈ 2 module, KHÔNG thêm module "Tìm kiếm" thứ 3 chỉ vì tên gần giống).
 KHÔNG tạo module cho breadcrumb (vd "Trang chủ > Quản lý danh mục > ..."). Breadcrumb chỉ là điều hướng tĩnh, KHÔNG phải chức năng cần test case riêng — bỏ qua hoàn toàn, không đưa vào "modules".
 Toàn bộ phân trang (nút Next/Prev/First/Last, số trang 1/2/3..., "Hiển thị X trên tổng số Y") gộp chung vào DUY NHẤT 1 module tên "Phân trang" — KHÔNG tạo thêm module phân trang thứ 2 với tên khác (vd "phân trang: 1 2"). Trong module "Phân trang" này PHẢI có đúng 4 TC riêng biệt (Next/Prev/Last/First — xem mẫu scenario/expected bên dưới), KHÔNG được gộp 4 hành động vào 1 TC duy nhất.
- Tên module: chỉ lấy phần TEXT của nút, KHÔNG kèm ký tự icon (+, ←, →, ✓...). VD: "+ Thêm mới" → tên module "Thêm mới"; "← Quay lại" → tên module "Quay lại".

{_SCENARIO_EXPECTED}

{_CHECKLIST_UI}

{_ICON_ACTIONS}"""

SYSTEM_PROMPT_TEXT_ONLY = f"""QA Engineer. User mô tả hệ thống bằng TEXT THUẦN (KHÔNG có ảnh giao diện đính kèm). Sinh testcase cho TẤT CẢ chức năng được liệt kê trong mô tả. Chỉ JSON, không markdown.

=== SCHEMA ===
{_SCHEMA}

=== QUY TẮC ===
{_BASE_RULES}
- Mỗi chức năng nghiệp vụ user liệt kê (vd "quản lý nhân viên", "chấm công", "tính lương", "báo cáo") PHẢI có ÍT NHẤT 1 module tương ứng trong "modules" — KHÔNG được bỏ sót bất kỳ chức năng nào user đã nêu tên.
- Với chức năng dạng "Quản lý [đối tượng]" (CRUD), PHẢI tách thành các module con riêng biệt: Thêm mới / Cập nhật (hoặc Sửa) / Xóa / Tìm kiếm / Xem chi tiết-danh sách — áp dụng cho các thao tác hợp lý với đối tượng đó (KHÔNG bắt buộc đủ cả 5 nếu bản chất nghiệp vụ không có, vd không phải đối tượng nào cũng cho phép Xóa).
- Với chức năng nghiệp vụ đặc thù KHÔNG thuộc CRUD chuẩn (vd "Chấm công", "Tính lương", "Báo cáo"), tự suy luận thao tác/kịch bản hợp lý theo nghiệp vụ thực tế Việt Nam (vd Chấm công: check-in/check-out đúng giờ, đi trễ, quên chấm công; Tính lương: tính đúng công thức theo ngày công, sai lệch dữ liệu chấm công; Báo cáo: xem theo khoảng thời gian, xuất file, không có dữ liệu) — mỗi module tối thiểu 3 TC, gồm tối thiểu 1 kịch bản thành công và 1 kịch bản lỗi/ngoại lệ.
- KHÔNG suy diễn thêm chức năng NGOÀI những gì user đã liệt kê (vd user không nhắc "phân quyền" thì không tự thêm module phân quyền riêng, trừ khi role/quyền được nêu rõ trong mô tả).
- Tên module: đặt tên NGẮN GỌN theo đúng cách user gọi chức năng đó (vd user viết "quản lý nhân viên" → tách "Thêm nhân viên"/"Sửa nhân viên"/"Xóa nhân viên"/"Tìm kiếm nhân viên", KHÔNG đặt tên chung chung "Quản lý nhân viên" bao trùm tất cả thao tác trong 1 module).

{_SCENARIO_EXPECTED}

{_CHECKLIST_UI}"""
_FULL_JSON_CHAR_LIMIT = 24_000


def _log_generation_runtime(func):
    """Log tổng thời gian và kết quả generate_test_cases mà không đổi logic pipeline."""
    def wrapper(
        self, description, previous_test_cases=None, image_blocks=None,
        domain=None, context_mode="new",
    ):
        started_at = time.perf_counter()
        project_label = (description or "Dự án").strip()
        print("\n" + "=" * 64)
        print("🚀 AI TESTCASE GENERATOR - START")
        print(f"📁 Yêu cầu : {project_label[:120]}")
        print(f"🖼️  Số ảnh  : {len(image_blocks or [])}")
        print("=" * 64)
        try:
            result = func(
                self,
                description,
                previous_test_cases=previous_test_cases,
                image_blocks=image_blocks,
                domain=domain,
                context_mode=context_mode,
            )
            elapsed = time.perf_counter() - started_at
            modules = result.get("modules", {}) if isinstance(result, dict) else {}
            # Đếm SỐ CHỨC NĂNG GỐC cho mục đích thống kê/log — "Đăng nhập
            # thành công" và "Đăng nhập không thành công" phải tính là 1
            # chức năng, không phải 2. Chỉ dùng base_name để ĐẾM, KHÔNG ghi
            # đè lên result/tc (result giữ nguyên module key có hậu tố).
            if isinstance(modules, dict) and modules:
                _base_names_seen: set[str] = set()
                for _mod_name in modules.keys():
                    try:
                        _base, _ = self._determine_base_business_function(_mod_name)
                    except Exception:
                        _base = _mod_name
                    _base_names_seen.add((_base or _mod_name or '').strip())
                total_modules = len(_base_names_seen)
            else:
                total_modules = 0
            total_testcases = (
                sum(len(items) for items in modules.values() if isinstance(items, list))
                if isinstance(modules, dict) else 0
            )
            project_name = (
                result.get("project_name")
                if isinstance(result, dict) else None
            ) or project_label
            print("\n" + "=" * 64)
            print("🎉 GENERATE TEST CASE SUCCESS")
            print(f"📦 Project         : {project_name}")
            print(f"🧩 Tổng chức năng  : {total_modules}")
            print(f"📦 Total TestCases : {total_testcases}")
            print(f"⏱️  Total Time      : {elapsed:.2f} s")
            print("=" * 64 + "\n")
            return result
        except Exception as exc:
            elapsed = time.perf_counter() - started_at
            print("\n" + "=" * 64)
            print("❌ GENERATE TEST CASE FAILED")
            print(f"📍 Stage        : generate_test_cases")
            print(f"⚠️  Reason       : {exc}")
            print(f"⏱️  Elapsed Time : {elapsed:.2f} s")
            print("=" * 64 + "\n")
            raise
    return wrapper


class AIService:
    def __init__(self, web2519_excel_path: str | None = None):
        api_key = os.environ.get('OPENAI_API_KEY')
        if not api_key:
            raise ValueError("OPENAI_API_KEY chưa được cấu hình trong file .env")
        self.client = OpenAI(api_key=api_key)
        self.model = os.environ.get('OPENAI_MODEL', 'gpt-4o')
        self.max_retries = 3
        self.vision_service = VisionService()
        self.workflow_service = WorkflowService()
        self._current_screen_context: dict | None = None
        self._current_workflow_relation: dict | None = None
        self._current_workflow_context: dict | None = None
        self._scan_cache: dict[str, str] = {}
        self._last_coverage_rounds_used: int | None = None
        self._current_form_structure: dict | None = None
        self._current_crud_context: dict | None = None
        self.rag_service = RAGService() if RAGService else None
        _excel = (
            web2519_excel_path
            or os.environ.get('WEB2519_EXCEL_PATH')
        )
        if _excel:
            examples = load_web2519_examples(_excel)
            if examples:
                self._web2519_examples = examples
                return
        self._web2519_examples = _WEB2519_EXAMPLES

    def _detect_request_crud_context(
        self,
        scanned: str | None,
        description: str | None = None,
        previous_test_cases: dict | None = None,
    ) -> dict:
        """
        Phân biệt CRUD ở MÀN HÌNH DANH SÁCH với FORM/POPUP chi tiết.

        Quy tắc cứng:
        - Có table/danh sách + icon/nút CRUD, nhưng chỉ có ô tìm kiếm/lọc:
          screen_type=list.
        - Chỉ coi là form khi có title modal/form hoặc có >=2 field không phải
          tìm kiếm/lọc và có nút submit tương ứng.
        - Popup xóa cần câu xác nhận hoặc tiêu đề xác nhận xóa.
        """
        raw = str(scanned or "")
        lower = raw.lower()
        combined = f"{description or ''}\n{raw}"

        action = ""
        if _re_detect_crud_action is not None:
            try:
                action = _re_detect_crud_action(combined) or ""
            except Exception:
                action = ""

        if not action:
            if re.search(r"(?<!\w)(cập nhật|chỉnh sửa|edit|update)(?!\w)", combined, re.I):
                action = "update"
            elif re.search(r"(?<!\w)(thêm mới|tạo mới|add new|create)(?!\w)", combined, re.I):
                action = "create"
            elif re.search(r"(?<!\w)(xóa|xoá|delete|remove)(?!\w)", combined, re.I):
                action = "delete"

        if not action:
            return {
                "action": "",
                "screen_type": "unknown",
                "has_form_fields": False,
                "required_fields": [],
                "parent_module": None,
                "parent_exists": False,
                "confidence": 0.0,
            }
        base_context = {}
        if _re_detect_crud_context is not None:
            try:
                previous_modules = (
                    previous_test_cases.get("modules", {})
                    if isinstance(previous_test_cases, dict) else None
                )
                base_context = _re_detect_crud_context(
                    search_text=combined,
                    action=action,
                    previous_modules=previous_modules,
                ) or {}
            except Exception:
                base_context = {}

        has_table = bool(re.search(
            r"(?im)\|\s*(table|data-table|datatable|pagination)\b|"
            r"\b(stt|cột thao tác|danh sách|phân trang|số dòng/trang|"
            r"trên tổng số|dòng dữ liệu|bản ghi|trang đầu|trang cuối)\b",
            raw,
        ))
        detected_crud_actions = {
            key for key, pattern in {
                "create": r"(?<!\w)(thêm mới|tạo mới)(?!\w)",
                "update": r"(?<!\w)(cập nhật|chỉnh sửa|edit|update)(?!\w)",
                "delete": r"(?<!\w)(xóa|xoá|delete|remove)(?!\w)",
            }.items()
            if re.search(pattern, combined, re.I)
        }
        has_row_actions = bool(re.search(
            r"(?im)(icon-button|action-icon|cột thao tác|tại dòng|mỗi dòng|"
            r"bút vàng|thùng đỏ|icon cập nhật|icon xóa|icon xoá)",
            raw,
        ))
        has_search_or_filter = bool(re.search(
            r"(?im)\|\s*(search|input|dropdown|select)\b.*"
            r"(tìm kiếm|search|lọc|filter|từ khóa|từ khoá)|"
            r"(tìm kiếm|search|lọc|filter|từ khóa|từ khoá)",
            raw,
        ))
        strong_list_evidence = (
            has_table
            or has_row_actions
            or (len(detected_crud_actions) >= 2 and has_search_or_filter)
        )
        has_modal_title = bool(re.search(
            r"(?im)\|\s*(modal-title|popup-title|dialog-title|form-title)\b|"
            r"\b(modal|popup|dialog|form)\b.{0,50}\b"
            r"(thêm mới|tạo mới|cập nhật|chỉnh sửa|xác nhận xóa|xác nhận xoá)\b",
            raw,
        ))
        delete_confirmation = action == "delete" and any(
            phrase in lower for phrase in (
                "bạn có chắc", "có chắc chắn", "xác nhận xóa",
                "xác nhận xoá", "không thể hoàn tác",
            )
        )

        field_lines = []
        for line in raw.splitlines():
            if not re.search(
                r"\|\s*(input|textarea|dropdown|select|combobox|datepicker|"
                r"date-picker|radio|checkbox|file-upload|upload)\b",
                line,
                re.I,
            ):
                continue
            line_lower = line.lower()
            if any(k in line_lower for k in (
                "tìm kiếm", "search", "lọc", "filter", "từ khóa", "từ khoá",
            )):
                continue
            field_lines.append(line)

        submit_patterns = {
            "create": r"(?im)^\s*[-+•]?\s*(thêm mới|tạo mới|lưu|thêm mới và tiếp tục)\s*\|\s*(button|submit)",
            "update": r"(?im)^\s*[-+•]?\s*(cập nhật|lưu thay đổi|lưu)\s*\|\s*(button|submit)",
            "delete": r"(?im)^\s*[-+•]?\s*(xác nhận xóa|xác nhận xoá|xóa|xoá|đồng ý)\s*\|\s*(button|submit)",
        }
        has_submit = bool(re.search(submit_patterns.get(action, r"$^"), raw, re.I | re.M))
        if delete_confirmation:
            screen_type = "confirm_popup"
            confidence = 1.0
        elif (
            has_modal_title
            and bool(field_lines)
            and has_submit
            and not strong_list_evidence
        ):
            screen_type = "form"
            confidence = 0.99
        elif len(field_lines) >= 2 and has_submit and not strong_list_evidence:
            screen_type = "form"
            confidence = 0.95
        else:
            screen_type = "list"
            confidence = 0.99 if strong_list_evidence else 0.86

        display = {"create": "Thêm mới", "update": "Cập nhật", "delete": "Xóa"}[action]
        required_fields = base_context.get("required_fields") or []
        return {
            **base_context,
            "action": action,
            "screen_type": screen_type,
            "has_form_fields": bool(field_lines),
            "required_fields": required_fields,
            "parent_module": display,
            "detected_actions": sorted(detected_crud_actions),
            "strong_list_evidence": strong_list_evidence,
            "confidence": confidence,
        }

    def _crud_is_list_only(self, action: str | None = None) -> bool:
        """True khi request hiện tại là màn hình danh sách.

        Một ảnh danh sách có thể đồng thời chứa Thêm mới, Cập nhật và Xóa.
        Vì vậy KHÔNG được yêu cầu action phải trùng action chính mà detector
        chọn. Nếu screen_type=list thì mọi chức năng CRUD trong ảnh đều chỉ
        kiểm tra hành động mở form/popup.
        """
        context = self._current_crud_context or {}
        return context.get("screen_type") == "list"

    def _crud_required_count(self, action: str, form_required: int) -> int:
        return 2 if self._crud_is_list_only(action) else form_required

    def _build_system_prompt(self, base_prompt: str) -> str:
        """Ghép base prompt với ví dụ WEB2519 của instance này."""
        if self._web2519_examples:
            return base_prompt + "\n\n" + self._web2519_examples
        return base_prompt

    def _extract_field_labels_for_rules(self, elements_text: str) -> tuple[list[str], bool]:
        """
        Trích danh sách nhãn field (input/textarea/datepicker) từ chuỗi elements
        đã quét (định dạng "- [Tên label] | [loại] | [chi tiết]").
        Trả về (danh sách nhãn, có ít nhất 1 input/textarea tự do hay không).
        Bỏ qua button/icon-button/table/tab/badge/pagination vì rule engine chỉ
        áp dụng cho field NHẬP LIỆU.
        """
        labels: list[str] = []
        has_free_text_input = False
        if not elements_text:
            return labels, has_free_text_input

        for line in elements_text.splitlines():
            line = line.strip()
            if not line.startswith('-'):
                continue
            parts = line.lstrip('- ').split('|')
            label = parts[0].strip() if parts else ''
            if not label:
                continue
            kind = parts[1].strip().lower() if len(parts) > 1 else ''
            if any(k in kind for k in ('input', 'textarea')):
                has_free_text_input = True
            if any(k in kind for k in ('input', 'textarea', 'datepicker')):
                labels.append(label)

        return labels, has_free_text_input

    def _detect_date_range_hint(self, labels: list[str]) -> str:
        """Phát hiện cặp field 'Ngày bắt đầu'/'Ngày kết thúc' (hoặc 'Từ ngày'/'Đến ngày')
        và sinh hint validate quan hệ giữa 2 field — quy tắc nghiệp vụ phổ biến
        mà template hành động chung (_SCENARIO_EXPECTED) không bao phủ."""
        start_kw = ("ngày bắt đầu", "từ ngày")
        end_kw = ("ngày kết thúc", "đến ngày")
        start_field = next((l for l in labels if any(k in l.lower() for k in start_kw)), None)
        end_field = next((l for l in labels if any(k in l.lower() for k in end_kw)), None)
        if not (start_field and end_field):
            return ""
        return (
            f'- Cặp trường "{start_field}" / "{end_field}" (quan hệ ngày bắt đầu–kết thúc):\n'
            f'  * Validation quan hệ: chọn "{end_field}" trước "{start_field}" '
            f'=> "Xuất hiện thông báo lỗi ngày kết thúc phải sau ngày bắt đầu, không lưu thông tin"\n'
            f'  * Boundary: chọn "{start_field}" và "{end_field}" trùng cùng 1 ngày '
            f'=> "Cho phép lưu hợp lệ (khoảng thời gian 1 ngày)"'
        )

    def _build_generic_security_hint(self, has_free_text_input: bool) -> str:
        """Rule bảo mật chung (injection) áp dụng khi màn hình có ít nhất 1 field
        nhập liệu tự do (input/textarea)."""
        if not has_free_text_input:
            return ""
        return (
            "- Bảo mật chung cho các trường nhập liệu dạng text:\n"
            '  * Nhập chuỗi script/SQL injection (vd "<script>alert(1)</script>", "\' OR \'1\'=\'1") '
            'vào trường bất kỳ => "Hệ thống không thực thi mã, dữ liệu được xử lý an toàn '
            '(escape/sanitize), không phát sinh lỗi hệ thống"'
        )

    def _build_rule_engine_hints(self, elements_text: str) -> str:
        """
        Rule Engine (bảng rule CỐ ĐỊNH theo loại field): quét elements_text,
        khớp tên/nhãn field với _FIELD_RULE_TABLE (email/sđt/mật khẩu/số-tiền tệ/
        CMND-CCCD/mã), cộng thêm quan hệ ngày bắt đầu–kết thúc và 1 rule bảo mật
        chung, rồi trả về block hint để nhúng vào prompt — buộc AI sinh thêm TC
        validation/boundary/bảo mật cho ĐÚNG field thực sự xuất hiện trong ảnh
        (KHÔNG suy diễn field không tồn tại).

        Trả về "" nếu không có field nào khớp rule (không làm phình prompt vô ích).
        """
        labels, has_free_text_input = self._extract_field_labels_for_rules(elements_text)
        if not labels:
            return ""

        hint_blocks: list[str] = []
        matched_keys: set[str] = set()

        for label in labels:
            label_lower = label.lower()
            words = re.split(r'[\s/,()\-]+', label_lower)
            for rule in _FIELD_RULE_TABLE:
                if rule["key"] in matched_keys:
                    continue
                for pat in rule["patterns"]:
                    is_match = (pat in label_lower) if ' ' in pat else (pat in words)
                    if is_match:
                        hint_blocks.append(f'- Trường "{label}" (kiểu {rule["type_name"]}):\n{rule["hint"]}')
                        matched_keys.add(rule["key"])
                        break
        date_hint = self._detect_date_range_hint(labels)
        if date_hint:
            hint_blocks.append(date_hint)

        security_hint = self._build_generic_security_hint(has_free_text_input)
        if security_hint:
            hint_blocks.append(security_hint)

        if not hint_blocks:
            return ""

        return (
            "\n=== RULE ENGINE: TC VALIDATION/BOUNDARY/BẢO MẬT BẮT BUỘC BỔ SUNG THEO FIELD ===\n"
            "(Áp dụng CHỈ cho field thực sự có trong danh sách UI ELEMENTS bên trên, "
            "KHÔNG suy diễn field không tồn tại. Nếu chức năng của field này đã có 1 TC "
            '"không thành công" gộp chung các lỗi input, có thể GỘP THÊM các trường hợp '
            "dưới đây vào CÙNG TC đó thay vì tạo TC mới, miễn giữ đúng số TC tối đa đã quy định ở trên.)\n"
            + "\n".join(hint_blocks)
        )

    def _select_relevant_rules(self, description: str, scanned: str, targeted: bool) -> tuple[str, str]:
        """
        BƯỚC 2 — chọn elements liên quan + rule engine áp dụng cho đúng phạm vi:
        - targeted=True: lọc `scanned` theo mô tả (_extract_targeted_elements) —
          chỉ giữ elements của các chức năng user yêu cầu.
        - targeted=False: dùng nguyên toàn bộ `scanned` (full scan).
        Rule Engine hints luôn được tính TRÊN PHẦN ĐÃ CHỌN (không tính trên toàn
        màn hình khi targeted), để không sinh hint cho field ngoài phạm vi yêu cầu.

        Trả về (relevant_elements, rule_hints).
        """
        if targeted:
            relevant_elements = self._extract_targeted_elements(description, scanned or '')
        else:
            relevant_elements = scanned or ''
        rule_hints = self._build_rule_engine_hints(relevant_elements)
        return relevant_elements, rule_hints

    def _hash_image_blocks(self, image_blocks: list[dict] | None) -> str | None:
        """
        Sinh hash ổn định đại diện cho NỘI DUNG ảnh (base64/url trong
        image_blocks) để làm cache key cho BƯỚC 1 (quét ảnh) — dùng để nhận
        biết "ảnh giống ảnh trước" mà không cần so sánh pixel.
        """
        if not image_blocks:
            return None
        try:
            raw = json.dumps(image_blocks, sort_keys=True, ensure_ascii=False)
        except Exception:
            raw = str(image_blocks)
        return hashlib.sha256(raw.encode('utf-8', errors='ignore')).hexdigest()

    def _cache_covers_request(self, cached_scanned: str | None, description: str, targeted: bool) -> bool:
        """
        Kiểm tra kết quả scan ĐÃ CACHE (từ lần quét ảnh trước, cùng hash ảnh)
        có đủ đáp ứng yêu cầu HIỆN TẠI không, để quyết định có cần quét lại
        ảnh hay không:
        - targeted: TẤT CẢ chức năng được liệt kê trong `description` phải
          đã xuất hiện trong cached_scanned (tách cụm giống logic
          _coverage_checker: dấu phẩy/";"/"+"/"/"/"và"). Nếu cache THIẾU dù
          chỉ 1 chức năng → coi như KHÔNG đủ, phải quét lại.
        - full (không targeted): chỉ cần cache có tối thiểu 2 dòng element
          hợp lệ (không phải scan lỗi/rỗng do OCR fail).
        """
        if not cached_scanned:
            return False
        if targeted:
            raw_terms = re.split(r'[,;+/]|(?:\bvà\b)', description or '')
            terms = [t.strip() for t in raw_terms if t.strip()]
            if not terms:
                return True
            low = cached_scanned.lower()
            return all(t.lower() in low for t in terms)
        element_lines = [l for l in cached_scanned.splitlines() if l.strip().startswith('-')]
        return len(element_lines) >= 2

    def _coverage_checker(
        self,
        description: str,
        targeted: bool,
        scanned: str,
        image_blocks: list[dict] | None,
    ) -> tuple[str, str, str]:
        """
        BƯỚC 3 — kiểm tra ĐỘ ĐẦY ĐỦ CỦA KẾT QUẢ QUÉT ẢNH (khác với
        _enforce_min_coverage vốn kiểm tra SỐ LƯỢNG TC sau khi AI đã sinh xong).
        Mục tiêu: bắt sớm case OCR bỏ sót element TRƯỚC KHI tốn 1 lượt gọi API
        sinh TC — đây chính là hướng điều tra cho bug "Thêm mới" chỉ ra ~3 TC.

        - targeted: tách từng chức năng được yêu cầu trong `description`
          (theo dấu phẩy / "và" / "+" / ";" / "/"), kiểm tra CÓ xuất hiện trong
          `scanned` gốc (chưa lọc) hay không. Nếu thiếu → re-scan 1 lần bằng
          pipeline legacy (1 lượt Vision+phân loại — khác pipeline 2-lượt
          chính) để thử bắt lại element đã bị OCR bỏ sót, rồi gộp thêm dòng mới
          vào `scanned`.
        - full scan: nếu `scanned` có quá ít dòng element (< 2, nghi ngờ OCR
          đọc thiếu/lỗi) → re-scan 1 lần bằng legacy, thay thế nếu kết quả mới
          nhiều element hơn.

        Trả về (scanned đã cập nhật, relevant_elements, rule_hints) — đã tính
        lại BƯỚC 2 trên `scanned` sau khi coverage-check.
        """
        updated_scanned = scanned or ''

        if targeted:
            raw_terms = re.split(r'[,;+/]|(?:\bvà\b)', description or '')
            terms = [t.strip() for t in raw_terms if t.strip()]
            missing_terms = [t for t in terms if t.lower() not in updated_scanned.lower()]

            if missing_terms and image_blocks:
                print(
                    "=== COVERAGE CHECKER: nghi ngờ OCR bỏ sót ===\n"
                    f"Chức năng được yêu cầu nhưng KHÔNG thấy trong scan: {missing_terms}\n"
                    "→ Re-scan bằng pipeline legacy (1 lượt Vision+phân loại)...\n"
                )
                legacy_scanned = self._scan_image_elements_legacy(image_blocks, description)
                if legacy_scanned:
                    existing_lines = {l.strip() for l in updated_scanned.splitlines()}
                    new_lines = [
                        l for l in legacy_scanned.splitlines()
                        if l.strip().startswith('-') and l.strip() not in existing_lines
                    ]
                    if new_lines:
                        updated_scanned = updated_scanned.rstrip() + "\n" + "\n".join(new_lines)
                        print(
                            "=== COVERAGE CHECKER: bổ sung element mới từ legacy scan ===\n"
                            + "\n".join(new_lines) + "\n"
                        )
                    still_missing = [t for t in missing_terms if t.lower() not in updated_scanned.lower()]
                    if still_missing:
                        print(f"=== COVERAGE CHECKER: vẫn KHÔNG thấy sau re-scan: {still_missing} ===")
        else:
            element_lines = [l for l in updated_scanned.splitlines() if l.strip().startswith('-')]
            if len(element_lines) < 2 and image_blocks:
                print(
                    "=== COVERAGE CHECKER: scan quá ít element, nghi ngờ OCR lỗi ===\n"
                    "→ Re-scan bằng pipeline legacy...\n"
                )
                legacy_scanned = self._scan_image_elements_legacy(image_blocks, description)
                legacy_lines = [
                    l for l in (legacy_scanned or '').splitlines() if l.strip().startswith('-')
                ]
                if len(legacy_lines) > len(element_lines):
                    updated_scanned = legacy_scanned

        relevant_elements, rule_hints = self._select_relevant_rules(description, updated_scanned, targeted)
        return updated_scanned, relevant_elements, rule_hints

    def _summarize_previous_test_cases(self, previous_test_cases: dict) -> str:
        project_name = previous_test_cases.get('project_name') or 'Project'
        modules = previous_test_cases.get('modules', {})
        module_summaries = []
        for module_name, test_cases in modules.items():
            if not isinstance(test_cases, list):
                continue
            features = [tc.get('feature', '') for tc in test_cases if isinstance(tc, dict)]
            features = [f for f in features if f]
            if features:
                module_summaries.append(f"- {module_name}: {', '.join(features[:6])}")
        if not module_summaries:
            module_summaries.append('- Không có chức năng nào được ghi nhận.')
        return (
            f"Tên dự án: {project_name}\n"
            f"Các module hiện tại:\n" + "\n".join(module_summaries[:10])
        )

    def _retrieve_rag_context(self, query: str, top_k: int = 5, targeted: bool = False) -> str:
        """
        Lấy ngữ cảnh RAG ngắn gọn để đưa vào prompt.
        Nếu chưa cấu hình RAG hoặc lỗi đọc dữ liệu thì trả về chuỗi rỗng để hệ thống vẫn chạy bình thường.
        targeted: truyền xuống RAGService.retrieve để chỉ lọc cứng theo
        đúng chức năng được yêu cầu khi đang ở chế độ TARGETED (TH1);
        mặc định False (FULL/TEXT) để không đổi hành vi cũ.
        """
        if not query or not self.rag_service:
            return ""
        try:
            return self.rag_service.retrieve(query=query, top_k=top_k, targeted=targeted)
        except Exception as exc:
            print(f"[RAG] Bỏ qua RAG vì lỗi: {exc}")
            return ""

    def _append_rag_context(self, prompt: str, rag_context: str) -> str:
        """Nhúng RAG vào cuối prompt, kèm ràng buộc để AI không sinh chức năng dư."""
        if not rag_context:
            return prompt
        return (
            f"{prompt}\n\n"
            "=== RAG CONTEXT - THAM KHẢO NGHIỆP VỤ ===\n"
            f"{rag_context}\n\n"
            "=== QUY TẮC DÙNG RAG ===\n"
            "- Chỉ dùng RAG để tham khảo business rules, workflow và testcase mẫu.\n"
            "- KHÔNG copy máy móc testcase trong RAG.\n"
            "- KHÔNG tạo chức năng ngoài yêu cầu người dùng hoặc ngoài UI elements đã chọn.\n"
            "- Nếu RAG mâu thuẫn với ảnh/mô tả người dùng, ưu tiên ảnh và mô tả người dùng.\n"
        )

    def _build_messages(
        self,
        description: str,
        previous_test_cases: dict | None = None,
        image_blocks: list[dict] | None = None,
        system_prompt: str | None = None,
        is_full_scan: bool = False,
        context_mode: str = "new",
    ) -> list[dict]:
        if system_prompt is None:
            system_prompt = self._build_system_prompt(SYSTEM_PROMPT_TARGETED)
        has_images = bool(image_blocks)

        if not previous_test_cases:
            if has_images:
                if is_full_scan:
                    # TH2: Full scan 
                    text = (
                        "Phân tích TẤT CẢ các ảnh giao diện được đính kèm và sinh testcase "
                        "cho TẤT CẢ chức năng/thành phần UI xuất hiện trong ảnh.\n\n"
                        "BƯỚC 1 – Quét ảnh từ trái sang phải, trên xuống dưới, "
                        "nhận diện TOÀN BỘ UI elements (button, input, icon, pagination, v.v.).\n"
                        "BƯỚC 2 – Tạo 1 chức năng riêng cho MỖI thành phần tương tác tìm được. "
                        "Với form/modal phải tách riêng từng input, Sinh mã, từng nút footer, Hủy bỏ và nút X đóng. "
                        "KHÔNG bỏ sót bất kỳ button, icon, field, ô tìm kiếm, phân trang nào.\n"
                        "BƯỚC 3 – Sinh đủ testcase cho TỪNG chức năng theo checklist trong system prompt.\n"
                        "BƯỚC 4 – Đảm bảo JSON đóng hoàn chỉnh, id TC tăng dần liên tục.\n\n"
                        "TUYỆT ĐỐI KHÔNG giới hạn số chức năng — "
                        "sinh đủ chức năng cho TẤT CẢ thành phần UI thấy trong ảnh.\n\n"
                        f"Thông tin bổ sung: {description}"
                    )
                else:
                    # TH1: Targeted 
                    text = (
                        "Phân tích TẤT CẢ các ảnh giao diện được đính kèm và sinh testcase.\n"
                        "BƯỚC 1 – Quét ảnh từ trái sang phải, trên xuống dưới để nhận diện UI elements "
                        "nhằm lấy thông tin phục vụ testcase.\n"
                        "BƯỚC 2 – Đọc mô tả bổ sung của người dùng.\n"
                        "BƯỚC 3 – Chỉ tạo chức năng dựa trên CÁC CHỨC NĂNG người dùng yêu cầu. "
                        "Không tạo chức năng cho từng UI element.\n"
                        "Các chức năng trong yêu cầu được xem là các chức năng riêng biệt, "
                        "không tự động gộp các chức năng khác nhau.\n"
                        "BƯỚC 4 – Các UI element chỉ dùng để bổ sung dữ liệu test, "
                        "không dùng để tạo thêm chức năng.\n"
                        "BƯỚC 5 – Sinh đủ testcase theo quy tắc trong system prompt.\n"
                        "PHẢI sinh testcase cho TẤT CẢ chức năng đã được yêu cầu. "
                        "JSON phải đóng hoàn chỉnh.\n\n"
                        f"Mô tả bổ sung: {description}"
                    )
            else:
                text = (
                    "Sinh testcase cho hệ thống sau, trả về đúng cấu trúc JSON đã mô tả.\n"
                    "PHẢI sinh ĐỦ testcase cho TẤT CẢ chức năng, JSON phải đóng hoàn chỉnh.\n\n"
                    f"Mô tả: {description}"
                )
        else:
            previous_json_str = json.dumps(previous_test_cases, ensure_ascii=False)
            if has_images:
                context_block = (
                    "Context bộ testcase cũ (tóm tắt):\n"
                    + self._summarize_previous_test_cases(previous_test_cases)
                )
            elif len(previous_json_str) <= _FULL_JSON_CHAR_LIMIT:
                context_block = f"Bộ testcase hiện tại (JSON đầy đủ):\n{previous_json_str}"
            else:
                summary = self._summarize_previous_test_cases(previous_test_cases)
                context_block = f"Tóm tắt bộ testcase hiện tại:\n{summary}"

            if has_images:
                existing_modules_note = (
                    "Danh sách chức năng đã có (KHÔNG tạo lại):\n"
                    + "\n".join(f"- {m}" for m in (previous_test_cases or {}).get("modules", {}).keys())
                )
                if self._is_targeted_request(description):
                    instruction = (
                        f"User CHỈ yêu cầu sinh testcase cho chức năng sau: \"{description}\".\n"
                        "TUYỆT ĐỐI KHÔNG phân tích hay sinh chức năng cho các thành phần UI khác "
                        "trong ảnh (vd KHÔNG tự thêm chức năng 'Hiển thị', 'Mô tả', cột bảng, hay bất "
                        "kỳ chức năng nào không được yêu cầu ở trên).\n"
                        "CHỈ tạo chức năng cho ĐÚNG chức năng được yêu cầu.\n"
                        "Đảm bảo id KHÔNG trùng với chức năng đã có trong context bên dưới.\n"
                        + existing_modules_note
                    )
                else:
                    instruction = (
                        "Phân tích TẤT CẢ ảnh giao diện được đính kèm và sinh testcase ĐỦ cho các "
                        "CHỈ tạo chức năng cho các chức năng được người dùng yêu cầu.Không sinh chức năng dư từ ảnh.\n"
                        "CHỈ trả về các chức năng MỚI từ ảnh này. "
                        "Đảm bảo id KHÔNG trùng với các chức năng đã có trong context bên dưới.\n"
                        + existing_modules_note
                    )
            elif len(previous_json_str) <= _FULL_JSON_CHAR_LIMIT:
                instruction = (
                    "Cập nhật và mở rộng bộ testcase hiện tại theo yêu cầu mới. "
                    "Giữ TOÀN BỘ testcase cũ, chỉ bổ sung testcase cho chức năng mới/đã thay đổi. "
                    "Trả về JSON đầy đủ với TẤT CẢ chức năng (cũ + mới)."
                )
            else:
                instruction = (
                    "Sinh testcase MỚI cho các chức năng được yêu cầu. "
                    "Đảm bảo id không trùng với chức năng đã liệt kê. "
                    "Chỉ trả về chức năng MỚI/CẦN CẬP NHẬT."
                )

            if context_mode == "screen_only":
                instruction = (
                    "CHẾ ĐỘ: CHỈ PHÂN TÍCH MÀN HÌNH HIỆN TẠI.\n"
                    "Testcase cũ bên dưới CHỈ dùng để hiểu tên dự án, đối tượng nghiệp vụ, "
                    "màn hình cha và chức năng liên quan.\n"
                    "KHÔNG sao chép testcase cũ. KHÔNG trả lại chức năng cũ. KHÔNG gộp kết quả. "
                    "KHÔNG nối thành workflow.\n"
                    "CHỈ sinh testcase cho ảnh/mô tả MỚI. Kết quả là một bộ testcase riêng "
                    "và id bắt đầu lại từ TC_001."
                )
            elif context_mode == "workflow":
                instruction = (
                    "CHẾ ĐỘ: TIẾP TỤC WORKFLOW.\n"
                    "Dùng testcase cũ để hiểu màn hình trước, nối màn hình mới vào cùng luồng "
                    "nghiệp vụ và chỉ sinh phần testcase mới phù hợp. Hệ thống sẽ gộp phần mới "
                    "với bộ testcase cũ sau khi AI trả kết quả."
                )
            text = (
                f"{instruction}\n\n"
                f"Yêu cầu mới: {description}\n\n"
                f"{context_block}"
            )
        if image_blocks:
            user_content = [
                {"type": "text", "text": text},
                *image_blocks,
            ]
        else:
            user_content = text

        return [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content},
        ]

    def _parse_json_response(self, content: str) -> dict:
        if not content:
            raise json.JSONDecodeError("Empty response", "", 0)
        content = content.strip()
        content = re.sub(r'^```(?:json)?\s*', '', content)
        content = re.sub(r'\s*```$', '', content)
        content = content.strip()
        try:
            return json.loads(content)
        except json.JSONDecodeError:
            pass
        if _JSON_REPAIR_AVAILABLE:
            try:
                return json.loads(_repair_json(content))
            except Exception:
                pass
        repaired = self._repair_truncated_json(content)
        return json.loads(repaired)

    def _repair_truncated_json(self, content: str) -> str:
        """Đóng các bracket/brace/string còn thiếu khi JSON bị cắt do max_tokens."""
        stack = []      
        in_string = False
        escape = False
        pos = len(content)  
        i = 0
        last_safe = 0  
        for i, ch in enumerate(content):
            if escape:
                escape = False
                continue
            if ch == '\\' and in_string:
                escape = True
                continue
            if ch == '"':
                in_string = not in_string
                if not in_string:
                    last_safe = i + 1  
                continu
            if in_string:
                continue
            if ch in '[{':
                stack.append(']' if ch == '[' else '}')
            elif ch in ']}':
                if stack and stack[-1] == ch:
                    stack.pop()
                    if not stack:
                        last_safe = i + 1
            elif ch not in ' \t\n\r,:':                pass
            else:
                if not stack:
                    last_safe = i + 1
        if in_string:
            content = content[:last_safe]
            content = re.sub(r'[,:]\s*$', '', content.rstrip())
        content = re.sub(r',\s*$', '', content.rstrip())
        stack = []
        in_string = False
        escape = False
        for ch in content:
            if escape:
                escape = False
                continue
            if ch == '\\' and in_string:
                escape = True
                continue
            if ch == '"':
                in_string = not in_string
                continue
            if in_string:
                continue
            if ch in '[{':
                stack.append(']' if ch == '[' else '}')
            elif ch in ']}':
                if stack and stack[-1] == ch:
                    stack.pop()
        closing = ''.join(reversed(stack))
        return content + closing

    def _unwrap_modules(self, raw: dict) -> dict:
        """
        Chuẩn hóa raw response về dạng {"modules": {"ModuleName": [tc, ...]}}

        Xử lý tất cả các biến thể AI hay trả về:
        1. Chuẩn:     {"modules": {"ModA": [...]}}                   ✓
        2. List:      {"modules": [{"name":"ModA","test_cases":[]}]}  → convert
        3. Flat TC:   {"modules": {"ModA": {"id":...}}}              → wrap list
        4. Nested:    {"modules": {"modules": {...}}}                 → unwrap
        5. Alt key:   {"test_cases": {...}}                           → rename
        6. No wrap:   {"ModA": [...], "ModB": [...]}                  → add wrapper
        """
        modules_raw = raw.get('modules')

        if modules_raw is None:
            for alt in ('testcases', 'test_cases', 'test_suites', 'suites'):
                if raw.get(alt) is not None:
                    modules_raw = raw[alt]
                    break

        if modules_raw is None:
            meta = {'project_name', 'description', 'name', 'title', 'version'}
            candidates = {k: v for k, v in raw.items() if k not in meta}
            if candidates and all(isinstance(v, list) for v in candidates.values()):
                return {
                    'project_name': raw.get('project_name', ''),
                    'description': raw.get('description', ''),
                    'modules': candidates,
                }

        if modules_raw is None:
            raise ValueError(
                f"Không tìm thấy trường modules. Keys nhận được: {list(raw.keys())}"
            )

        if isinstance(modules_raw, dict) and list(modules_raw.keys()) == ['modules']:
            modules_raw = modules_raw['modules']

        if isinstance(modules_raw, list):
            first = modules_raw[0] if modules_raw else {}
            is_list_of_tc = isinstance(first, dict) and any(
                k in first for k in ('scenario', 'given', 'when', 'then', 'expected_result')
            )

            if is_list_of_tc:
                module_name = raw.get('project_name') or 'Chức năng'
                modules_raw = {module_name: modules_raw}
            else:
                converted: dict[str, list] = {}
                for item in modules_raw:
                    if not isinstance(item, dict):
                        continue
                    name = (
                        item.get('name') or item.get('chức năng') or
                        item.get('module_name') or item.get('title') or 'Chức năng'
                    )
                    tcs = (
                        item.get('test_cases') or item.get('testcases') or
                        item.get('tests') or item.get('cases') or []
                    )
                    if not isinstance(tcs, list):
                        tcs = [tcs] if tcs else []
                    converted.setdefault(str(name), []).extend(tcs)
                modules_raw = converted

        if not isinstance(modules_raw, dict):
            raise ValueError(
                f"modules có kiểu không hợp lệ: {type(modules_raw).__name__}. "
                f"Preview: {str(modules_raw)[:300]}"
            )

        clean_modules: dict[str, list] = {}
        for mod_name, mod_val in modules_raw.items():
            if isinstance(mod_val, list):
                clean_modules[mod_name] = mod_val
            elif isinstance(mod_val, dict):
                if any(k in mod_val for k in ('id', 'scenario', 'given')):
                    clean_modules[mod_name] = [mod_val]
                elif all(isinstance(v, dict) for v in mod_val.values()):
                    clean_modules[mod_name] = list(mod_val.values())
                else:
                    continue

        if not clean_modules:
            raise ValueError(
                "modules dict tồn tại nhưng không có chức năng hợp lệ nào. "
                f"Preview: {str(modules_raw)[:300]}"
            )

        return {
            'project_name': raw.get('project_name') or raw.get('name') or 'Dự án',
            'description': raw.get('description', ''),
            'modules': clean_modules,
        }
    _MODULE_CANONICAL_GROUPS = [
        ['tìm kiếm theo mã', 'search by code'],
        ['tìm kiếm theo tên', 'search by name'],
        ['tìm kiếm theo kho'],
        ['tìm kiếm', 'ô tìm kiếm', 'search'],
        ['tìm', 'nút tìm', 'button tìm'],
        ['quay lại', 'back', 'nút quay lại'],
        ['thêm mới và tiếp tục', 'lưu và tiếp tục', 'thêm và tiếp tục', 'save and continue'],
        ['thêm mới', 'thêm', 'tạo mới', 'add'],
        ['xóa', 'xoá', 'delete'],
        ['cập nhật', 'nút cập nhật', 'icon cập nhật'],
        ['chỉnh sửa', 'edit', 'update'],
        ['đóng', 'x'],
        ['xem chi tiết', 'xem', 'view', 'chi tiết'],
        ['xem & kiểm kê', 'xem và kiểm kê', 'kiểm kê'],
        ['sửa kế hoạch', 'chỉnh sửa kế hoạch', 'edit kế hoạch'],
        ['xuất file excel', 'xuất excel', 'excel'],
        ['xuất file word', 'xuất word', 'word'],
        ['xóa đợt kiểm kê', 'xoá đợt kiểm kê', 'delete đợt kiểm kê'],
        ['xóa kế hoạch', 'xoá kế hoạch'],
        ['phân trang', 'pagination', 'trang'],
        ['số dòng mỗi trang', 'page-size-dropdown', 'hiển thị số dòng', 'page size'],
        ['xóa điều kiện tìm kiếm', 'clear search', 'xóa tìm kiếm', 'nút x tìm kiếm'],
    ]
    _SITUATIONAL_QUALIFIER_KEYWORDS = [
        'sql injection', 'sqli', 'xss', 'mã độc', 'script độc', 'injection', 'csrf', 'security',
        'validation', 'bỏ trống', 'để trống', 'không nhập', 'thiếu trường', 'khoảng trắng', 'bắt buộc',
        'sai định dạng', 'không đúng định dạng', 'trùng dữ liệu', 'trùng',
        'vượt quá', 'vượt độ dài', 'quá dài', 'quá ngắn',
        'boundary', 'ngoài khoảng', 'giới hạn', 'giá trị tối đa', 'giá trị tối thiểu', 'giá trị biên',
        'permission', 'không đủ quyền', 'không có quyền', 'phân quyền', 'không được phép',
        'business rule', 'không đủ số dư', 'đã tồn tại', 'không tồn tại',
        'trạng thái không hợp lệ', 'mã đã tồn tại',
        'system', 'timeout', 'hết thời gian', 'lỗi hệ thống', 'lỗi api', 'lỗi database',
        'mất mạng', 'mất kết nối', 'lỗi kết nối', 'lỗi mạng',
        'hủy thao tác', 'huỷ thao tác', 'đóng popup', 'hủy bỏ', 'huỷ bỏ', 'cancel',
        'exception', 'lỗi không xác định', 'ngoại lệ', 'unexpected error',
        'không thành công', 'thành công', 'thất bại',
    ]

    def _determine_base_business_function(self, module_name: str) -> tuple[str, str]:
        """
        Tách 1 tên chức năng bất kỳ do AI sinh ra thành:
          (chức_năng_gốc, qualifier_tình_huống)
        vd: "Đăng nhập SQL Injection" -> ("Đăng nhập", "SQL Injection")
            "Thêm mới"                -> ("Thêm mới", "")
            "Tìm kiếm theo mã"        -> ("Tìm kiếm theo mã", "") — không có
              từ khóa tình huống nào trong _SITUATIONAL_QUALIFIER_KEYWORDS
              khớp "theo mã" nên giữ nguyên, đúng vì đây là biến thể field
              cụ thể của chức năng Tìm kiếm, không phải tình huống lỗi.
        KHÔNG hard-code tên chức năng cụ thể nào — chỉ dựa vào từ khóa tình
        huống tổng quát nên áp dụng được cho mọi chức năng.
        """
        name = (module_name or '').strip()
        if not name:
            return name, ''
        lower = re.sub(r'\s+', ' ', name.lower())
        for suffix in (' không thành công', ' thành công'):
            if lower.endswith(suffix):
                return name[: -len(suffix)].strip(), ''

        earliest_pos = None
        for kw in self._SITUATIONAL_QUALIFIER_KEYWORDS:
            idx = lower.find(kw)
            if idx == -1:
                continue
            if earliest_pos is None or idx < earliest_pos:
                earliest_pos = idx
        if earliest_pos is None or earliest_pos == 0:
            return name, ''
        base = name[:earliest_pos].strip(' -–:,.')
        qualifier = name[earliest_pos:].strip(' -–:,.')
        if not base:
            return name, ''
        return base, qualifier

    _FAILURE_SIGNAL_KEYWORDS = [
        'không thành công', 'thất bại', 'không hợp lệ', 'từ chối', 'không lưu',
        'không thực hiện được', 'không tồn tại', 'không đủ quyền', 'không có quyền',
        'bắt buộc', 'bỏ trống', 'để trống', 'sai định dạng', 'trùng',
        'vượt quá', 'vượt độ dài',
        'xss', 'sql injection', 'sqli', 'injection', 'mã độc',
        'timeout', 'hết thời gian', 'mất mạng', 'mất kết nối',
        'lỗi hệ thống', 'lỗi api', 'lỗi database', 'lỗi kết nối', 'lỗi mạng',
        'hủy', 'huỷ', 'đóng popup', 'không mở được', 'không tìm thấy', 'hết hạn',
        'ngoại lệ', 'không đủ số dư', 'đã tồn tại', 'trạng thái không hợp lệ',
        'lỗi ',
    ]
    _SUCCESS_SIGNAL_KEYWORDS = [
        'thành công', 'hợp lệ', 'hiển thị đúng', 'mở đúng', 'lưu đúng', 'hoạt động đúng',
    ]

    def _classify_tc_outcome(self, tc: dict) -> str:
        """
        Phân loại 1 TC theo đúng quy tắc nghiệp vụ: 'success' CHỈ KHI TC
        hoàn thành mục tiêu nghiệp vụ (dữ liệu hợp lệ, lưu/cập nhật/xóa/
        đăng nhập/chuyển khoản/xuất file/in/mở màn hình/xem chi tiết/tìm
        kiếm... thành công). MẶC ĐỊNH là 'failure' cho MỌI trường hợp còn
        lại (Validation/Boundary/Permission/Security/Business Rule/System/
        User Action/Exception) — kể cả khi không khớp tín hiệu nào, để
        không bao giờ xếp nhầm 1 TC mơ hồ vào nhóm thành công.
        KHÔNG dùng test_type (Kiểm thử biên/phân quyền/xác thực/bảo mật...)
        làm tín hiệu outcome — đã thử và bị bác bỏ bằng dữ liệu thật: TC_002
        "Đăng nhập bằng tài khoản Admin" có test_type "Kiểm thử phân quyền"
        nhưng là ca THÀNH CÔNG; TC_008 "Chuyển toàn bộ số dư khả dụng hợp
        lệ" có test_type "Kiểm thử biên" nhưng cũng là ca THÀNH CÔNG. Loại
        test_type mô tả KỸ THUẬT kiểm thử, không mô tả KẾT QUẢ mong đợi,
        nên không thể suy outcome một chiều từ đó.
        Thay vào đó, quét từ khóa THẤT BẠI/THÀNH CÔNG trên TOÀN BỘ nội
        dung nghiệp vụ của TC — scenario, description, expected_result,
        then, test_data, precondition, steps (không chỉ title/tên module)
        — có chuẩn hoá bỏ dấu (ascii-fold) để không phụ thuộc cách gõ dấu
        tiếng Việt (vd "Ðăng nhập" dựng sẵn khác "Đăng nhập" tổ hợp).
        Kiểm tra failure TRƯỚC success vì "không thành công" chứa sẵn
        chuỗi con "thành công" — nếu kiểm tra success trước sẽ khớp nhầm.
        """
        if not isinstance(tc, dict):
            return 'failure'

        def _fold(value: object) -> str:
            import unicodedata
            s = str(value or '').strip().lower()
            s = unicodedata.normalize('NFD', s)
            return ''.join(ch for ch in s if unicodedata.category(ch) != 'Mn')

        text = _fold(' '.join(
            str(tc.get(k) or '') for k in (
                'scenario', 'description', 'title', 'expected_result', 'then',
                'test_data', 'precondition', 'steps',
            )
        ))
        if any(_fold(kw) in text for kw in self._FAILURE_SIGNAL_KEYWORDS):
            return 'failure'
        if any(_fold(kw) in text for kw in self._SUCCESS_SIGNAL_KEYWORDS):
            return 'success'
        return 'failure'

    def _group_tcs_by_base_function(self, modules: dict) -> dict:
        """
        Gộp TC của cặp "<Base> thành công" / "<Base> không thành công" (đã
        được _finalize_success_failure_grouping tách ra) trở lại theo
        CHỨC NĂNG GỐC — dùng bởi _validate_testcase_count/_evaluate_coverage
        để tính ngưỡng số lượng/loại kịch bản trên ĐÚNG phạm vi mà các
        ngưỡng đó vốn được thiết kế (1 chức năng gộp chung mọi loại kịch
        bản), thay vì tính nhầm trên từng nửa đã tách riêng.
        """
        grouped: dict[str, list] = {}
        for module_name, tcs in modules.items():
            if not isinstance(tcs, list):
                continue
            base = module_name.strip()
            lower = base.lower()
            for suffix in (' không thành công', ' thành công'):
                if lower.endswith(suffix):
                    base = module_name.strip()[: -len(suffix)].strip()
                    break
            grouped.setdefault(base, []).extend(tcs)
        return grouped

    def _dedup_tcs_in_bucket(self, tcs: list) -> list:
        """Loại TC trùng lặp nội dung trong CÙNG 1 nhóm thành công/không
        thành công đã gộp từ nhiều tên module khác nhau — trùng theo id
        (exact) hoặc trùng theo nội dung chuẩn hoá (exact hoặc fuzzy ratio
        rất cao, >= 0.92, để tránh gộp nhầm 2 TC khác ý nghĩa như đã ghi
        nhận ở bài học "category beats ratio" — ngưỡng 0.92 cao hơn hẳn mốc
        0.82 từng gây false-merge)."""
        def _norm_text(tc: dict) -> str:
            raw = ' '.join(
                str(tc.get(k) or '') for k in
                ('scenario', 'description', 'expected_result', 'then')
            )
            return re.sub(r'\s+', ' ', raw.strip().lower())

        seen_ids: set = set()
        seen_texts: list[str] = []
        result: list = []
        for tc in tcs:
            if not isinstance(tc, dict):
                continue
            tc_id = tc.get('id')
            if tc_id and tc_id in seen_ids:
                continue
            text = _norm_text(tc)
            is_dup = bool(text) and any(
                text == prev or difflib.SequenceMatcher(None, text, prev).ratio() >= 0.92
                for prev in seen_texts
            )
            if is_dup:
                continue
            result.append(tc)
            if tc_id:
                seen_ids.add(tc_id)
            if text:
                seen_texts.append(text)
        return result

    def _build_generic_outcome_tc(self, base: str, outcome: str) -> dict:
        """Sinh 1 TC tối thiểu hợp lý khi 1 chức năng gốc chỉ có TC ở 1
        trong 2 nhóm (vd chỉ có 'thành công' mà thiếu 'không thành công')
        — bổ sung để đảm bảo MỌI chức năng đều có đủ 2 nhóm, đúng yêu cầu
        "bổ sung các trường thành công và không thành công của tất cả các
        chức năng để sinh ra cho hợp lý"."""
        base_clean = base.strip()
        if outcome == 'success':
            scenario = f"{base_clean} thành công với dữ liệu/điều kiện hợp lệ."
            expected = f"Hệ thống thực hiện {base_clean.lower()} thành công, hiển thị/lưu kết quả đúng."
            test_type = 'Kiểm thử dương'
            title = f"{base_clean} thành công"
        else:
            scenario = f"{base_clean} không thành công do dữ liệu/điều kiện không hợp lệ."
            expected = (
                f"Hệ thống từ chối thực hiện {base_clean.lower()}, hiển thị thông báo lỗi phù hợp "
                "và không thay đổi dữ liệu."
            )
            test_type = 'Kiểm thử âm'
            title = f"{base_clean} không thành công"
        return {
            'scenario': scenario, 'description': scenario, 'title': title,
            'given': '', 'when': '', 'then': expected,
            'precondition': '', 'steps': '', 'test_data': '',
            'expected_result': expected, 'actual_result': '', 'status': 'Chưa chạy',
            'priority': 'Trung bình', 'test_type': test_type,
            'note': 'Tự động bổ sung để đảm bảo đủ 2 nhóm thành công/không thành công.',
        }

    def _finalize_success_failure_grouping(self, data: dict) -> dict:
        """
        BƯỚC BẮT BUỘC CUỐI CÙNG trước khi re-index TC ID (xem điểm gọi
        trong _normalize_test_cases). Với TOÀN BỘ chức năng hiện có (không
        phân biệt chức năng nào, không hard-code tên):
          1. Xác định chức năng gốc + qualifier tình huống cho từng module.
          2. Với từng TC: phân loại thành công/không thành công, giữ lại
             qualifier (nếu tên module cũ có) vào scenario/description.
          3. Gộp TC vào đúng 2 nhóm "<gốc> thành công" / "<gốc> không
             thành công" — các module khác tên nhưng CÙNG chức năng gốc
             (vd "Đăng nhập", "Đăng nhập SQL Injection", "Đăng nhập
             Validation") sẽ tự động hội tụ về cùng 1 trong 2 nhóm này.
          4. Khử TC trùng lặp trong từng nhóm.
          5. Bổ sung TC tối thiểu cho nhóm còn thiếu (chỉ có 1 trong 2 nhóm).
        Re-index TC ID được thực hiện NGAY SAU bởi vòng lặp có sẵn ở cuối
        _normalize_test_cases — không cần viết lại ở đây.
        """
        modules = data.get('modules', {})
        if not isinstance(modules, dict) or not modules:
            return data
        grouped: dict[str, list] = {}
        order: list[str] = []
        for module_name, tcs in modules.items():
            if not isinstance(tcs, list):
                continue
            if self._is_navigation_ui_module(module_name):
                for tc in tcs:
                    if not isinstance(tc, dict):
                        continue
                    tc.setdefault('chức năng', module_name)
                    tc.setdefault('feature', module_name)
                if module_name not in grouped:
                    grouped[module_name] = []
                    order.append(module_name)
                grouped[module_name].extend(tcs)
                continue
            base, qualifier = self._determine_base_business_function(module_name)
            base = base.strip() or module_name.strip()
            for tc in tcs:
                if not isinstance(tc, dict):
                    continue
                outcome = self._classify_tc_outcome(tc)
                suffix = 'thành công' if outcome == 'success' else 'không thành công'
                final_name = f"{base} {suffix}".strip()
                if qualifier:
                    existing_text = ' '.join(
                        str(tc.get(k) or '') for k in ('scenario', 'description')
                    ).lower()
                    if qualifier.lower() not in existing_text:
                        current_scenario = (tc.get('scenario') or tc.get('description') or '').strip()
                        tc_scenario = (
                            f"{qualifier[0].upper()}{qualifier[1:]}. {current_scenario}".strip()
                            if current_scenario else
                            f"{qualifier[0].upper()}{qualifier[1:]}"
                        )
                        tc['scenario'] = tc_scenario
                        tc['description'] = tc_scenario
                tc['chức năng'] = final_name
                tc['feature'] = final_name
                if final_name not in grouped:
                    grouped[final_name] = []
                    order.append(final_name)
                grouped[final_name].append(tc)
        for name in list(grouped.keys()):
            grouped[name] = self._dedup_tcs_in_bucket(grouped[name])
        bases_seen: dict[str, dict[str, bool]] = {}
        for name in order:
            if name.endswith(' không thành công'):
                base = name[: -len(' không thành công')].strip()
                bases_seen.setdefault(base, {})['failure'] = True
            elif name.endswith(' thành công'):
                base = name[: -len(' thành công')].strip()
                bases_seen.setdefault(base, {})['success'] = True
        for base, sides in bases_seen.items():
            if 'success' not in sides:
                name = f"{base} thành công"
                if name not in grouped:
                    grouped[name] = [self._build_generic_outcome_tc(base, 'success')]
                    order.append(name)
            if 'failure' not in sides:
                name = f"{base} không thành công"
                if name not in grouped:
                    grouped[name] = [self._build_generic_outcome_tc(base, 'failure')]
                    order.append(name)
        data['modules'] = {name: grouped[name] for name in order if grouped.get(name)}
        print("MODULES AFTER SUCCESS/FAILURE GROUPING:", list(data['modules'].keys()))
        return data

    def _dedupe_similar_modules(self, data: dict) -> dict:
        """
        Gộp các chức năng có CÙNG ý nghĩa nhưng tên khác nhau (vd "Tìm" và
        "Tìm kiếm theo mã hoặc tên chu kỳ") thành 1 chức năng duy nhất, để
        tránh sinh "dư" testcase trùng lặp chức năng.
        Test case trùng id giữa các chức năng được gộp sẽ chỉ giữ 1 bản.
        """
        modules = data.get('modules', {})
        if not isinstance(modules, dict) or len(modules) < 2:
            return data
        print("MODULES BEFORE MERGE:", list(modules.keys()))
        def _norm(name: str) -> str:
            cleaned = re.sub(r'[:,.\-–—…+←→↑↓✓✗]', ' ', name.strip().lower())
            return re.sub(r'\s+', ' ', cleaned).strip()
        _PREFIX_ONLY_GROUPS = {'phân trang', 'pagination'}
        def _group_index(name_lower: str) -> int | None:
            name_tokens = name_lower.split()
            for idx, group in enumerate(self._MODULE_CANONICAL_GROUPS):
                for g in group:
                    if g == name_lower:
                        return idx
                    if g in _PREFIX_ONLY_GROUPS and name_lower.startswith(g):
                        return idx
                    g_tokens = g.split()
                    n, t = len(name_tokens), len(g_tokens)
                    if t >= 2 and t <= n:  
                        if 'tiếp tục' in name_lower and 'tiếp tục' not in g:
                            continue
                        for i in range(n - t + 1):
                            if name_tokens[i:i+t] == g_tokens:
                                return idx
            return None
        merged: dict[str, list] = {}
        canonical_name_for_group: dict[int, str] = {}
        def _clean_display_name(name: str) -> str:
            return re.sub(r'^[+\-←→↑↓✓✗•▸›»\s]+', '', name).strip() or name
        for mod_name, tcs in modules.items():
            if not isinstance(tcs, list):
                continue
            mod_name = _clean_display_name(mod_name)
            name_lower = _norm(mod_name)
            group_idx = _group_index(name_lower)
            if group_idx is None:
                target_name = mod_name
            else:
                target_name = canonical_name_for_group.setdefault(group_idx, mod_name)
            if target_name not in merged:
                merged[target_name] = list(tcs)
            else:
                existing_ids = {tc.get('id') for tc in merged[target_name] if isinstance(tc, dict)}
                for tc in tcs:
                    if not isinstance(tc, dict):
                        continue
                    if tc.get('id') and tc.get('id') in existing_ids:
                        continue  
                    merged[target_name].append(tc)
                    if tc.get('id'):
                        existing_ids.add(tc.get('id'))
        data['modules'] = self._fold_generic_search_duplicates(merged)
        print("MODULES AFTER MERGE:", list(data['modules'].keys()))
        return data

    def _fold_generic_search_duplicates(self, modules: dict) -> dict:
        """
        Lưới an toàn CUỐI CÙNG cho canonical group Search — chạy SAU
        _group_index() ở trên (vốn vẫn coi "tìm kiếm"/"ô tìm kiếm"/"search"
        chung chung là 1 group RIÊNG với các group "tìm kiếm theo [X]" cụ
        thể, để không phá vỡ trường hợp hợp lệ có nhiều ô lọc field khác
        nhau thật sự). Hàm này bắt đúng case gây bug thực tế: chức năng
        "Tìm kiếm" (tên chung chung, KHÔNG tương ứng UI element nào riêng)
        bị sinh THỪA bên cạnh 1 chức năng cụ thể-hơn đã có (vd "Tìm kiếm theo
        mã hoặc tên chu kỳ") — gộp TC của chức năng chung chung đó vào chức năng
        cụ thể rồi xoá chức năng chung chung, thay vì giữ cả 2 làm 2 chức năng
        khác nhau.
        Nút "Tìm" (SEARCH_BUTTON_NAMES) KHÔNG bị đụng tới — đó là 1 UI
        element khác (button, không phải input), vẫn giữ làm chức năng riêng
        nếu UI thực sự có nút đó tách biệt.
        Dùng canonical_module_key/is_search_module DÙNG CHUNG với
        coverage_checker.py (xem import _cc_is_search_module ở đầu file)
        để đảm bảo generate/normalize và Coverage Checker không lệch nhau
        về định nghĩa "cùng 1 chức năng Search".
        """
        if not isinstance(modules, dict) or len(modules) < 2:
            return modules
        generic_names: list[str] = []
        specific_names: list[str] = []
        for name in modules.keys():
            if not isinstance(modules.get(name), list):
                continue
            n = name.strip().lower()
            if n in _cc_search_generic_names:
                generic_names.append(name)
            elif n not in _cc_search_button_names and _cc_is_search_module(n):
                specific_names.append(name)
        if not generic_names or not specific_names:
            return modules
        target = specific_names[0]
        existing_ids = {
            tc.get('id') for tc in modules[target]
            if isinstance(tc, dict) and tc.get('id')
        }
        for g_name in generic_names:
            for tc in modules.get(g_name, []):
                if not isinstance(tc, dict):
                    continue
                tc_id = tc.get('id')
                if tc_id and tc_id in existing_ids:
                    continue
                modules[target].append(tc)
                if tc_id:
                    existing_ids.add(tc_id)
            del modules[g_name]

        return modules
    _FILTER_CANONICAL_GROUPS: list[set[str]] = [
        {'tìm kiếm theo mã', 'tìm theo mã', 'search by code',
         'tìm kiếm theo mã hoặc kho', 'tìm kiếm theo mã hoặc tên'},
        {'tìm kiếm theo tên', 'tìm theo tên', 'search by name'},
        {'tìm kiếm theo kho', 'tìm theo kho'},
        {'tìm kiếm', 'ô tìm kiếm', 'search'},
        {'tìm', 'nút tìm', 'button tìm'},
        {'quay lại', 'back', 'nút quay lại'},
        {'thêm mới', 'thêm', 'add', 'tạo mới'},
        {'xóa', 'xoá', 'delete'},
        {'cập nhật', 'nút cập nhật', 'icon cập nhật'},
        {'chỉnh sửa', 'edit', 'update'},
        {'xuất excel', 'excel', 'xuất file excel', 'xuất file'},
        {'xuất word', 'word', 'xuất file word'},
        {'phân trang', 'pagination', 'trang'},
        {'số dòng mỗi trang', 'page-size-dropdown', 'hiển thị số dòng', 'page size'},
        {'xóa điều kiện tìm kiếm', 'clear search', 'xóa tìm kiếm', 'nút x tìm kiếm'},
        {'sinh mã', 'tạo mã', 'generate code'},
        {'hủy bỏ', 'huỷ bỏ', 'hủy', 'huỷ', 'cancel'},
        {'đóng popup', 'đóng', 'close', 'nút x'},
        {'thêm mới và tiếp tục', 'lưu và tiếp tục', 'thêm và tiếp tục', 'save and continue'},
        {'xem chi tiết', 'xem', 'view', 'chi tiết'},
        {'xem & kiểm kê', 'xem và kiểm kê', 'kiểm kê'},
        {'sửa kế hoạch', 'chỉnh sửa kế hoạch'},
        {'xóa đợt kiểm kê', 'xoá đợt kiểm kê'},
        {'xóa kế hoạch', 'xoá kế hoạch'},
    ]
    def _parse_description_phrases(self, description: str) -> list[str]:
        """
        Tách description thành từng phrase chức năng riêng biệt, dùng CHUNG
        1 quy tắc tách cho mọi nơi cần đọc "người dùng yêu cầu đúng những
        chức năng nào" (_filter_modules_by_description, _resolve_requested_
        canonicals...). VD: "phân quyền" -> ["phân quyền"];
        "Website ...: đăng nhập, phân quyền" -> ["đăng nhập", "phân quyền"].
        """
        if not description or not description.strip():
            return []
        phrase_source = description.split(':', 1)[1] if ':' in description else description
        phrases = [p.strip() for p in re.split(r'[,\n]', phrase_source) if p.strip()]
        if not phrases:
            stripped = description.strip()
            phrases = [stripped] if stripped else []
        return phrases

    def _resolve_requested_canonicals(self, description: str) -> dict[str, str]:
        """
        Map mỗi phrase mà người dùng yêu cầu (vd "phân quyền", "chấm công",
        "phân quyền thành công", "tạo testcase cho chức năng phân quyền")
        sang canonical key ĐÃ có fixed template tập trung VÀ đang nằm trong
        tập enforce (_SRE_DEFAULT_ENFORCED_CANONICALS — gồm cả "phan_quyen",
        "cham_cong"). Trả về {canonical: phrase_gốc_khớp_đầu_tiên}.
        Đây là nguồn canonical DUY NHẤT lấy trực tiếp từ yêu cầu người dùng,
        không suy diễn từ tên dự án, mô tả cũ, lịch sử hội thoại, hay ảnh.
        """
        result: dict[str, str] = {}
        if _sre_normalize_function_name is None:
            return result
        for phrase in self._parse_description_phrases(description):
            canonical = _sre_normalize_function_name(phrase)
            if canonical and canonical in _SRE_DEFAULT_ENFORCED_CANONICALS:
                result.setdefault(canonical, phrase)
        return result

    def _apply_known_canonical_templates(self, data: dict, description: str) -> dict:
        """
        Lớp chặn cứng CUỐI CÙNG cho targeted mode: với các chức năng người
        dùng yêu cầu mà đã có fixed template tập trung trong
        scenario_rule_engine.py (vd "phan_quyen", "cham_cong", "dang_nhap",
        "quay_lai"...), LUÔN build trực tiếp bằng build_testcases_for_module()
        — KHÔNG phụ thuộc việc AI có sinh đúng tên module tương ứng hay
        không, và KHÔNG cho phép các module AI tự suy diễn thêm (vd "Kiểm
        tra dị ứng", "Quản lý đơn thuốc" khi user chỉ yêu cầu "phân quyền")
        lọt qua.

        - Nếu TOÀN BỘ phrase trong description đều là canonical đã enforce
          (request "thuần" — như input chỉ "phân quyền"): XÓA HẲN mọi module
          khác AI sinh ra, CHỈ giữ đúng các module fixed template tương ứng.
        - Nếu chỉ MỘT PHẦN phrase khớp canonical đã enforce (vd user vừa
          yêu cầu "phân quyền" vừa yêu cầu "thêm mới" — "thêm mới" thuộc
          nhóm CRUD field-aware, cần AI đọc field thật từ ảnh, không có
          template cố định để enforce cứng): CHỈ ép canonical đã enforce,
          giữ nguyên module khác do AI sinh — không đụng tới pipeline CRUD
          field-aware hiện có.
        - Nếu description không khớp canonical nào đã enforce: giữ nguyên
          data, không thay đổi gì (an toàn cho mọi flow khác).
        """
        if _sre_normalize_function_name is None or _sre_build_testcases_for_module is None:
            return data
        if not isinstance(data, dict):
            return data
        phrases = self._parse_description_phrases(description)
        if not phrases:
            return data
        requested_canonicals = self._resolve_requested_canonicals(description)
        print(f"[TargetedParse] Chức năng người dùng yêu cầu (phrases): {phrases}")
        print(f"[TargetedParse] Canonical đã parse (có fixed template): {requested_canonicals}")
        if not requested_canonicals:
            return data
        modules = data.get('modules', {})
        if not isinstance(modules, dict):
            modules = {}
        forced_modules: dict[str, list[dict]] = {}
        for canonical, phrase in requested_canonicals.items():
            built = _sre_build_testcases_for_module(phrase, canonical, "") or {}
            for disp_name, tcs in built.items():
                forced_modules[disp_name] = tcs
        full_scope = len(requested_canonicals) == len(phrases)
        if full_scope:
            data['modules'] = forced_modules
            print(
                "[TargetedParse] Toàn bộ chức năng yêu cầu đều có fixed template "
                f"→ CHỈ giữ đúng {len(forced_modules)} module: {list(forced_modules.keys())} "
                "(loại bỏ mọi module khác AI có thể đã tự sinh thêm)"
            )
            return data
        new_modules: dict[str, list[dict]] = dict(forced_modules)
        for mod_name, tcs in modules.items():
            mod_canonical = _sre_normalize_function_name(mod_name)
            if mod_canonical in requested_canonicals:
                continue
            new_modules.setdefault(mod_name, tcs)
        data['modules'] = new_modules
        print(
            "[TargetedParse] Một phần chức năng có fixed template, phần còn lại giữ "
            f"nguyên AI sinh → danh sách module sau khi ép/lọc: {list(new_modules.keys())}"
        )
        return data

    def _filter_modules_by_description(self, data: dict, description: str) -> dict:
        """
        Lớp chặn cứng cho targeted request: chỉ giữ lại các chức năng mà AI sinh ra
        nếu tên chức năng thực sự khớp với (1 trong) các phrase mà user yêu cầu.
        Chức năng nào AI tự thêm vào nhưng KHÔNG khớp yêu cầu (vd "Hiển thị", "Mô tả"
        khi user chỉ yêu cầu "tìm kiếm, tìm, quay lại") sẽ bị loại bỏ ngay,
        không chờ tới bước normalize/merge.
        Matching dùng word-boundary (exact word), không phải substring lỏng:
        - term "tìm" CHỈ khớp chức năng nếu "tìm" là từ riêng trong tên chức năng
          (vd "Tìm kiếm" → OK, "Tìm kiếm và lọc" → OK, "Kết quả tìm" → OK,
           nhưng "Tìm kiếm theo mã" khi term là "tìm" → cũng OK vì "tìm" xuất
           hiện ở đầu; ngược lại "Định dạng" không chứa từ "tìm" → loại).
        - Ngoài ra, canonical groups đảm bảo "tìm kiếm theo mã hoặc kho"
         được coi là cùng 1 chức năng → chức năng nào khớp cũng được giữ.
        """
        modules = data.get('modules', {})
        if not isinstance(modules, dict) or not modules:
            return data
        phrase_source = description.split(':', 1)[1] if ':' in description else description
        phrases = [p.strip().lower() for p in re.split(r'[,\n]', phrase_source) if p.strip()]
        if not phrases:
            return data

        synonym_map = {
            'tìm kiếm theo mã': ['tìm theo mã', 'search by code'],
            'tìm kiếm theo tên': ['tìm theo tên', 'search by name'],
            'tìm kiếm': ['ô tìm kiếm', 'search box'],
            'tìm': ['nút tìm', 'button tìm'],
            'quay lại': ['back', 'nút quay lại'],
            'thêm mới': ['thêm', 'add', 'tạo mới'],
            'xóa': ['delete', 'xoá', 'thùng rác'],
            'cập nhật': ['sửa', 'chỉnh sửa', 'edit', 'update'],
            'xuất excel': ['excel', 'xuất file'],
            'xuất word': ['word'],
            'phân trang': ['pagination'],
            'sinh mã': ['tạo mã', 'generate code'],
            'phân quyền': ['permission', 'role', 'phan quyen'],
            'chấm công': ['attendance', 'check in', 'check-in', 'cham cong'],
        }
        raw_terms: set[str] = set()
        for phrase in phrases:
            clean = re.sub(r'^(nút|ô|icon|chức năng)\s+', '', phrase).strip()
            raw_terms.add(clean)
            for key, syns in synonym_map.items():
                if key in clean or clean in key:
                    raw_terms.add(key)
                    raw_terms.update(syns)
        allowed_terms: set[str] = set(raw_terms)
        for group in self._FILTER_CANONICAL_GROUPS:
            if raw_terms & group:       
                allowed_terms |= group   
        allowed_terms = {t for t in allowed_terms if len(t) >= 3}
        def _word_match(term: str, name_lower: str) -> bool:
            """
            Trả về True nếu term xuất hiện ở word-boundary trong name_lower,
            HOẶC name_lower là substring/exact của term (cho phép tên chức năng
            ngắn hơn như "Tìm kiếm" khớp term "tìm kiếm theo mã hoặc kho").
            Nếu mà nhận thấy màu vàng  thì " cập nhật " 
            Nếu mà nhận thấy màu đỏ thì "xóa"
            Nếu mà nhận thấy màu xanh thì "thêm mới"
            Nếu mà nhận thấy màu xanh lá  "xuất file Excel"
            """
            if name_lower == term or name_lower in term:
                return True
            name_tokens = re.split(r'\s+', name_lower)
            term_tokens = re.split(r'\s+', term)
            n, t = len(name_tokens), len(term_tokens)
            for i in range(n - t + 1):
                if name_tokens[i:i + t] == term_tokens:
                    return True
            if len(term) >= 5 and term in name_lower:
                return True
            return False
        kept = {}
        description_lower = description.lower()
        if "quay lại" in description_lower:
            allowed_terms.add("quay lại")
        for mod_name, tcs in modules.items():
            name_lower = mod_name.strip().lower()
            if any(_word_match(term, name_lower) for term in allowed_terms):
                kept[mod_name] = tcs
        if kept:
            data['modules'] = kept
        else:
            print(
                "=== FILTER MODULES BY DESCRIPTION: 0 chức năng khớp allowed_terms "
                f"{sorted(allowed_terms)} — GIỮ NGUYÊN {len(modules)} module gốc "
                "(có thể lọt chức năng dư, kiểm tra lại parsing description) ==="
            )
        return data
    _ACTION_WHITELIST = [
        'tìm', 'kiếm', 'thêm', 'sửa', 'cập nhật', 'chỉnh sửa', 'xóa', 'xoá',
        'xem', 'chi tiết', 'phân trang', 'trang', 'dòng', 'quay lại', 'back',
        'xuất', 'nhập', 'lọc', 'filter', 'sắp xếp', 'sort', 'chọn', 'đăng',
        'lưu', 'hủy', 'huỷ', 'export', 'import', 'search', 'thao tác',
        'duyệt', 'gửi', 'tải', 'upload', 'download', 'in ', 'print', 'khóa',
        'khoá', 'kích hoạt', 'vô hiệu', 'sao chép', 'copy', 'mở', 'đóng',
        'reset', 'làm mới', 'refresh', 'validate', 'kiểm tra', 'xác thực',
        'phê duyệt', 'từ chối', 'gán', 'phân quyền', 'đổi', 'cấu hình',
        'sinh', 'popup', 'chấm công', 'chấm',
    ]
    def _drop_static_data_modules(self, data: dict) -> dict:
        """
        Lớp chặn cứng: loại bỏ MỌI chức năng không liên quan tới một hành động
        UI thực sự (không nằm trong _ACTION_WHITELIST). Đây thường là các
        chức năng AI "bịa" ra để kiểm tra hiển thị 1 giá trị dữ liệu cụ thể có
        sẵn trong hệ thống (vd chức năng tên "KHO-VTNN", "Mã", "Tên", hoặc chức năng
        kiểm tra tiêu đề màn hình) — KHÔNG phải yêu cầu kiểm thử hợp lệ.
        Áp dụng cho CẢ targeted và full-scan, không chỉ targeted.
        """
        modules = data.get('modules', {})
        if not isinstance(modules, dict) or not modules:
            return data
        kept = {}
        for mod_name, tcs in modules.items():
            name_lower = mod_name.strip().lower()
            if any(kw in name_lower for kw in self._ACTION_WHITELIST):
                kept[mod_name] = tcs
        if kept:
            data['modules'] = kept
        return data
    _THEM_MOI_NOTE = "Trường hợp nhập dữ liệu không phù hợp, không lưu thông tin và có thông báo lỗi phù hợp"
    _POPUP_ACTION_4TC_KEYS = ('sinh mã', 'hủy bỏ', 'huỷ bỏ', 'đóng popup')
    # BUG ĐÃ SỬA: trước đây chỉ khớp "hủy bỏ"/"huỷ bỏ" (substring), KHÔNG
    # khớp "hủy"/"huỷ"/"cancel" trần trụi — dẫn tới việc module tên ĐÚNG
    # "Hủy" (không có drift/chêm object gì) vẫn bị _finalize_success_failure_
    # grouping tách thành "Hủy thành công"/"Hủy không thành công" TRƯỚC khi
    # scenario_rule_engine.replace_generated_cases_with_template kịp chạy,
    # khiến normalize_function_name("Hủy thành công") không khớp exact nữa
    # và fixed template (3 kịch bản chuẩn) KHÔNG BAO GIỜ được áp dụng.
    # Dùng \b word-boundary (không phải substring thường) cho các từ NGẮN
    # "hủy"/"huỷ"/"cancel" để tránh khớp nhầm vào chữ khác chứa chuỗi con
    # tương tự (vd "hủy hoại"); các key dài đã có ('sinh mã', 'đóng popup')
    # đủ đặc trưng nên giữ nguyên kiểu substring như cũ.
    _POPUP_ACTION_4TC_BARE_WORD_RE = re.compile(r'\b(hủy|huỷ|cancel)\b')
    def _is_popup_action_module_4tc(self, module_name_lower: str) -> bool:
        """True nếu chức năng là 1 trong các action button modal cần đúng 4 TC:
        Sinh mã / Hủy (hủy bỏ) / Đóng popup. KHÔNG khớp "Thêm mới và tiếp tục"
        (chức năng đó cần đúng 2 TC, xử lý riêng bởi logic Thêm mới/generic)."""
        lower = re.sub(r'\s+', ' ', module_name_lower.strip())
        if any(key in lower for key in self._POPUP_ACTION_4TC_KEYS):
            return True
        return bool(self._POPUP_ACTION_4TC_BARE_WORD_RE.search(lower))
    def _is_navigation_ui_module(self, module_name: str) -> bool:
        """
        True nếu chức năng là điều hướng/UI (Navigation/UI Actions): Quay
        lại, Phân trang, Đóng popup/Hủy popup/Sinh mã tự động, Mở popup...
        — KHÔNG phải chức năng nghiệp vụ (Business Actions: Đăng nhập,
        Thêm, Cập nhật, Xóa, Tìm kiếm, Chuyển khoản, Thanh toán, Xuất
        Excel, In...).
        Các chức năng điều hướng/UI này đã có rule số lượng/kịch bản CỐ
        ĐỊNH riêng (_enforce_quay_lai_exact_four, _is_popup_action_module_4tc,
        _enforce_generic_module_exact_two...) nên PHẢI loại trừ khỏi
        _finalize_success_failure_grouping — nếu không, việc gộp về đúng 2
        nhóm "thành công/không thành công" sẽ xóa mất ý nghĩa từng kịch bản
        riêng (vd 4 kịch bản: chưa đổi dữ liệu/đã đổi dữ liệu/từ màn hình
        chi tiết/không có màn hình trước của "Quay lại").
        Dùng substring match (không phải exact match) để nhất quán với
        cách _validate_testcase_count/_ensure_final_testcase_counts đang
        nhận diện các chức năng này — tránh lệch tên như đã gặp ở
        _is_back_module (chỉ exact-match 4 chuỗi cứng, dễ bỏ sót biến thể).
        """
        lower = re.sub(r'\s+', ' ', (module_name or '').strip().lower())
        if not lower:
            return False
        if 'quay lại' in lower or 'quay lai' in lower or lower == 'back':
            return True
        if 'phân trang' in lower or 'phan trang' in lower:
            return True
        if self._is_popup_action_module_4tc(lower):
            return True
        if 'mở popup' in lower or 'mo popup' in lower:
            return True
        return False

    # canonical (theo scenario_rule_engine.normalize_function_name) của các
    # nút hành động UI CHUNG, không gắn với 1 đối tượng nghiệp vụ cụ thể —
    # đây là những canonical mà nếu AI tự chêm thêm tên đối tượng nghiệp vụ
    # (vd "Hủy lớp học" thay vì "Hủy") thì chắc chắn là LỖI, cần rollback.
    _GENERIC_ACTION_PAD_GUARD_CANONICALS = frozenset({"huy", "dong_popup", "sinh_ma"})

    def _realign_generic_action_module_names(self, description: str | None, modules: dict) -> dict:
        """
        SAFETY NET: chống lỗi "Hủy" (hoặc "Đóng popup"/"Sinh mã") bị AI tự
        SINH RA với tên module đã bị CHÊM THÊM tên đối tượng nghiệp vụ lấy
        từ ngữ cảnh dự án — vd người dùng yêu cầu đúng "hủy" nhưng AI trả
        về module "Hủy lớp học" vì dự án đang là "Quản lý lớp học".
        scenario_rule_engine.normalize_function_name() KHÔNG tự bắt được
        case này vì nó CHỈ khớp CHÍNH XÁC "hủy"/"huỷ"/"cancel"... (đúng
        theo yêu cầu — để KHÔNG nuốt nhầm 1 hành động nghiệp vụ THẬT SỰ
        được đặt tên rõ ràng, vd "Hủy lịch hẹn" do người dùng tự gõ) —
        "Hủy lớp học" không khớp exact nên bị normalize_function_name bỏ
        qua, và module cứ thế trôi qua replace_generated_cases_with_template
        mà KHÔNG được thay bằng fixed template đúng.
        Hàm này chạy TRƯỚC _finalize_success_failure_grouping / trước khi
        scenario_rule_engine override, và CHỈ đổi tên module khi TẤT CẢ
        điều kiện sau đúng — mỗi điều kiện tương ứng 1 rủi ro cụ thể cần
        tránh:
          1. Người dùng có yêu cầu (trong `description`, tách theo dấu
             phẩy/;/+//và — CÙNG cách _coverage_checker đang tách) một mục
             mà bản thân nó CHÍNH XÁC là 1 canonical hành động UI chung
             (huy/dong_popup/sinh_ma) — vd đúng "hủy", không kèm từ nào
             khác. Nếu người dùng không hề yêu cầu như vậy thì không có gì
             để đối chiếu, không đụng vào.
          2. Modules hiện tại CHƯA có module nào khớp CHÍNH XÁC canonical
             đó — nếu đã có rồi thì không cần sửa, tránh làm trùng/mất dữ
             liệu.
          3. Đúng 1 (không phải 0, không phải nhiều hơn 1 — mơ hồ thì
             không đoán) module có tên bắt đầu bằng đúng từ người dùng đã
             gõ, kèm hậu tố phía sau (vd "Hủy lớp học" bắt đầu bằng "hủy").
          4. Tên ĐẦY ĐỦ của module nghi ngờ đó, tự nó, KHÔNG PHẢI là 1
             canonical khác đã biết (an toàn — không đụng vào 1 chức năng
             cố định khác chẳng may trùng tiền tố).
          5. Cụm từ ĐẦY ĐỦ đó (vd "hủy lớp học") KHÔNG xuất hiện trong
             chính `description` gốc của người dùng — tức người dùng CHƯA
             từng tự gõ đúng cụm đó. Nếu người dùng ĐÃ tự gõ "hủy lớp học"
             thì đây là 1 hành động nghiệp vụ RIÊNG do người dùng chủ định,
             PHẢI giữ nguyên, TUYỆT ĐỐI không rollback (đúng quy tắc: chỉ
             coi là hành động hủy nghiệp vụ khi người dùng nói rõ đầy đủ).
        Nếu cả 5 điều kiện đúng: đổi tên KEY của module đó về đúng tên hiển
        thị chuẩn (viết hoa chữ cái đầu của đúng cụm người dùng đã gõ, vd
        "Hủy"), gộp toàn bộ TC bên trong vào key mới, và cập nhật lại field
        'chức năng'/'feature'/'module' của từng TC đang trỏ tới tên cũ.
        """
        if not description or not isinstance(modules, dict) or not modules:
            return modules
        if _sre_normalize_function_name is None:
            return modules

        raw_terms = re.split(r'[,;+/]|(?:\bvà\b)', description)
        terms = [t.strip() for t in raw_terms if t.strip()]
        if not terms:
            return modules

        desc_lower = description.strip().lower()

        for term in terms:
            term_canonical = _sre_normalize_function_name(term)
            if term_canonical is None or term_canonical not in self._GENERIC_ACTION_PAD_GUARD_CANONICALS:
                continue
            already_exact = any(
                _sre_normalize_function_name(name) == term_canonical for name in modules.keys()
            )
            if already_exact:
                continue
            term_norm = term.strip().lower()
            candidates = []
            for name in list(modules.keys()):
                name_lower = re.sub(r'\s+', ' ', str(name).strip().lower())
                if not name_lower.startswith(term_norm):
                    continue
                suffix = name_lower[len(term_norm):].strip()
                if not suffix:
                    continue
                if _sre_normalize_function_name(name) is not None:
                    continue
                if name_lower in desc_lower:
                    continue
                candidates.append(name)
            if len(candidates) != 1:
                continue
            drifted_name = candidates[0]
            display_name = term.strip()
            display_name = display_name[0].upper() + display_name[1:]
            if drifted_name == display_name:
                continue
            tcs = modules.pop(drifted_name)
            if not isinstance(tcs, list):
                tcs = []
            for tc in tcs:
                if isinstance(tc, dict):
                    for key in ('chức năng', 'feature', 'module'):
                        if tc.get(key) == drifted_name:
                            tc[key] = display_name
            if display_name in modules and isinstance(modules.get(display_name), list):
                modules[display_name].extend(tcs)
            else:
                modules[display_name] = tcs
            print(
                f"[RealignGenericAction] Module '{drifted_name}' bị AI tự chêm thêm "
                f"đối tượng nghiệp vụ dù người dùng chỉ yêu cầu '{term.strip()}' "
                f"-> đổi lại tên module thành '{display_name}'."
            )
        return modules

    def _is_them_moi_module(self, module_name: str) -> bool:
        """
        True CHỈ KHI chức năng thực sự là "Thêm mới" (nút lưu chính), KHÔNG
        khớp "Thêm mới và tiếp tục" (nút lưu-và-giữ-popup-mở — chức năng
        HOÀN TOÀN RIÊNG, có 2 TC với ý nghĩa khác hẳn: "lưu thành công và
        giữ popup mở" / "lỗi validation, không lưu").
        BUG ĐÃ SỬA: trước đây dùng `'thêm' in module_name.lower()` — vì
        "Thêm mới và tiếp tục" cũng chứa chữ "thêm" nên bị enforce nhầm
        theo format 2-TC của "Thêm mới" (thành công/không thành công +
        note lỗi merge bullet), làm sai lệch nội dung và có thể trộn lẫn
        TC giữa 2 chức năng khi dedupe theo category ở _merge_test_cases.
        """
        lower = re.sub(r'\s+', ' ', module_name.strip().lower())
        if 'tiếp tục' in lower:
            return False
        return 'thêm' in lower
    _AUTOGEN_HINT_KEYWORDS = ('tự sinh', 'tự động sinh', 'auto generate', 'sinh mã tự động')
    def _categorize_them_moi_detail(self, text: str) -> str:
        """Phân loại CHI TIẾT 1 TC của chức năng 'Thêm mới' theo NHÓM NGHIỆP VỤ
        (không còn gộp chung mọi lỗi vào 1 nhóm 'negative' duy nhất):
        - 'add_more'         : lưu rồi thêm tiếp (nút "Thêm mới và tiếp tục")
        - 'cancel'            : hủy thao tác — không thuộc phạm vi chức năng này
        - 'negative_security' : XSS / SQL Injection
        - 'negative_format'   : trùng dữ liệu / sai định dạng / vượt độ dài
        - 'negative_required' : bỏ trống / chỉ nhập khoảng trắng ở trường bắt buộc
        - 'negative_other'    : lỗi input khác không rơi vào 3 nhóm trên
        - 'positive'          : thành công (mặc định)
        Thứ tự kiểm tra ưu tiên nhóm ĐẶC TRƯNG nhất trước (security → format
        → required) để tránh 1 câu vừa có "trống" vừa có "XSS" bị xếp nhầm
        nhóm ít đặc trưng hơn.
        """
        t = text.lower()
        if any(k in t for k in ('tiếp tục', 'thêm tiếp', 'lưu rồi thêm')):
            return 'add_more'
        if any(k in t for k in ('hủy', 'huỷ')):
            return 'cancel'
        if any(k in t for k in ('xss', 'sql injection', 'sqli', 'mã độc', 'script', 'injection')):
            return 'negative_security'
        if any(k in t for k in ('trùng', 'sai định dạng', 'vượt quá', 'vượt độ dài', 'độ dài', 'ký tự đặc biệt')):
            return 'negative_format'
        if any(k in t for k in ('để trống', 'bỏ trống', 'không nhập', 'thiếu trường', 'khoảng trắng', 'bắt buộc')):
            return 'negative_required'
        if any(k in t for k in ('không hợp lệ', 'không thành công', 'không tồn tại')):
            return 'negative_other'
        return 'positive'
    def _categorize_add_tc(self, text: str) -> str:
        """Bản rút gọn của _categorize_them_moi_detail, gộp mọi nhóm
        'negative_*' về chung 'negative' — giữ lại cho các nơi khác trong
        file chỉ cần phân biệt thành công/không thành công/thêm tiếp/hủy."""
        cat = self._categorize_them_moi_detail(text)
        return 'negative' if cat.startswith('negative') else cat
    def _vn_join_names(self, names: list) -> str:
        """Nối danh sách tên trường theo văn phong tiếng Việt: "A và B",
        "A, B và C"... đồng thời khử trùng lặp, giữ nguyên thứ tự xuất hiện."""
        seen = set()
        uniq = []
        for n in names:
            n = (n or '').strip()
            if n and n not in seen:
                seen.add(n)
                uniq.append(n)
        if not uniq:
            return ''
        if len(uniq) == 1:
            return uniq[0]
        if len(uniq) == 2:
            return f"{uniq[0]} và {uniq[1]}"
        return ", ".join(uniq[:-1]) + f" và {uniq[-1]}"
    def _them_moi_field_groups(self):
        """Phân tích Form Structure (self._current_form_structure) thành các
        nhóm field dùng để tách testcase của chức năng "Thêm mới" theo NHÓM
        NGHIỆP VỤ hợp lý:
        - fields            : toàn bộ field thực tế của form
        - required_fields   : field bắt buộc (Required/Whitespace)
        - autogen_fields     : field định danh được phép bỏ trống để hệ
          thống tự sinh (vd "Mã kho") — nhận diện qua business_rules HOẶC
          field không bắt buộc có tên chứa "mã"
        - dup_format_fields  : field có khả năng bị trùng/sai định dạng/vượt
          độ dài (ưu tiên field dạng "mã"/định danh, fallback field bắt
          buộc, fallback toàn bộ field)
        """
        form_structure = self._current_form_structure or {}
        fields = [
            f for f in (form_structure.get('fields') or [])
            if isinstance(f, dict) and (f.get('name') or '').strip()
        ]
        rules = [str(r) for r in (form_structure.get('business_rules') or [])]
        required_fields = [f for f in fields if f.get('required')]
        def _is_autogen(f) -> bool:
            name = (f.get('name') or '').strip().lower()
            if not name:
                return False
            rule_hit = any(
                name in r.lower() and any(k in r.lower() for k in self._AUTOGEN_HINT_KEYWORDS)
                for r in rules
            )
            if rule_hit:
                return True
            return (not f.get('required')) and 'mã' in name
        autogen_fields = [f for f in fields if _is_autogen(f)]
        dup_format_fields = (
            [f for f in fields if 'mã' in (f.get('name') or '').lower()]
            or required_fields
            or fields
        )
        return fields, required_fields, autogen_fields, dup_format_fields
    def _them_moi_fallback_two_tc(self, filtered: list) -> list:
        """Tạo đúng 2 TC cho nút Thêm mới khi CHƯA thấy form chi tiết.
        Không tái sử dụng scenario/expected_result do AI sinh vì nội dung đó
        có thể đã lẫn Required, Duplicate, XSS hoặc thao tác Lưu.
        """
        return [
            {
                "title": "Mở form thêm mới thành công",
                "scenario": "Nhấn nút Thêm mới trên màn hình danh sách",
                "description": "Nhấn nút Thêm mới trên màn hình danh sách",
                "given": "Người dùng đang ở màn hình danh sách và có quyền thêm mới",
                "when": "Người dùng nhấn nút Thêm mới",
                "then": "Hệ thống mở đúng form thêm mới",
                "precondition": "Màn hình danh sách hiển thị bình thường",
                "steps": "1. Mở màn hình danh sách\n2. Nhấn nút Thêm mới",
                "test_data": "Không áp dụng",
                "expected_result": "Hệ thống mở đúng form thêm mới; chưa thực hiện lưu dữ liệu",
                "priority": "Cao",
                "test_type": "Kiểm thử dương",
                "actual_result": "",
                "status": "Chưa chạy",
                "note": "",
            },
            {
                "title": "Mở form thêm mới không thành công",
                "scenario": "Nhấn Thêm mới khi không có quyền hoặc form không thể tải",
                "description": "Nhấn Thêm mới khi không có quyền hoặc form không thể tải",
                "given": "Người dùng không có quyền thêm mới hoặc hệ thống gặp lỗi tải form",
                "when": "Người dùng nhấn nút Thêm mới",
                "then": "Hệ thống không mở form và hiển thị thông báo phù hợp",
                "precondition": "Người dùng đang ở màn hình danh sách",
                "steps": "1. Mở màn hình danh sách\n2. Nhấn nút Thêm mới",
                "test_data": "Tài khoản không có quyền hoặc dịch vụ tải form lỗi",
                "expected_result": "Hệ thống không mở form, hiển thị thông báo phù hợp và giữ nguyên dữ liệu danh sách",
                "priority": "Cao",
                "test_type": "Kiểm thử âm",
                "actual_result": "",
                "status": "Chưa chạy",
                "note": "",
            },
        ]

    def _enforce_them_moi_format(self, data: dict) -> dict:
        """Chuẩn hoá chức năng 'Thêm mới' — KHÔNG còn ép cứng đúng 2 TC.

        Với form có field thực tế (Form Understanding đã phân tích được
        self._current_form_structure.fields), tách testcase theo NHÓM
        NGHIỆP VỤ hợp lý, mỗi TC chỉ mô tả 1 nhóm lỗi liên quan (không nhồi
        checklist 8-9 loại validation vào 1 ô Mô tả):
          TC1  Thành công — nhập đầy đủ dữ liệu hợp lệ
          TC2  (nếu có field định danh cho phép bỏ trống) — để trống, hệ
               thống tự sinh mã hợp lệ, không trùng, lưu thành công
          TC3  (nếu có field bắt buộc) — nhóm lỗi Required/Whitespace: bỏ
               trống hoặc chỉ nhập khoảng trắng → lỗi bắt buộc, không lưu
          TC4  (nếu có field liên quan) — nhóm lỗi Duplicate/Format/Length:
               trùng, sai định dạng hoặc vượt độ dài → lỗi phù hợp, không lưu
          TC5  Nhóm lỗi Security: XSS/SQL Injection → không thực thi mã,
               không lưu dữ liệu nguy hiểm

        Với form kho hiện tại (Mã kho + Tên kho), kết quả ĐÚNG 5 TC theo
        đúng 5 nhóm trên. Với form khác, số TC co giãn theo field thực tế
        (vd form không có field tự sinh mã thì không có TC2).

        Nếu chức năng không có Form Structure/field cụ thể (vd chỉ là nút bấm
        không có form nhập liệu), fallback về 2 TC gọn (thành công/không
        thành công) — xem _them_moi_fallback_two_tc.
        """
        modules = data.get('modules', {})
        for module_name, tcs in list(modules.items()):
            if not isinstance(tcs, list) or not tcs:
                continue
            if not self._is_them_moi_module(module_name):
                continue
            filtered = []
            for tc in tcs:
                if not isinstance(tc, dict):
                    continue
                text = ' '.join(str(tc.get(k) or '') for k in ('scenario', 'description', 'title'))
                if self._categorize_them_moi_detail(text) == 'cancel':
                    continue
                filtered.append(tc)
            fields, required_fields, autogen_fields, dup_format_fields = self._them_moi_field_groups()
            if self._crud_is_list_only('create'):
                modules[module_name] = self._them_moi_fallback_two_tc(filtered)
                continue
            if not fields:
                modules[module_name] = self._them_moi_fallback_two_tc(filtered)
                continue
            buckets: dict[str, list] = {
                'positive': [], 'negative_required': [], 'negative_format': [], 'negative_security': [],
            }
            for tc in filtered:
                text = ' '.join(str(tc.get(k) or '') for k in ('scenario', 'description', 'title'))
                cat = self._categorize_them_moi_detail(text)
                if cat == 'add_more':
                    cat = 'positive'
                if cat not in buckets:
                    cat = 'negative_other'
                bucket_key = cat if cat in buckets else 'negative_format'
                buckets.setdefault(bucket_key, []).append(tc)
            all_names = self._vn_join_names([(f.get('name') or '').strip() for f in fields])
            required_names = self._vn_join_names([(f.get('name') or '').strip() for f in required_fields])
            dup_names = self._vn_join_names([(f.get('name') or '').strip() for f in dup_format_fields])
            result: list = []
            # ── TC1: thành công ──
            tc1 = dict(buckets['positive'][0]) if buckets['positive'] else {}
            tc1['scenario'] = tc1['description'] = f"Thêm mới thành công khi nhập {all_names} hợp lệ."
            tc1.setdefault('title', "Thêm mới thành công")
            tc1['steps'] = f"1. Mở form Thêm mới\n2. Nhập {all_names} hợp lệ\n3. Nhấn nút \"Thêm mới\""
            tc1['test_data'] = "; ".join(f"{(f.get('name') or '').strip()}: giá trị hợp lệ" for f in fields)
            tc1['expected_result'] = tc1['then'] = (
                "Xuất hiện thông báo \"Thêm thành công\". Dữ liệu được lưu và hiển thị đúng tại danh sách."
            )
            result.append(tc1)
            if autogen_fields:
                fname = (autogen_fields[0].get('name') or '').strip()
                tc2: dict = {}
                tc2['scenario'] = tc2['description'] = (
                    f"Để trống {fname}, hệ thống tự sinh mã hợp lệ, không trùng và lưu thành công."
                )
                tc2.setdefault('title', "Thêm mới - tự sinh mã")
                tc2['steps'] = f"1. Mở form Thêm mới\n2. Để trống {fname}, nhập hợp lệ các trường còn lại\n3. Nhấn nút \"Thêm mới\""
                tc2['test_data'] = f"{fname}: (để trống); các trường còn lại: giá trị hợp lệ"
                tc2['expected_result'] = tc2['then'] = (
                    f"Hệ thống tự sinh {fname} hợp lệ, không trùng dữ liệu hiện có. "
                    "Xuất hiện thông báo \"Thêm thành công\", dữ liệu được lưu."
                )
                result.append(tc2)
            if required_fields:
                tc3 = dict(buckets['negative_required'][0]) if buckets['negative_required'] else {}
                tc3['scenario'] = tc3['description'] = (
                    f"Bỏ trống {required_names} hoặc chỉ nhập khoảng trắng, hiển thị lỗi bắt buộc và không lưu."
                )
                tc3.setdefault('title', "Thêm mới - thiếu trường bắt buộc")
                tc3['steps'] = f"1. Mở form Thêm mới\n2. Bỏ trống hoặc chỉ nhập khoảng trắng vào {required_names}\n3. Nhấn nút \"Thêm mới\""
                tc3['test_data'] = f"{required_names}: (để trống hoặc chỉ nhập khoảng trắng)"
                tc3['expected_result'] = tc3['then'] = (
                    "Xuất hiện thông báo lỗi bắt buộc nhập tương ứng từng trường, dữ liệu không được lưu."
                )
                result.append(tc3)
            if dup_format_fields:
                tc4 = dict(buckets['negative_format'][0]) if buckets['negative_format'] else {}
                tc4['scenario'] = tc4['description'] = (
                    f"{dup_names} trùng, sai định dạng hoặc vượt độ dài, hiển thị lỗi phù hợp và không lưu."
                )
                tc4.setdefault('title', "Thêm mới - trùng/sai định dạng/vượt độ dài")
                tc4['steps'] = (
                    f"1. Mở form Thêm mới\n2. Nhập {dup_names} trùng dữ liệu đã có, sai định dạng hoặc vượt độ dài cho phép\n"
                    "3. Nhấn nút \"Thêm mới\""
                )
                tc4['test_data'] = f"{dup_names}: giá trị trùng / sai định dạng / vượt độ dài cho phép"
                tc4['expected_result'] = tc4['then'] = (
                    "Xuất hiện thông báo lỗi phù hợp theo từng trường hợp, dữ liệu không được lưu."
                )
                result.append(tc4)
            tc5 = dict(buckets['negative_security'][0]) if buckets['negative_security'] else {}
            tc5['scenario'] = tc5['description'] = (
                "Nhập XSS hoặc SQL Injection, hệ thống không thực thi mã và không lưu dữ liệu nguy hiểm."
            )
            tc5.setdefault('title', "Thêm mới - XSS/SQL Injection")
            tc5['steps'] = f"1. Mở form Thêm mới\n2. Nhập mã XSS hoặc SQL Injection vào {all_names}\n3. Nhấn nút \"Thêm mới\""
            tc5['test_data'] = f"{all_names}: <script>alert(1)</script> hoặc ' OR '1'='1"
            tc5['expected_result'] = tc5['then'] = (
                "Hệ thống không thực thi mã độc, từ chối hoặc escape an toàn; xuất hiện thông báo "
                "lỗi phù hợp (nếu có) và không lưu dữ liệu nguy hiểm."
            )
            tc5.setdefault('test_type', 'Kiểm thử bảo mật')
            result.append(tc5)
            modules[module_name] = result
        data['modules'] = modules
        return data
    def _enforce_cap_nhat_tim_kiem_exact_two(self, data: dict) -> dict:
        """Chỉ chuẩn hóa Tìm kiếm/Tìm về đúng 2 TC.
        Cập nhật KHÔNG còn bị ép 2 TC vì form Cập nhật chi tiết cần 7–9 TC
        theo nhóm nghiệp vụ. Việc gọn hóa CRUD được xử lý bởi
        _enforce_crud_compact_ranges().
        """
        modules = data.get('modules', {})
        search_pattern = re.compile(r'^(tìm kiếm( theo .+)?|tìm( theo .+)?|search( by .+)?)$', re.I)
        def is_negative(tc: dict) -> bool:
            text = ' '.join(str(tc.get(k) or '') for k in (
                'scenario', 'description', 'expected_result', 'then', 'title'
            )).lower()
            return any(k in text for k in (
                'không thành công', 'thất bại', 'không hợp lệ', 'lỗi',
                'không có kết quả', 'không tìm thấy', 'không có dữ liệu',
                'không lưu', 'rỗng', 'không phù hợp',
            ))
        for module_name, tcs in list(modules.items()):
            if not isinstance(tcs, list) or not search_pattern.match(module_name.strip().lower()):
                continue
            valid = [tc for tc in tcs if isinstance(tc, dict) and (
                tc.get('scenario') or tc.get('description') or tc.get('title')
            )]
            positives = [tc for tc in valid if not is_negative(tc)]
            negatives = [tc for tc in valid if is_negative(tc)]
            final = []
            if positives:
                final.append(positives[0])
            if negatives:
                keep = negatives[0]
                extras = []
                for extra in negatives[1:]:
                    value = (extra.get('scenario') or extra.get('description') or '').strip()
                    if value and value not in (keep.get('scenario') or ''):
                        extras.append(value)
                if extras:
                    base = (keep.get('scenario') or keep.get('description') or '').rstrip()
                    keep['scenario'] = base + ('\n' if base else '') + '\n'.join(f'- {x}' for x in extras)
                final.append(keep)
            modules[module_name] = final
        data['modules'] = modules
        return data
    def _enforce_crud_compact_ranges(self, data: dict) -> dict:
        """Giữ CRUD gọn theo nhóm nghiệp vụ, không cap về 2 TC.
        - Thêm mới: tối đa 8 TC.
        - Cập nhật: tối đa 9 TC.
        - Xóa: tối đa 6 TC.
        Hàm chỉ bỏ dòng rỗng và testcase trùng gần như hoàn toàn; KHÔNG tự
        xóa nhóm nghiệp vụ khác nhau. Số tối thiểu được Coverage Checker và
        _ensure_final_testcase_counts bảo đảm ở bước sau.
        """
        modules = data.get('modules', {}) if isinstance(data, dict) else {}

        def action_of(name: str) -> str:
            n = re.sub(r'\s+', ' ', (name or '').strip().lower())
            if self._is_them_moi_module(name) and 'tiếp tục' not in n:
                return 'create'
            if n in {'cập nhật', 'nút cập nhật', 'icon cập nhật', 'chỉnh sửa', 'sửa', 'edit', 'update'} or n.startswith(('cập nhật ', 'chỉnh sửa ')):
                return 'update'
            if n in {'xóa', 'xoá', 'icon xóa', 'nút xóa', 'button xóa', 'delete'} or n.startswith(('xóa ', 'xoá ')):
                return 'delete'
            return ''

        max_by_action = {'create': 8, 'update': 9, 'delete': 6}
        for module_name, tcs in list(modules.items()):
            action = action_of(module_name)
            if not action or not isinstance(tcs, list):
                continue
            valid = [tc for tc in tcs if isinstance(tc, dict) and (
                tc.get('scenario') or tc.get('description') or tc.get('title')
            )]
            seen = set()
            compact = []
            for tc in valid:
                raw = ' | '.join(str(tc.get(k) or '') for k in ('scenario', 'description', 'expected_result'))
                key = re.sub(r'[^a-z0-9à-ỹ]+', ' ', raw.lower()).strip()
                if key and key in seen:
                    continue
                if key:
                    seen.add(key)
                compact.append(tc)
                max_allowed = 2 if self._crud_is_list_only(action) else max_by_action[action]
                if len(compact) >= max_allowed:
                    break
            modules[module_name] = compact
        data['modules'] = modules
        return data
    def _enforce_xoa_exact_two(self, data: dict) -> dict:
        """Tương thích tên hàm cũ; Xóa không còn bị ép đúng 2 TC.
        Rule mới cần 5–6 TC theo nhóm nghiệp vụ. Việc làm gọn và giới hạn tối
        đa đã được _enforce_crud_compact_ranges xử lý.
        """
        return data
    def _enforce_generic_module_exact_two(self, data: dict) -> dict:
        """
        Chuẩn hoá CỨNG MỌI chức năng (trừ "Quay lại" và "Phân trang", đã có
        hàm enforce riêng 4 TC) về ĐÚNG 2 TC: Thành công / Không thành công.
        Trước đây chỉ có enforce riêng cho "Thêm mới"/"Cập nhật"/"Tìm kiếm"/
        "Xóa" — các chức năng nghiệp vụ khác (vd "Nhập điểm", "Quản lý bệnh
        nhân", "Chuyển khoản"...) không được cap, nên nếu model lỡ tách
        nhiều biến thể lỗi thành nhiều TC riêng (vd "Nhập điểm" ra 3 TC:
        thành công + 2 loại không hợp lệ) thì vẫn bị giữ nguyên dư TC.
        Hàm này chạy SAU CÙNG, sau khi các enforce riêng đã xử lý các
        chức năng đặc thù, nên các chức năng đó (đã ≤ 2 TC) sẽ không bị ảnh
        hưởng gì thêm — an toàn để áp dụng chung cho toàn bộ modules còn lại.
        """
        modules = data.get('modules', {})
        def _categorize(tc: dict) -> str:
            text = ' '.join(
                str(tc.get(k) or '') for k in ('scenario', 'description', 'expected_result', 'then', 'title')
            ).lower()
            negative_kw = (
                'không thành công', 'thất bại', 'không hợp lệ', 'lỗi',
                'không có kết quả', 'không tìm thấy', 'không có dữ liệu',
                'không lưu', 'rỗng', 'không phù hợp', 'không đủ', 'từ chối',
            )
            if any(k in text for k in negative_kw):
                return 'negative'
            return 'positive'
        for module_name, tcs in list(modules.items()):
            if not isinstance(tcs, list) or not tcs:
                continue
            lower = module_name.lower()
            if 'quay lại' in lower or 'phân trang' in lower:
                continue
            exact_two = (
                lower == 'tìm'
                or lower.startswith('tìm kiếm')
                or lower.startswith('cập nhật')
                or lower.startswith('chỉnh sửa')
            )
            if not exact_two:
                continue
            if self._is_popup_action_module_4tc(lower):
                continue
            if self._is_them_moi_module(module_name):
                continue
            valid_tcs = [
                tc for tc in tcs
                if isinstance(tc, dict) and (tc.get('scenario') or tc.get('description') or tc.get('title'))
            ]
            if not valid_tcs or len(valid_tcs) <= 2:
                modules[module_name] = valid_tcs
                continue
            positives = [tc for tc in valid_tcs if _categorize(tc) == 'positive']
            negatives = [tc for tc in valid_tcs if _categorize(tc) == 'negative']
            final: list = []
            if positives:
                final.append(positives[0])
            if negatives:
                keep = negatives[0]
                if len(negatives) > 1:
                    extra_desc = []
                    for extra in negatives[1:]:
                        d = (extra.get('scenario') or extra.get('description') or '').strip()
                        if d and d not in (keep.get('scenario') or ''):
                            extra_desc.append(d)
                    if extra_desc:
                        base = (keep.get('scenario') or keep.get('description') or '').rstrip()
                        merged = base + ('\n' if base else '') + '\n'.join(f"- {d}" for d in extra_desc)
                        keep['scenario'] = merged
                        if keep.get('description'):
                            keep['description'] = merged
                final.append(keep)
            if not final:
                final = valid_tcs[:2]
            modules[module_name] = final
        data['modules'] = modules
        return data
    def _clean_contradictory_lines(self, data: dict) -> dict:
        """
        Loại bỏ các DÒNG trong scenario/description bị "lạc" — mâu thuẫn với
        chính loại TC (thành công/không thành công) mà nó đang thuộc về.
        Ví dụ thực tế: TC "Tìm kiếm — không thành công" nhưng scenario lại
        bị lẫn 1 dòng phụ "Tìm kiếm với từ khóa hợp lệ" (đây là câu mô tả
        case THÀNH CÔNG, không thuộc phạm vi TC không thành công này) — hậu
        quả của việc AI/gộp nhiều biến thể mô tả nhưng giữ sót 1 dòng thuộc
        case ngược lại.
        Heuristic: với TC được phân loại 'negative' (dựa theo toàn bộ nội
        dung TC), 1 dòng trong scenario/description bị coi là "lạc" nếu dòng
        đó chứa 'hợp lệ' nhưng KHÔNG chứa bất kỳ từ phủ định/lỗi nào (như
        'không', 'sai', 'thiếu', 'trống', 'lỗi'...). Chỉ xoá khi field còn
        lại ÍT NHẤT 1 dòng khác sau khi loại, tránh làm rỗng field.
        """
        modules = data.get('modules', {})
        negative_kw = (
            'không thành công', 'thất bại', 'không hợp lệ', 'lỗi',
            'không có kết quả', 'không tìm thấy', 'không có dữ liệu',
            'không lưu', 'rỗng', 'không phù hợp', 'không đủ', 'từ chối',
        )
        line_negative_markers = (
            'không', 'sai', 'thiếu', 'trống', 'lỗi', 'vượt quá',
            'không tồn tại', 'ký tự đặc biệt',
        )
        for tcs in modules.values():
            if not isinstance(tcs, list):
                continue
            for tc in tcs:
                if not isinstance(tc, dict):
                    continue
                full_text = ' '.join(
                    str(tc.get(k) or '') for k in
                    ('scenario', 'description', 'expected_result', 'then', 'title')
                ).lower()
                tc_category = 'negative' if any(k in full_text for k in negative_kw) else 'positive'
                if tc_category != 'negative':
                    continue  
                for field in ('scenario', 'description'):
                    text = tc.get(field)
                    if not isinstance(text, str) or '\n' not in text:
                        continue
                    lines = text.split('\n')
                    if len(lines) <= 1:
                        continue
                    kept_lines = []
                    for line in lines:
                        line_lower = line.lower()
                        has_positive_marker = 'hợp lệ' in line_lower
                        has_negative_marker = any(k in line_lower for k in line_negative_markers)
                        if has_positive_marker and not has_negative_marker:
                            continue
                        kept_lines.append(line)
                    if kept_lines and len(kept_lines) != len(lines):
                        tc[field] = '\n'.join(kept_lines)
        data['modules'] = modules
        return data
    def _enforce_quay_lai_exact_four(self, data: dict) -> dict:
        """Chuẩn hoá CỨNG chức năng 'Quay lại' về ĐÚNG 4 TC — 1 TC / mỗi kịch
        bản bắt buộc (chưa thay đổi / đã thay đổi / từ màn hình chi tiết-sửa /
        không có màn hình trước) — đối phó case thực tế đã gặp: AI sinh 2 TC
        cùng thuộc kịch bản "chưa thay đổi dữ liệu" (diễn đạt khác nhau,
        không bị dedupe), khiến sheet "Quay lại" dư ra 5 TC thay vì 4.
        Khác với 2 hàm cap ở trên (chỉ có 2 nhóm positive/negative), chức năng
        này có 4 NHÓM riêng biệt — check theo thứ tự ưu tiên từ đặc trưng
        nhất đến chung nhất để tránh nhận nhầm nhóm (case "chưa thay đổi" là
        mặc định/fallback vì mô tả của nó ít từ khóa đặc trưng nhất).
        """
        modules = data.get('modules', {})
        def _is_back_module(name_lower: str) -> bool:
            """
            BUG ĐÃ SỬA: trước đây dùng exact-match với 4 chuỗi cứng
            ('quay lại', 'nút quay lại', 'button quay lại', 'back'), nên
            bất kỳ biến thể tên nào AI sinh ra (thừa khoảng trắng, kèm
            icon còn sót, viết hoa khác chuẩn, hoặc chỉ cần khác 1 ký tự)
            đều KHÔNG khớp — khiến hàm bỏ qua hoàn toàn việc ép về 4 TC,
            và vì _ensure_final_testcase_counts chạy sau đó chỉ CỘNG THÊM
            TC khi thiếu chứ không bao giờ cắt bớt khi dư, nên số TC AI
            sinh ra (có thể là 6) lọt thẳng ra ngoài không bị cap.
            Sửa thành substring-match, NHẤT QUÁN với cách
            _validate_testcase_count (dòng ~3600) đang nhận diện module
            này (`'quay lại' in lower`) — để 2 hàm không còn lệch tiêu chí.
            """
            n = name_lower.strip()
            return 'quay lại' in n or 'quay lai' in n or n == 'back'
        def _categorize(tc: dict) -> str:
            text = ' '.join(
                str(tc.get(k) or '') for k in ('scenario', 'description', 'expected_result', 'then', 'title')
            ).lower()
            if any(k in text for k in ('chi tiết', 'màn hình sửa', 'từ màn hình')):
                return 'detail'
            if any(k in text for k in (
                'không có màn hình trước', 'truy cập trực tiếp', 'trang mặc định', 'không hành động',
            )):
                return 'no_prev'
            if any(k in text for k in ('đã thay đổi', 'đã đổi', 'đã nhập', 'đã sửa', 'thông báo thành công')):
                return 'changed'
            return 'unchanged'
        order = ['unchanged', 'changed', 'detail', 'no_prev']
        for module_name, tcs in list(modules.items()):
            if not isinstance(tcs, list) or not tcs:
                continue
            if not _is_back_module(module_name.lower()):
                continue
            valid_tcs = [
                tc for tc in tcs
                if isinstance(tc, dict) and (tc.get('scenario') or tc.get('description') or tc.get('title'))
            ]
            if not valid_tcs:
                continue
            buckets: dict[str, list] = {k: [] for k in order}
            for tc in valid_tcs:
                buckets[_categorize(tc)].append(tc)
            final: list = []
            for key in order:
                group = buckets[key]
                if not group:
                    continue
                keep = group[0]
                if len(group) > 1:
                    extra_desc = []
                    for extra in group[1:]:
                        d = (extra.get('scenario') or extra.get('description') or '').strip()
                        if d and d not in (keep.get('scenario') or ''):
                            extra_desc.append(d)
                    if extra_desc:
                        base = (keep.get('scenario') or keep.get('description') or '').rstrip()
                        merged = base + ('\n' if base else '') + '\n'.join(f"- {d}" for d in extra_desc)
                        keep['scenario'] = merged
                        if keep.get('description'):
                            keep['description'] = merged
                final.append(keep)
            modules[module_name] = final
        data['modules'] = modules
        return data
    def _enforce_list_screen_action_only(self, data: dict) -> dict:
        """Chốt kết quả cuối cho màn hình danh sách.
        Khi chưa thấy form/popup:
        - Thêm mới: chỉ mở form thành công/không thành công.
        - Cập nhật: chỉ mở đúng form bản ghi/không mở được.
        - Xóa: chỉ mở popup xác nhận/không mở được.
        - Quay lại: chỉ điều hướng thành công/không có lịch sử trước.
        Hàm thay thế hoàn toàn nội dung AI cũ để không còn sót validation,
        nhập liệu, lưu dữ liệu hoặc XSS của form chưa được cung cấp.
        """
        if not self._crud_is_list_only():
            return data
        modules = data.get("modules", {})
        if not isinstance(modules, dict):
            return data
        def base_tc(module_name: str, title: str, scenario: str, expected: str,
                    test_type: str, given: str, test_data: str = "Không áp dụng") -> dict:
            return {
                "id": None,
                "chức năng": module_name,
                "feature": module_name,
                "title": title,
                "scenario": scenario,
                "description": scenario,
                "given": given,
                "when": scenario,
                "then": expected,
                "precondition": "Màn hình danh sách hiển thị bình thường",
                "steps": f"1. Mở màn hình danh sách\n2. {scenario}",
                "test_data": test_data,
                "expected_result": expected,
                "priority": "Cao",
                "test_type": test_type,
                "actual_result": "",
                "status": "Chưa chạy",
                "note": "",
            }
        for module_name in list(modules.keys()):
            lower = re.sub(r"\s+", " ", str(module_name).strip().lower())
            if self._is_them_moi_module(module_name) and "tiếp tục" not in lower:
                modules[module_name] = [
                    base_tc(
                        module_name,
                        "Mở form thêm mới thành công",
                        "Nhấn nút Thêm mới",
                        "Hệ thống mở đúng form thêm mới; chưa thực hiện lưu dữ liệu",
                        "Kiểm thử dương",
                        "Người dùng có quyền thêm mới",
                    ),
                    base_tc(
                        module_name,
                        "Mở form thêm mới không thành công",
                        "Nhấn nút Thêm mới khi không có quyền hoặc form không thể tải",
                        "Hệ thống không mở form, hiển thị thông báo phù hợp và giữ nguyên danh sách",
                        "Kiểm thử âm",
                        "Người dùng không có quyền hoặc hệ thống gặp lỗi tải form",
                    ),
                ]
                continue
            if lower.startswith(("cập nhật", "chỉnh sửa")) or lower in {
                "sửa", "update", "edit", "icon cập nhật", "nút cập nhật"
            }:
                modules[module_name] = [
                    base_tc(
                        module_name,
                        "Mở form cập nhật thành công",
                        "Nhấn biểu tượng Cập nhật tại một dòng dữ liệu",
                        "Hệ thống mở đúng form và hiển thị đúng dữ liệu của bản ghi được chọn",
                        "Kiểm thử dương",
                        "Người dùng có quyền cập nhật và bản ghi tồn tại",
                        "Một bản ghi hợp lệ trong danh sách",
                    ),
                    base_tc(
                        module_name,
                        "Mở form cập nhật không thành công",
                        "Nhấn Cập nhật khi không có quyền, bản ghi không tồn tại hoặc form không thể tải",
                        "Hệ thống không mở form, hiển thị thông báo phù hợp và không thay đổi dữ liệu",
                        "Kiểm thử âm",
                        "Người dùng không có quyền, bản ghi không tồn tại hoặc hệ thống gặp lỗi",
                        "Bản ghi không tồn tại hoặc tài khoản không có quyền",
                    ),
                ]
                continue
            if lower.startswith(("xóa", "xoá")) or lower in {
                "delete", "icon xóa", "icon xoá", "nút xóa", "nút xoá"
            }:
                modules[module_name] = [
                    base_tc(
                        module_name,
                        "Mở popup xác nhận xóa thành công",
                        "Nhấn biểu tượng Xóa tại một dòng dữ liệu",
                        "Hệ thống mở popup xác nhận và hiển thị đúng đối tượng được chọn",
                        "Kiểm thử dương",
                        "Người dùng có quyền xóa và bản ghi tồn tại",
                        "Một bản ghi hợp lệ trong danh sách",
                    ),
                    base_tc(
                        module_name,
                        "Mở popup xác nhận xóa không thành công",
                        "Nhấn Xóa khi không có quyền, bản ghi không tồn tại hoặc popup không thể tải",
                        "Hệ thống không mở popup, hiển thị thông báo phù hợp và giữ nguyên dữ liệu",
                        "Kiểm thử âm",
                        "Người dùng không có quyền, bản ghi không tồn tại hoặc hệ thống gặp lỗi",
                        "Bản ghi không tồn tại hoặc tài khoản không có quyền",
                    ),
                ]
                continue
        data["modules"] = modules
        return data
    def _normalize_test_cases(
        self, data: dict, apply_static_filter: bool = True, description: str | None = None,
    ) -> dict:
        if apply_static_filter:
            data = self._drop_static_data_modules(data)
        data = self._dedupe_similar_modules(data)
        if self._crud_is_list_only():
            data = self._enforce_list_screen_action_only(data)
        else:
            data = self._enforce_them_moi_format(data)
            data = self._enforce_crud_compact_ranges(data)
        data = self._enforce_cap_nhat_tim_kiem_exact_two(data)
        data = self._enforce_xoa_exact_two(data)
        print("[DEBUG-QUAYLAI] module keys:", repr(list(data.get('modules', {}).keys())))
        data = self._enforce_quay_lai_exact_four(data)
        data = self._enforce_generic_module_exact_two(data)
        cleaner = getattr(self, '_clean_contradictory_lines', None)
        if callable(cleaner):
            data = cleaner(data)
        data = self._ensure_final_testcase_counts(data)
        data = self._enforce_list_screen_action_only(data)
        valid_priorities = {'Cao', 'Trung bình', 'Thấp'}
        valid_types = {
            'Kiểm thử chức năng', 'Kiểm thử giao diện', 'Kiểm thử xác thực',
            'Kiểm thử bảo mật', 'Kiểm thử phân quyền', 'Kiểm thử âm',
            'Kiểm thử dương', 'Kiểm thử biên', 'Kiểm thử tích hợp'
        }
        test_type_aliases = {
            'chức năng': 'Kiểm thử chức năng',
            'kiem thu chuc nang': 'Kiểm thử chức năng',
            'giao diện': 'Kiểm thử giao diện',
            'kiem thu giao dien': 'Kiểm thử giao diện',
            'xác thực': 'Kiểm thử xác thực',
            'validation': 'Kiểm thử xác thực',
            'kiem thu xac thuc': 'Kiểm thử xác thực',
            'bảo mật': 'Kiểm thử bảo mật',
            'security': 'Kiểm thử bảo mật',
            'kiem thu bao mat': 'Kiểm thử bảo mật',
            'phân quyền': 'Kiểm thử phân quyền',
            'authorization': 'Kiểm thử phân quyền',
            'kiem thu phan quyen': 'Kiểm thử phân quyền',
            'âm': 'Kiểm thử âm',
            'negative': 'Kiểm thử âm',
            'kiem thu am': 'Kiểm thử âm',
            'dương': 'Kiểm thử dương',
            'positive': 'Kiểm thử dương',
            'kiem thu duong': 'Kiểm thử dương',
            'biên': 'Kiểm thử biên',
            'boundary': 'Kiểm thử biên',
            'kiem thu bien': 'Kiểm thử biên',
            'tích hợp': 'Kiểm thử tích hợp',
            'integration': 'Kiểm thử tích hợp',
            'kiem thu tich hop': 'Kiểm thử tích hợp',
        }
        def _ascii_lower(value: object) -> str:
            import unicodedata
            text = str(value or '').strip().lower()
            text = unicodedata.normalize('NFD', text)
            text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
            return re.sub(r'\s+', ' ', text)
        def _infer_test_type(tc: dict) -> str:
            """Suy ra loại test từ nội dung khi model trả thiếu hoặc quá chung."""
            current = str(tc.get('test_type') or '').strip()
            if current in valid_types and current != 'Kiểm thử chức năng':
                return current
            alias_key = _ascii_lower(current)
            if alias_key in test_type_aliases:
                mapped = test_type_aliases[alias_key]
                if mapped != 'Kiểm thử chức năng':
                    return mapped
            text = _ascii_lower(' | '.join(str(tc.get(k) or '') for k in (
                'title', 'feature', 'scenario', 'description', 'test_data',
                'expected_result', 'precondition', 'steps'
            )))
            if any(x in text for x in (
                'xss', 'sql injection', 'sqli', 'script', 'ma doc',
                'tan cong', 'bao mat', 'csrf', 'session', 'token'
            )):
                return 'Kiểm thử bảo mật'
            if any(x in text for x in (
                'khong co quyen', 'khong du quyen', 'phan quyen', 'vai tro',
                'role ', 'admin', 'user khong duoc phep', 'truy cap trai phep'
            )):
                return 'Kiểm thử phân quyền'
            if any(x in text for x in (
                'bat buoc', 'bo trong', 'de trong', 'khoang trang',
                'sai dinh dang', 'khong hop le', 'validation', 'xac thuc',
                'email sai', 'so dien thoai sai'
            )):
                return 'Kiểm thử xác thực'
            if any(x in text for x in (
                'toi da', 'toi thieu', 'gioi han', 'vuot do dai', 'do dai',
                'gia tri bien', 'bien tren', 'bien duoi', '0 ky tu',
                'ky tu cuoi', 'trang dau', 'trang cuoi'
            )):
                return 'Kiểm thử biên'
            if any(x in text for x in (
                'mat mang', 'loi server', 'api', 'database', 'dong bo',
                'tich hop', 'nhieu lan lien tiep', 'gui lap', 'double click',
                'tai lai', 'ket noi dich vu'
            )):
                return 'Kiểm thử tích hợp'
            if any(x in text for x in (
                'mau sac', 'can le', 'font', 'placeholder', 'icon', 'hien thi',
                'responsive', 'giao dien', 'kich thuoc', 'bo cuc'
            )):
                return 'Kiểm thử giao diện'
            if any(x in text for x in (
                'khong thanh cong', 'that bai', 'khong ton tai', 'tu choi',
                'loi ', 'khong mo duoc', 'khong tim thay', 'huy ', 'dong popup'
            )):
                return 'Kiểm thử âm'
            if any(x in text for x in (
                'thanh cong', 'hop le', 'mo dung', 'luu dung', 'hien thi dung'
            )):
                return 'Kiểm thử dương'
            return 'Kiểm thử chức năng'
        all_existing_ids: set[str] = set()
        for tcs in data.get('modules', {}).values():
            if isinstance(tcs, list):
                for tc in tcs:
                    if isinstance(tc, dict) and tc.get('id'):
                        all_existing_ids.add(tc['id'])
        auto_counter = [1] 
        def _next_unique_id() -> str:
            while True:
                candidate = f"TC_{auto_counter[0]:03d}"
                auto_counter[0] += 1
                if candidate not in all_existing_ids:
                    all_existing_ids.add(candidate)
                    return candidate
        for module_name, test_cases in data.get('modules', {}).items():
            if not isinstance(test_cases, list):
                continue
            for tc in test_cases:
                if not isinstance(tc, dict):
                    continue
                if not tc.get('id'):
                    tc['id'] = _next_unique_id()
                tc['chức năng'] = tc.get('chức năng') or module_name
                feature_name = (
                    tc.get('feature')
                    or tc.get('module')
                    or tc.get('chức năng')
                    or module_name
                )
                tc['feature'] = feature_name
                tc['module'] = tc.get('module') or module_name
                tc.setdefault('title', tc.get('feature') or '')
                tc.setdefault('description', tc.get('scenario') or '')
                tc.setdefault('scenario', '')
                tc.setdefault('given', '')
                tc.setdefault('when', '')
                tc.setdefault('then', '')
                tc.setdefault('precondition', '')
                if not tc.get('steps'):
                    parts = []
                    if tc.get('given'):
                        parts.append(f"1. {tc['given']}")
                    if tc.get('when'):
                        parts.append(f"2. {tc['when']}")
                    if tc.get('then'):
                        parts.append(f"3. {tc['then']}")
                    tc['steps'] = '\n'.join(parts)
                tc.setdefault('test_data', '')
                tc.setdefault('expected_result', tc.get('then') or '')
                tc.setdefault('actual_result', '')
                tc.setdefault('note', '')
                tc.setdefault('status', 'Chưa chạy')
                if tc.get('priority') not in valid_priorities:
                    tc['priority'] = 'Trung bình'
                tc['test_type'] = _infer_test_type(tc)
                tc['status'] = 'Chưa chạy'
        data['modules'] = self._realign_generic_action_module_names(
            description, data.get('modules', {}) or {}
        )
        data = self._finalize_success_failure_grouping(data)
        if description and self._is_targeted_request(description):
            data = self._apply_known_canonical_templates(data, description)
        if _sre_replace_generated_cases_with_template is not None:
            _sre_modules_before = data.get('modules', {}) or {}
            _sre_names_before = list(_sre_modules_before.keys())
            _sre_tc_count_before = sum(
                len(v) for v in _sre_modules_before.values() if isinstance(v, list)
            )
            print(f"[ScenarioRuleEngine] Số module đầu vào: {len(_sre_names_before)} | Tổng TC trước: {_sre_tc_count_before}")
            try:
                for _sre_name in _sre_names_before:
                    _sre_canonical = _sre_normalize_function_name(_sre_name) if _sre_normalize_function_name else None
                    if _sre_canonical is not None and _sre_canonical in _SRE_DEFAULT_ENFORCED_CANONICALS:
                        print(f"[ScenarioRuleEngine] Module '{_sre_name}' -> canonical '{_sre_canonical}' (sẽ bị thay bằng fixed template)")
                    elif _sre_canonical is not None:
                        print(f"[ScenarioRuleEngine] Module '{_sre_name}' -> canonical '{_sre_canonical}' nhưng KHÔNG trong tập enforce (field-aware, giữ TC AI sinh)")
            except Exception:
                pass
            try:
                data['modules'] = _sre_replace_generated_cases_with_template(
                    _sre_modules_before,
                    enforced_canonicals=_SRE_DEFAULT_ENFORCED_CANONICALS,
                )
                _sre_modules_after = data.get('modules', {}) or {}
                _sre_names_after = list(_sre_modules_after.keys())
                _sre_tc_count_after = sum(
                    len(v) for v in _sre_modules_after.values() if isinstance(v, list)
                )
                print(
                    f"[ScenarioRuleEngine] Số module sau: {len(_sre_names_after)} | "
                    f"Tổng TC sau: {_sre_tc_count_after} (trước: {_sre_tc_count_before})"
                )
            except Exception as _exc_sre_apply:
                print(f"[ScenarioRuleEngine] Lỗi khi áp fixed template, giữ nguyên modules hiện có: {_exc_sre_apply}")
        seq = 1
        for test_cases in data.get('modules', {}).values():
            if not isinstance(test_cases, list):
                continue
            for tc in test_cases:
                if not isinstance(tc, dict):
                    continue
                tc['id'] = f"TC_{seq:03d}"
                seq += 1
        # KHÔNG collapse data['modules'] về base_name và KHÔNG ghi đè
        # tc['module']/tc['chức năng']/tc['feature'] bằng base_name ở đây.
        # _finalize_success_failure_grouping() ở trên đã tạo đúng 2 nhóm
        # "<gốc> thành công" / "<gốc> không thành công" và đã gán đúng
        # tc['module']/tc['chức năng']/tc['feature'] CÓ hậu tố — bước này
        # trước đây gọi lại _determine_base_business_function() để CẮT hậu
        # tố rồi ghi đè lên chính testcase + gộp luôn data['modules'], khiến
        # "Đăng nhập thành công" và "Đăng nhập không thành công" cùng bị rút
        # gọn còn "Đăng nhập" ngay tại tầng backend (trước khi trả JSON cho
        # app.py/app.js), nên dù frontend đã sửa đúng vẫn vô nghĩa.
        # base_name (qua _determine_base_business_function) CHỈ được dùng ở
        # nơi cần GOM NHÓM/THỐNG KÊ (vd log "Tổng chức năng" bên dưới trong
        # _log_generation_runtime), tuyệt đối không dùng để thay đổi dữ liệu
        # testcase hay cấu trúc data['modules'].
        print("MODULES AFTER NORMALIZE:", list(data.get('modules', {}).keys()))
        return data
    def _ensure_final_testcase_counts(self, data: dict) -> dict:
        """Bảo đảm count cuối cùng SAU mọi bước normalize/dedupe.
        Không gọi AI, không tăng số vòng Coverage. Hàm chỉ thêm các TC còn
        thiếu bằng template cố định và không xoá TC hợp lệ đang có.
        """
        if not isinstance(data, dict):
            return data
        modules = data.get('modules', {})
        if not isinstance(modules, dict):
            return data
        def required_count(name: str, current_len: int) -> int:
            n = re.sub(r'\s+', ' ', (name or '').strip().lower())
            if n == 'tìm' or n.startswith('tìm kiếm'):
                return 2
            if n.startswith('cập nhật') or n.startswith('chỉnh sửa') or n == 'sửa':
                return self._crud_required_count('update', 7)
            if n.startswith('xóa') or n.startswith('xoá') or n == 'delete':
                return self._crud_required_count('delete', 5)
            if 'quay lại' in n:
                return 4
            if 'phân trang' in n:
                return 4
            if self._is_popup_action_module_4tc(n):
                return 4
            if 'thêm mới và tiếp tục' in n:
                return 4
            if self._is_them_moi_module(name) and 'tiếp tục' not in n:
                return self._crud_required_count('create', 6)
            return current_len

        def final_templates(name: str) -> list[dict]:
            n = re.sub(r'\s+', ' ', (name or '').strip().lower())

            if self._is_them_moi_module(name) and 'tiếp tục' not in n and self._crud_is_list_only('create'):
                return [
                    {'title': 'Thêm mới thành công', 'scenario': 'Nhấn nút Thêm mới trên màn hình danh sách', 'expected_result': 'Hệ thống mở đúng form thêm mới để người dùng nhập dữ liệu', 'test_type': 'Kiểm thử dương'},
                    {'title': 'Thêm mới không thành công', 'scenario': 'Nhấn Thêm mới khi không có quyền hoặc hệ thống không thể tải form', 'expected_result': 'Hệ thống không mở form, hiển thị thông báo phù hợp và giữ nguyên danh sách', 'test_type': 'Kiểm thử âm'},
                ]
            if (n.startswith(('cập nhật', 'chỉnh sửa')) or n in {'sửa', 'update', 'edit'}) and self._crud_is_list_only('update'):
                return [
                    {'title': 'Cập nhật thành công', 'scenario': 'Nhấn biểu tượng Cập nhật tại một dòng dữ liệu', 'expected_result': 'Hệ thống mở đúng form và hiển thị đúng dữ liệu của bản ghi được chọn', 'test_type': 'Kiểm thử dương'},
                    {'title': 'Cập nhật không thành công', 'scenario': 'Nhấn Cập nhật khi không có quyền, bản ghi không tồn tại hoặc form không tải được', 'expected_result': 'Hệ thống không mở form, hiển thị thông báo phù hợp và không thay đổi dữ liệu', 'test_type': 'Kiểm thử âm'},
                ]
            if (n.startswith(('xóa', 'xoá')) or n == 'delete') and self._crud_is_list_only('delete'):
                return [
                    {'title': 'Xóa thành công', 'scenario': 'Nhấn biểu tượng Xóa tại một dòng dữ liệu', 'expected_result': 'Hệ thống mở popup xác nhận đúng đối tượng được chọn', 'test_type': 'Kiểm thử dương'},
                    {'title': 'Xóa không thành công', 'scenario': 'Nhấn Xóa khi không có quyền, bản ghi không tồn tại hoặc popup không tải được', 'expected_result': 'Hệ thống không mở popup, hiển thị thông báo phù hợp và giữ nguyên dữ liệu', 'test_type': 'Kiểm thử âm'},
                ]
            if n == 'tìm':
                return [
                    {'title': 'Tìm thành công', 'scenario': 'Nhập từ khóa hợp lệ rồi nhấn nút Tìm', 'expected_result': 'Danh sách hiển thị đúng các bản ghi khớp từ khóa', 'test_type': 'Kiểm thử dương'},
                    {'title': 'Tìm không thành công', 'scenario': 'Để trống, nhập từ khóa không khớp hoặc ký tự không hợp lệ rồi nhấn Tìm', 'expected_result': 'Hệ thống hiển thị toàn bộ dữ liệu hoặc thông báo không có kết quả phù hợp, không phát sinh lỗi', 'test_type': 'Kiểm thử âm'},
                ]
            if n.startswith('tìm kiếm'):
                return [
                    {'title': 'Tìm kiếm thành công', 'scenario': 'Nhập mã hoặc tên kho tồn tại vào ô tìm kiếm', 'expected_result': 'Danh sách chỉ hiển thị các kho khớp điều kiện tìm kiếm', 'test_type': 'Kiểm thử dương'},
                    {'title': 'Tìm kiếm không thành công', 'scenario': 'Nhập từ khóa không tồn tại, để trống hoặc nhập ký tự đặc biệt vào ô tìm kiếm', 'expected_result': 'Danh sách rỗng hoặc hiển thị thông báo phù hợp, hệ thống không lỗi', 'test_type': 'Kiểm thử âm'},
                ]
            if self._is_them_moi_module(name) and 'tiếp tục' not in n:
                return [
                    {'title': 'Thêm mới thành công', 'scenario': 'Nhập dữ liệu hợp lệ rồi nhấn Thêm mới', 'expected_result': 'Hệ thống lưu thành công và hiển thị bản ghi tại danh sách', 'test_type': 'Kiểm thử dương'},
                    {'title': 'Tự sinh mã', 'scenario': 'Để trống trường mã được phép tự sinh rồi lưu', 'expected_result': 'Hệ thống tạo mã hợp lệ, không trùng và lưu thành công', 'test_type': 'Kiểm thử chức năng'},
                    {'title': 'Thiếu trường bắt buộc', 'scenario': 'Bỏ trống hoặc chỉ nhập khoảng trắng vào trường có dấu *', 'expected_result': 'Hiển thị lỗi tại trường bắt buộc và không lưu dữ liệu', 'test_type': 'Kiểm thử xác thực'},
                    {'title': 'Sai định dạng', 'scenario': 'Nhập dữ liệu sai định dạng vào các field có quy tắc định dạng', 'expected_result': 'Hiển thị lỗi đúng field và không lưu dữ liệu', 'test_type': 'Kiểm thử âm'},
                    {'title': 'Trùng hoặc vượt giới hạn', 'scenario': 'Nhập định danh trùng hoặc dữ liệu vượt độ dài/giới hạn cho phép', 'expected_result': 'Hệ thống từ chối dữ liệu và giữ nguyên danh sách', 'test_type': 'Kiểm thử biên'},
                    {'title': 'Bảo mật và gửi lặp', 'scenario': 'Nhập XSS/SQL Injection hoặc nhấn lưu nhiều lần', 'expected_result': 'Không thực thi mã độc và không tạo bản ghi trùng', 'test_type': 'Kiểm thử bảo mật'},
                ]
            if n.startswith(('cập nhật', 'chỉnh sửa')) or n in {'sửa', 'update', 'edit'}:
                return [
                    {'title': 'Mở đúng bản ghi cập nhật', 'scenario': 'Nhấn Cập nhật tại một dòng dữ liệu', 'expected_result': 'Form mở đúng đối tượng và điền sẵn dữ liệu hiện tại', 'test_type': 'Kiểm thử chức năng'},
                    {'title': 'Cập nhật thành công', 'scenario': 'Thay đổi một hoặc nhiều field bằng dữ liệu hợp lệ rồi lưu', 'expected_result': 'Thông tin được cập nhật và hiển thị đúng tại danh sách', 'test_type': 'Kiểm thử dương'},
                    {'title': 'Không thay đổi dữ liệu', 'scenario': 'Mở form và nhấn Cập nhật khi không thay đổi field nào', 'expected_result': 'Hệ thống xử lý phù hợp, không tạo thay đổi dữ liệu ngoài mong đợi', 'test_type': 'Kiểm thử chức năng'},
                    {'title': 'Thiếu trường bắt buộc', 'scenario': 'Xóa giá trị hoặc nhập khoảng trắng vào field có dấu * rồi cập nhật', 'expected_result': 'Hiển thị lỗi đúng field và không cập nhật dữ liệu', 'test_type': 'Kiểm thử xác thực'},
                    {'title': 'Sai định dạng', 'scenario': 'Nhập dữ liệu sai định dạng vào các field có quy tắc', 'expected_result': 'Hiển thị lỗi định dạng và giữ nguyên dữ liệu cũ', 'test_type': 'Kiểm thử âm'},
                    {'title': 'Trùng hoặc vượt giới hạn', 'scenario': 'Nhập định danh trùng hoặc dữ liệu vượt độ dài/giới hạn', 'expected_result': 'Hệ thống từ chối cập nhật và giữ nguyên dữ liệu cũ', 'test_type': 'Kiểm thử biên'},
                    {'title': 'Bảo mật hoặc lỗi hệ thống', 'scenario': 'Nhập XSS/SQL Injection hoặc xảy ra lỗi server khi cập nhật', 'expected_result': 'Không thực thi mã độc, không mất dữ liệu cũ và hiển thị thông báo phù hợp', 'test_type': 'Kiểm thử bảo mật'},
                ]
            if n.startswith(('xóa', 'xoá')) or n == 'delete':
                return [
                    {'title': 'Mở xác nhận xóa', 'scenario': 'Nhấn biểu tượng Xóa tại một dòng dữ liệu', 'expected_result': 'Hệ thống hiển thị hộp thoại xác nhận đúng đối tượng cần xóa', 'test_type': 'Kiểm thử chức năng'},
                    {'title': 'Xóa thành công', 'scenario': 'Xác nhận xóa đối tượng không có dữ liệu tham chiếu', 'expected_result': 'Hệ thống thông báo xóa thành công và bản ghi không còn trong danh sách', 'test_type': 'Kiểm thử dương'},
                    {'title': 'Hủy xác nhận xóa', 'scenario': 'Nhấn Hủy hoặc đóng hộp thoại xác nhận xóa', 'expected_result': 'Hộp thoại đóng và bản ghi vẫn còn nguyên trong danh sách', 'test_type': 'Kiểm thử chức năng'},
                    {'title': 'Xóa không thành công do ràng buộc', 'scenario': 'Xác nhận xóa đối tượng đang có dữ liệu tham chiếu', 'expected_result': 'Hệ thống từ chối xóa, hiển thị thông báo phù hợp và giữ nguyên dữ liệu', 'test_type': 'Kiểm thử âm'},
                    {'title': 'Không quyền hoặc bản ghi không tồn tại', 'scenario': 'Xóa khi không có quyền hoặc bản ghi đã bị xóa bởi người khác', 'expected_result': 'Hệ thống từ chối thao tác, thông báo phù hợp và không ảnh hưởng dữ liệu khác', 'test_type': 'Kiểm thử phân quyền'},
                    {'title': 'Lỗi hệ thống hoặc xác nhận lặp', 'scenario': 'Mất mạng/lỗi server hoặc nhấn xác nhận xóa nhiều lần', 'expected_result': 'Hệ thống không xóa lặp, trạng thái dữ liệu nhất quán và có thông báo phù hợp', 'test_type': 'Kiểm thử tích hợp'},
                ]
            return self._local_coverage_templates(name)
        for module_name, current in list(modules.items()):
            if not isinstance(current, list):
                continue
            required = required_count(module_name, len(current))
            if len(current) >= required:
                continue
            def key(tc: dict) -> str:
                raw = ' | '.join(str(tc.get(k) or '') for k in ('title', 'scenario', 'expected_result'))
                return re.sub(r'[^a-z0-9à-ỹ]+', ' ', raw.lower()).strip()
            existing = {key(tc) for tc in current if isinstance(tc, dict)}
            for template in final_templates(module_name):
                if len(current) >= required:
                    break
                tc = self._build_local_coverage_testcase(module_name, template, '')
                k = key(tc)
                if k in existing:
                    continue
                tc['id'] = None
                current.append(tc)
                existing.add(k)
            if len(current) < required:
                print(
                    f"[CoverageWarning] Module '{module_name}' còn thiếu "
                    f"{required - len(current)} testcase so với yêu cầu tối thiểu "
                    f"({len(current)}/{required}) sau khi đã áp hết template cục bộ — "
                    "KHÔNG fabricate TC chung chung, giữ nguyên số lượng hiện có."
                )
        return data
    def _merge_test_cases(self, old_data: dict, new_data: dict) -> dict:
        import difflib
        def _content_key(tc: dict) -> str:
            """Chuẩn hoá nội dung TC để so khớp trùng lặp, không phụ thuộc id
            (vì các TC bổ sung từ _enforce_min_coverage luôn có id=None)."""
            raw = (tc.get('scenario') or tc.get('description') or tc.get('title') or '')
            return ' '.join(raw.lower().split())
        def _tc_category(text: str) -> str | None:
            """Phân loại TC 'Thêm mới' theo NHÓM NGHIỆP VỤ chi tiết (không
            còn gộp mọi biến thể lỗi vào 1 nhóm 'negative' duy nhất — mỗi
            nhóm lỗi khác nhau (required/format-duplicate/security) phải
            được coi là NHÓM RIÊNG khi dedupe, để không bị merge nhầm 2 TC
            khác ý nghĩa vào làm 1). Bắt các câu diễn đạt khác nhau nhưng
            CÙNG Ý (VD "thêm rồi tiếp tục" vs "thêm tiếp" chỉ ratio văn bản
            0.6, dưới ngưỡng dedup thường, nhưng cùng thuộc category
            'add_more')."""
            return self._categorize_them_moi_detail(text)
        def _is_duplicate(new_tc: dict, existing_list: list, module_lower: str = '', threshold: float = 0.95) -> bool:
            """Chỉ loại TC thật sự trùng nội dung/nghiệp vụ.
            BUG cũ: ngưỡng fuzzy 0.82 quá thấp nên hai testcase khác dữ liệu biên
            hoặc khác trường validation vẫn bị xem là trùng, đặc biệt sau khi gộp
            workflow. Kết quả là chức năng bị tụt dưới số TC tối thiểu.
            Quy tắc mới:
            - Trùng chính xác scenario/description/title => loại.
            - Chức năng "Thêm mới" vẫn được dedupe theo nhóm nghiệp vụ chi tiết.
            - Chức năng khác chỉ fuzzy-dedupe khi cả nội dung chính VÀ expected_result
              gần như giống hệt nhau; không loại chỉ vì câu chữ mở đầu giống nhau.
            """
            new_key = _content_key(new_tc)
            if not new_key:
                return False
            use_category = self._is_them_moi_module(module_lower)
            new_cat = _tc_category(new_key) if use_category else None
            new_expected = ' '.join(str(new_tc.get('expected_result') or new_tc.get('then') or '').lower().split())
            for old_tc in existing_list:
                if not isinstance(old_tc, dict):
                    continue
                old_key = _content_key(old_tc)
                if not old_key:
                    continue
                if new_key == old_key:
                    return True
                if use_category and new_cat and _tc_category(old_key) == new_cat:
                    return True
                ratio = difflib.SequenceMatcher(None, new_key, old_key).ratio()
                if ratio >= threshold:
                    old_expected = ' '.join(str(old_tc.get('expected_result') or old_tc.get('then') or '').lower().split())
                    expected_ratio = difflib.SequenceMatcher(None, new_expected, old_expected).ratio() if (new_expected and old_expected) else 0.0
                    if expected_ratio >= 0.90:
                        return True
            return False
        def _find_similar_key(name: str, existing_keys, threshold: float = 0.65):
            """Tìm chức năng key đã có gần giống tên mới (AI hay đặt tên hơi khác
            nhau giữa các vòng gọi, VD 'Thêm mới danh mục kho' vs 'Thêm danh
            mục kho') để gộp chung 1 bucket thay vì tạo chức năng trùng lặp."""
            name_l = ' '.join(name.lower().split())
            best_key, best_ratio = None, 0.0
            for k in existing_keys:
                k_l = ' '.join(k.lower().split())
                if name_l == k_l:
                    return k
                ratio = difflib.SequenceMatcher(None, name_l, k_l).ratio()
                if ratio > best_ratio:
                    best_ratio, best_key = ratio, k
            return best_key if best_ratio >= threshold else None
        merged_modules: dict[str, list] = {}
        for module_name, test_cases in old_data.get('modules', {}).items():
            merged_modules[module_name] = list(test_cases)
        print(
            "MODULES BEFORE MERGE:",
            list(old_data.get('modules', {}).keys()) + list(new_data.get('modules', {}).keys()),
        )
        for module_name, new_test_cases in new_data.get('modules', {}).items():
            if not isinstance(new_test_cases, list):
                continue
            target_key = module_name
            if module_name not in merged_modules:
                similar = _find_similar_key(module_name, merged_modules.keys())
                if similar:
                    target_key = similar  
            if target_key not in merged_modules:
                target_lower = target_key.lower()
                deduped: list = []
                for tc in new_test_cases:
                    if isinstance(tc, dict) and _is_duplicate(tc, deduped, target_lower):
                        continue
                    deduped.append(tc)
                merged_modules[target_key] = deduped
            else:
                target_lower = target_key.lower()
                existing_ids = {tc.get('id') for tc in merged_modules[target_key] if tc.get('id')}
                for tc in new_test_cases:
                    if not isinstance(tc, dict):
                        continue
                    tc_id = tc.get('id')
                    if tc_id and tc_id in existing_ids:
                        tc = dict(tc)
                        tc['id'] = None
                        tc_id = None
                    if _is_duplicate(tc, merged_modules[target_key], target_lower):
                        continue  
                    merged_modules[target_key].append(tc)
                    if tc_id:
                        existing_ids.add(tc_id)
        print("MODULES AFTER MERGE:", list(merged_modules.keys()))
        return {
            'project_name': new_data.get('project_name') or old_data.get('project_name', ''),
            'description': new_data.get('description') or old_data.get('description', ''),
            'modules': merged_modules,
        }
    _STATIC_MODULE_PATTERNS = [
        'stt', 'số thứ tự', 'số tt',
        'hiển thị', 'trên tổng số', 'tổng số',
        'thao tác', 
        'mã số', 'mã hàng', 'mã sản phẩm',
        'ngày tạo', 'người tạo', 'ngày cập nhật', 'người cập nhật',
    ]
    _STATIC_MODULE_WHITELIST = [
        'cập nhật', 'thêm mới', 'xóa', 'tìm kiếm', 'tìm', 'quay lại',
        'xem chi tiết', 'xuất file excel', 'xuất file word', 'tải xuống',
        'sinh mã', 'đóng', 'hủy bỏ', 'phân trang', 'số dòng',
    ]
    def _remove_static_modules(self, data: dict) -> dict:
        """
        Lọc cứng sau _unwrap_modules: loại bỏ chức năng tĩnh/hiển thị chỉ đọc.
        Chạy cho cả TH1 (targeted) lẫn TH2 (full scan).
        """
        modules = data.get('modules', {})
        kept = {}
        for mod_name, tcs in modules.items():
            name_lower = mod_name.strip().lower()
            is_breadcrumb = '>' in mod_name or name_lower.startswith('trang chủ')
            if is_breadcrumb:
                continue
            if any(w in name_lower for w in self._STATIC_MODULE_WHITELIST):
                kept[mod_name] = tcs
                continue
            is_static = any(p in name_lower for p in self._STATIC_MODULE_PATTERNS)
            is_empty_stub = (
                len(name_lower) <= 4
                and isinstance(tcs, list)
                and len(tcs) == 0
            )
            if not is_static and not is_empty_stub:
                kept[mod_name] = tcs
        if kept:
            data['modules'] = kept
        return data
    def _detect_missing_modules(self, scanned: str, modules: dict) -> list[str]:
        """
        Phát hiện chức năng UI xuất hiện trong ảnh (scanned) nhưng KHÔNG có
        chức năng tương ứng nào trong kết quả.
        Logic thật nằm ở coverage_checker.detect_missing_modules (đã tách
        ra file riêng — xem coverage_checker.py để biết chi tiết + lý do).
        Wrapper này chỉ giữ lại để không phải sửa các điểm gọi khác trong
        file. Nếu import coverage_checker thất bại, fallback trả về rỗng
        (không detect được gì, không raise lỗi làm gãy pipeline).
        """
        if _cc_detect_missing_modules is None:
            return []
        result = _cc_detect_missing_modules(scanned, modules)
        if result:
            print(
                "=== COVERAGE CHECKER (chức năng, có ảnh): PHÁT HIỆN THIẾU ===\n"
                f"{result}\n"
                "==========================================================="
            )
        else:
            print("=== COVERAGE CHECKER (chức năng, có ảnh): đủ chức năng, không thiếu gì ===")
        return result
    def _detect_missing_targeted_modules(self, description: str, modules: dict) -> list[str]:
        """
        Bản tương đương _detect_missing_modules nhưng dùng cho trường hợp
        KHÔNG CÓ ẢNH (mô tả text thuần) — vì khi đó `scanned` luôn rỗng nên
        _detect_missing_modules không detect được gì.
        Logic thật nằm ở coverage_checker.detect_missing_targeted_modules
        (đã tách ra file riêng). Wrapper này chỉ giữ lại để không phải sửa
        các điểm gọi khác trong file. Nếu import coverage_checker thất bại,
        fallback trả về rỗng (không detect được gì, không raise lỗi làm
        gãy pipeline).
        """
        if _cc_detect_missing_targeted_modules is None:
            return []
        result = _cc_detect_missing_targeted_modules(description, modules)
        if result:
            print(
                "=== COVERAGE CHECKER (chức năng, text-only): PHÁT HIỆN THIẾU ===\n"
                f"{result}\n"
                "===============================================================")
        else:
            print("=== COVERAGE CHECKER (chức năng, text-only): đủ chức năng, không thiếu gì ===")
        return result
    def _get_domain_hints(
        self,
        domain: str | None,
        description: str,
        proj: str,
        relevant_elements: str,
    ) -> str:
        """
        Domain Rule Engine: đọc BUSINESS_RULES + TEST_CASES mẫu từ đúng
        domain (bank/hospital/school/recruitment...) do UI truyền vào,
        match cứng theo tên chức năng — khác với _build_rule_engine_hints
        (match theo TÊN FIELD, không quan tâm domain) và RAG (semantic
        search, top-k, không đảm bảo đúng chức năng). Chạy SONG SONG với cả
        hai, không thay thế.
        Logic thật nằm ở rule_engine.select_domain_rules (file riêng).
        Trả về "" nếu không có domain, không có file, hoặc import lỗi
        (không raise, không làm gãy pipeline).
        """
        if not domain or _re_select_domain_rules is None:
            return ""
        hints = _re_select_domain_rules(domain, description, proj, relevant_elements)
        if hints:
            print(f"=== DOMAIN RULE ENGINE ({domain}): match được module, đã nhúng vào prompt ===")
        else:
            print(f"=== DOMAIN RULE ENGINE ({domain}): không match module nào, bỏ qua ===")
        return hints
    def _run_coverage_round(
        self,
        result: dict,
        scanned: str | None,
        image_blocks: list[dict] | None,
        targeted: bool,
        description: str | None,
    ) -> dict:
        """
        1 VÒNG kiểm tra Coverage Checker — dùng CHUNG cho cả Round 1 (trước
        khi gọi API bổ sung) lẫn Round 2 (SAU khi đã merge kết quả bổ
        sung), để đảm bảo 2 vòng KHÔNG BAO GIỜ lệch tiêu chí so khớp nhau
        (nguồn gốc 1 dạng bug đã gặp: Round 1 và report cuối dùng 2 hàm
        tính khác nhau → kết quả không nhất quán).
        Trả về {'missing_modules': [...], 'missing_counts': [...], 'missing_types': [...]}
        — luôn tính TRÊN `result` truyền vào tại đúng thời điểm gọi, KHÔNG
        cache/tái sử dụng số liệu từ vòng trước.
        """
        modules = result.get('modules', {}) if isinstance(result, dict) else {}
        if image_blocks:
            missing_modules = (
                [] if targeted else self._detect_missing_modules(scanned or '', modules)
            )
        else:
            missing_modules = self._detect_missing_targeted_modules(description or '', modules)
        return {
            'missing_modules': missing_modules,
            'missing_counts': self._validate_testcase_count(result),
            'missing_types': self._evaluate_coverage(result),
        }
    def _enforce_min_coverage(
        self,
        result: dict | None,
        scanned: str | None,
        image_blocks: list[dict] | None,
        system_prompt: str,
        proj: str,
        max_rounds: int = 2,
        targeted: bool = False,
        description: str | None = None,
    ) -> dict | None:
        """
        Coverage Checker — LUỒNG CỐ ĐỊNH, ĐÚNG TỐI ĐA 2 VÒNG, KHÔNG for/while
        retry:
            coverage_round_1
            → nếu ĐỦ (không thiếu chức năng/số lượng/loại): dừng ngay, KHÔNG
              gọi AI repair, KHÔNG có vòng 2. coverage_rounds_used = 1.
            → nếu THIẾU: gộp TẤT CẢ thiếu sót (chức năng thiếu hẳn + chức năng
              thiếu số lượng TC + chức năng thiếu loại kịch bản) vào ĐÚNG 1
              prompt repair, gọi AI ĐÚNG 1 LẦN → merge kết quả mới vào
              `result` cũ (không xoá/đổi TC cũ) → coverage_round_2 (kiểm
              tra lại, CHỈ để log/báo cáo — KHÔNG gọi thêm AI dù vẫn còn
              thiếu). coverage_rounds_used = 2.
        API retry do timeout/rate limit/lỗi mạng/JSON lỗi xảy ra BÊN TRONG
        _call_api KHÔNG được tính là 1 "vòng Coverage Checker" — vòng ở đây
        chỉ đếm số lần gọi _run_coverage_round (đúng 2 lần tối đa), không
        đếm theo số lần gọi API hay số loại thiếu sót.
        targeted=True (TH1, CÓ ẢNH): "thiếu chức năng so với ảnh" KHÔNG áp dụng
        — user chỉ yêu cầu một số chức năng cụ thể, các chức năng khác có
        trong ảnh (vd "Xóa", "Thêm mới"...) là CỐ Ý không được sinh, không
        phải thiếu sót. Nhánh _detect_missing_modules vì vậy bị TẮT khi
        targeted=True VÀ có ảnh, chỉ còn nhánh bổ sung SỐ LƯỢNG TC.
        Khi KHÔNG CÓ ẢNH (mô tả text thuần), không dùng _detect_missing_modules
        (scanned luôn rỗng nên vô nghĩa) mà dùng _detect_missing_targeted_modules
        để so khớp trực tiếp với các chức năng user đã liệt kê trong
        `description` — luôn bật kể cả khi targeted=True, vì ở đây targeted
        chỉ có nghĩa "user liệt kê cụ thể", không có khái niệm "ảnh có chức
        năng khác cố ý bỏ qua".
        `max_rounds` giữ lại làm tham số cho tương thích ngược với các nơi
        gọi cũ nhưng KHÔNG còn dùng để lặp — luồng luôn dừng tối đa ở vòng 2.
        """
        if not result or not isinstance(result, dict):
            return result
        round_1 = self._run_coverage_round(result, scanned, image_blocks, targeted, description)
        missing_modules = round_1['missing_modules']
        missing_counts = round_1['missing_counts']
        missing_types = round_1['missing_types']

        if not missing_modules and not missing_counts and not missing_types:
            self._last_coverage_rounds_used = 1
            print(
                "[EnforceMinCoverage] Round 1 ĐÃ ĐỦ — không gọi AI repair, "
                "không chạy Round 2. coverage_rounds_used=1"
            )
            return result
        missing_modules_resolved = self._resolve_missing_module_display_names(missing_modules, scanned)
        repair_prompt = self._build_combined_repair_prompt(
            proj, missing_modules_resolved, missing_counts, missing_types
        )
        try:
            addition = self._call_api(repair_prompt, image_blocks, system_prompt)
        except Exception:
            addition = None

        if addition and addition.get('modules'):
            for tcs in addition.get('modules', {}).values():
                if isinstance(tcs, list):
                    for tc in tcs:
                        if isinstance(tc, dict):
                            tc['id'] = None
            result = self._merge_test_cases(result, addition)
        round_2 = self._run_coverage_round(result, scanned, image_blocks, targeted, description)
        self._last_coverage_rounds_used = 2
        print(
            "[EnforceMinCoverage] Round 2 (kiểm tra lại SAU repair+merge) — "
            f"missing_modules={round_2['missing_modules'] or '(không có)'} | "
            f"missing_counts={round_2['missing_counts'] or '(không có)'} | "
            f"missing_types={round_2['missing_types'] or '(không có)'} | "
            "coverage_rounds_used=2"
        )
        if round_2['missing_modules'] or round_2['missing_counts'] or round_2['missing_types']:
            result = self._fill_remaining_coverage_locally(
                result=result,
                missing_modules=round_2['missing_modules'],
                missing_counts=round_2['missing_counts'],
                missing_types=round_2['missing_types'],
                scanned=scanned or '',
                project_name=proj,
            )
            print(
                "[EnforceMinCoverage] Round 2 còn thiếu — đã bù bằng rule/template "
                "nội bộ, KHÔNG gọi AI vòng 3."
            )
        return result
    def _fill_remaining_coverage_locally(
        self,
        result: dict,
        missing_modules: list[str],
        missing_counts: list[str],
        missing_types: list[str],
        scanned: str,
        project_name: str,
    ) -> dict:
        """
        Bù coverage còn thiếu sau Round 2 bằng template nội bộ.
        Hàm này KHÔNG gọi API nên không tạo Coverage Round 3. Nó chỉ bảo đảm:
        - chức năng Round 2 còn thiếu được tạo;
        - chức năng 1/4, 2/4... được bổ sung đủ số lượng;
        - testcase bổ sung chứa đúng từ khóa loại kịch bản để report cuối nhận diện.
        """
        if not isinstance(result, dict):
            return result
        modules = result.setdefault('modules', {})
        if not isinstance(modules, dict):
            result['modules'] = {}
            modules = result['modules']
        resolved_missing = self._resolve_missing_module_display_names(missing_modules or [], scanned)
        for module_name in resolved_missing:
            if not module_name:
                continue
            existing_name = self._find_existing_module_name(modules, module_name)
            if existing_name is None:
                modules[module_name] = []
        required_by_module: dict[str, int] = {}
        for item in missing_counts or []:
            text = str(item or '')
            match = re.match(r'^(.*?):\s*\d+/(\d+)', text)
            if match:
                required_by_module[match.group(1).strip()] = int(match.group(2))
        for item in missing_types or []:
            name = str(item or '').split(':', 1)[0].strip()
            if name and name not in required_by_module:
                existing_name = self._find_existing_module_name(modules, name) or name
                current = modules.get(existing_name, [])
                required_by_module[name] = max(len(current) if isinstance(current, list) else 0, 2)
        for name in resolved_missing:
            if name and name not in required_by_module:
                required_by_module[name] = self._local_required_count(name)
        for requested_name, required in required_by_module.items():
            module_name = self._find_existing_module_name(modules, requested_name) or requested_name
            current = modules.setdefault(module_name, [])
            if not isinstance(current, list):
                current = []
                modules[module_name] = current
            templates = self._local_coverage_templates(module_name)
            existing_keys = {
                self._local_tc_key(tc) for tc in current if isinstance(tc, dict)
            }
            for template in templates:
                if len(current) >= required:
                    break
                tc = self._build_local_coverage_testcase(
                    module_name, template, project_name
                )
                key = self._local_tc_key(tc)
                if key in existing_keys:
                    continue
                current.append(tc)
                existing_keys.add(key)
            while len(current) < required:
                index = len(current) + 1
                template = {
                    'title': f'{module_name} - trường hợp ngoại lệ {index}',
                    'scenario': f'Thực hiện {module_name} với dữ liệu ngoại lệ hợp lệ để kiểm tra lần {index}',
                    'expected_result': 'Hệ thống xử lý ổn định, không tạo dữ liệu trùng và hiển thị phản hồi phù hợp',
                    'test_type': 'Kiểm thử tích hợp',
                    'priority': 'Trung bình',
                }
                current.append(self._build_local_coverage_testcase(module_name, template, project_name))
        return result

    @staticmethod
    def _module_names_equivalent(name_a: str, name_b: str) -> bool:
        """So khớp an toàn hai tên chức năng khi canonical key chưa đủ.
        Không dùng substring thô cho các action độc lập như
        "Thêm mới" và "Thêm mới và tiếp tục".
        """
        a = re.sub(r'\s+', ' ', (name_a or '').strip().lower())
        b = re.sub(r'\s+', ' ', (name_b or '').strip().lower())
        if not a or not b:
            return False
        if a == b:
            return True
        if _cc_canonical_module_key is not None:
            try:
                key_a = _cc_canonical_module_key(a)
                key_b = _cc_canonical_module_key(b)
                if key_a in {'search_input', 'search_button'} or key_b in {'search_input', 'search_button'}:
                    return key_a == key_b
            except Exception:
                pass

        if ('tiếp tục' in a) != ('tiếp tục' in b):
            return False

        aliases = (
            {'xóa', 'xoá', 'delete', 'remove'},
            {'cập nhật', 'chỉnh sửa', 'sửa', 'edit', 'update'},
            {'hủy bỏ', 'huỷ bỏ', 'hủy', 'huỷ', 'cancel'},
            {'quay lại', 'back'},
            {'sinh mã', 'tạo mã', 'generate code'},
        )
        for group in aliases:
            if a in group and b in group:
                return True

        return min(len(a), len(b)) >= 5 and (a in b or b in a)

    def _find_existing_module_name(self, modules: dict, requested_name: str) -> str | None:
        """Tìm tên chức năng thật theo canonical key, tránh tạo chức năng gần-trùng."""
        wanted = _cc_canonical_module_key(requested_name) if _cc_canonical_module_key else requested_name.strip().lower()
        for name in modules.keys():
            key = _cc_canonical_module_key(name) if _cc_canonical_module_key else str(name).strip().lower()
            if key == wanted:
                return name
            if self._module_names_equivalent(str(name), requested_name):
                return name
        return None
    def _local_required_count(self, module_name: str) -> int:
        lower = (module_name or '').strip().lower()
        if 'quay lại' in lower or 'phân trang' in lower:
            return 4
        if self._is_popup_action_module_4tc(lower):
            return 4
        if self._is_them_moi_module(module_name):
            return 4
        if lower == 'tìm' or lower.startswith('tìm kiếm') or lower.startswith('cập nhật') or lower.startswith('chỉnh sửa'):
            return 2
        return 4
    @staticmethod
    def _local_tc_key(tc: dict) -> str:
        text = ' | '.join(str(tc.get(k) or '') for k in ('chức năng', 'title', 'scenario', 'expected_result'))
        return re.sub(r'\s+', ' ', text.strip().lower())
    def _build_local_coverage_testcase(
        self, module_name: str, template: dict, project_name: str
    ) -> dict:
        scenario = template['scenario']
        expected = template['expected_result']
        return {
            'id': None,
            'chức năng': module_name,
            'feature': module_name,
            'scenario': scenario,
            'title': template['title'],
            'description': scenario,
            'given': f'Người dùng đang ở màn hình {project_name or "dự án"}',
            'when': scenario,
            'then': expected,
            'precondition': 'Người dùng có quyền truy cập chức năng',
            'steps': f'1. Mở chức năng {module_name}\n2. {scenario}\n3. Quan sát kết quả',
            'test_data': template.get('test_data', 'Dữ liệu phù hợp với tình huống kiểm thử'),
            'expected_result': expected,
            'priority': template.get('priority', 'Cao'),
            'test_type': template.get('test_type', 'Kiểm thử chức năng'),
            'actual_result': '',
            'status': 'Chưa chạy',
            'note': '',
        }
    def _local_coverage_templates(self, module_name: str) -> list[dict]:
        """Template cố định để Round 2 bù đủ TC mà không gọi AI lần nữa."""
        lower = (module_name or '').strip().lower()
        if 'sinh mã' in lower:
            return [
                {'title': 'Sinh mã thành công', 'scenario': 'Nhấn Sinh mã khi trường mã đang trống để sinh mã thành công', 'expected_result': 'Hệ thống tạo mã thành công, đúng định dạng và điền vào trường Mã', 'test_type': 'Kiểm thử dương'},
                {'title': 'Mã sinh tự động không trùng', 'scenario': 'Sinh mã cho nhiều bản ghi liên tiếp và kiểm tra tính duy nhất', 'expected_result': 'Mỗi mã được tạo đúng định dạng, duy nhất và không trùng dữ liệu hiện có', 'test_type': 'Kiểm thử xác thực'},
                {'title': 'Nhấn Sinh mã nhiều lần', 'scenario': 'Nhấn nhiều lần nút Sinh mã liên tục', 'expected_result': 'Hệ thống xử lý ổn định, không sinh bản ghi trùng và chỉ giữ mã hợp lệ cuối cùng', 'test_type': 'Kiểm thử tích hợp'},
                {'title': 'Lỗi sinh mã', 'scenario': 'Thực hiện sinh mã khi dịch vụ tạo mã gặp lỗi sinh mã', 'expected_result': 'Hệ thống thông báo lỗi sinh mã, không điền mã sai và cho phép thử lại', 'test_type': 'Kiểm thử âm'},
            ]
        if 'hủy bỏ' in lower or 'huỷ bỏ' in lower:
            return [
                {'title': 'Hủy khi chưa nhập dữ liệu', 'scenario': 'Nhấn Hủy bỏ khi chưa nhập dữ liệu', 'expected_result': 'Popup đóng đúng, không lưu dữ liệu và trở về màn hình danh sách', 'test_type': 'Kiểm thử dương'},
                {'title': 'Hủy khi đã nhập dữ liệu', 'scenario': 'Đã nhập dữ liệu vào form rồi nhấn Hủy bỏ', 'expected_result': 'Popup đóng đúng, dữ liệu đã nhập không được lưu', 'test_type': 'Kiểm thử chức năng'},
                {'title': 'Không lưu dữ liệu sau khi hủy', 'scenario': 'Thay đổi nhiều trường rồi chọn Hủy bỏ và mở lại popup', 'expected_result': 'Không lưu dữ liệu; popup mở lại với giá trị mặc định', 'test_type': 'Kiểm thử âm'},
                {'title': 'Hủy bỏ đóng popup đúng', 'scenario': 'Nhấn Hủy bỏ nhiều lần hoặc khi popup đang xử lý', 'expected_result': 'Popup đóng đúng một lần, giao diện không treo và dữ liệu không thay đổi', 'test_type': 'Kiểm thử tích hợp'},
            ]
        if 'thêm mới và tiếp tục' in lower:
            return [
                {'title': 'Thêm mới và tiếp tục thành công', 'scenario': 'Nhập dữ liệu hợp lệ rồi nhấn Thêm mới và tiếp tục để thực hiện thành công', 'expected_result': 'Thêm mới thành công, bản ghi được lưu và form được làm trống để nhập tiếp', 'test_type': 'Kiểm thử dương'},
                {'title': 'Thêm mới và tiếp tục không thành công', 'scenario': 'Bỏ trống trường bắt buộc rồi nhấn Thêm mới và tiếp tục', 'expected_result': 'Thao tác không thành công, hiển thị lỗi bắt buộc, không lưu và popup vẫn mở', 'test_type': 'Kiểm thử xác thực'},
                {'title': 'Dữ liệu không hợp lệ', 'scenario': 'Nhập dữ liệu không hợp lệ hoặc vượt độ dài rồi nhấn Thêm mới và tiếp tục', 'expected_result': 'Hệ thống hiển thị lỗi phù hợp, không lưu dữ liệu không hợp lệ', 'test_type': 'Kiểm thử âm'},
                {'title': 'Ngăn tạo bản ghi trùng khi nhấn nhiều lần', 'scenario': 'Nhấn Thêm mới và tiếp tục nhiều lần liên tiếp với cùng dữ liệu hợp lệ', 'expected_result': 'Mỗi lần gửi hợp lệ chỉ tạo một bản ghi, không phát sinh dữ liệu trùng', 'test_type': 'Kiểm thử tích hợp'},
            ]
        if 'phân trang' in lower:
            return [
                {'title': 'Chuyển sang trang kế tiếp', 'scenario': 'Nhấn Next để chuyển sang trang kế tiếp', 'expected_result': 'Hiển thị dữ liệu trang sau đúng thứ tự', 'test_type': 'Kiểm thử chức năng'},
                {'title': 'Chuyển về trang trước', 'scenario': 'Nhấn Prev để chuyển về trang trước', 'expected_result': 'Hiển thị dữ liệu trang trước đó', 'test_type': 'Kiểm thử chức năng'},
                {'title': 'Chuyển sang trang đầu', 'scenario': 'Nhấn First để chuyển về trang đầu', 'expected_result': 'Hiển thị dữ liệu trang đầu tiên', 'test_type': 'Kiểm thử biên'},
                {'title': 'Chuyển sang trang cuối', 'scenario': 'Nhấn Last để chuyển sang trang cuối', 'expected_result': 'Hiển thị dữ liệu trang cuối cùng', 'test_type': 'Kiểm thử biên'},
            ]
        if 'quay lại' in lower:
            return [
                {'title': 'Quay lại khi chưa thay đổi dữ liệu', 'scenario': 'Nhấn Quay lại khi chưa thay đổi dữ liệu', 'expected_result': 'Về màn hình trước, không cảnh báo và không thay đổi dữ liệu', 'test_type': 'Kiểm thử dương'},
                {'title': 'Quay lại khi đã thay đổi dữ liệu', 'scenario': 'Đã thay đổi dữ liệu rồi nhấn Quay lại', 'expected_result': 'Hiển thị cảnh báo phù hợp và xử lý điều hướng theo lựa chọn người dùng', 'test_type': 'Kiểm thử chức năng'},
                {'title': 'Quay lại từ màn hình chi tiết hoặc sửa', 'scenario': 'Từ màn hình chi tiết hoặc sửa, nhấn Quay lại', 'expected_result': 'Chuyển đúng về màn hình danh sách vừa truy cập', 'test_type': 'Kiểm thử tích hợp'},
                {'title': 'Quay lại khi không có màn hình trước', 'scenario': 'Truy cập trực tiếp khi không có màn hình trước rồi nhấn Quay lại', 'expected_result': 'Chuyển về trang mặc định hoặc giữ nguyên màn hình an toàn', 'test_type': 'Kiểm thử biên'},
            ]
        return [
            {'title': f'{module_name} thành công', 'scenario': f'Thực hiện {module_name} với dữ liệu hợp lệ và thao tác thành công', 'expected_result': 'Hệ thống xử lý thành công và cập nhật dữ liệu đúng', 'test_type': 'Kiểm thử dương'},
            {'title': f'{module_name} không thành công', 'scenario': f'Thực hiện {module_name} với dữ liệu không hợp lệ hoặc điều kiện lỗi', 'expected_result': 'Thao tác không thành công, hiển thị lỗi phù hợp và không cập nhật dữ liệu sai', 'test_type': 'Kiểm thử âm'},
            {'title': f'{module_name} với dữ liệu biên', 'scenario': f'Thực hiện {module_name} tại giá trị giới hạn cho phép', 'expected_result': 'Hệ thống xử lý đúng giá trị biên theo quy định', 'test_type': 'Kiểm thử biên'},
            {'title': f'{module_name} khi thao tác lặp', 'scenario': f'Thực hiện {module_name} nhiều lần liên tiếp', 'expected_result': 'Hệ thống xử lý ổn định, không tạo dữ liệu trùng và không treo giao diện', 'test_type': 'Kiểm thử tích hợp'},
        ]
    def _resolve_missing_module_display_names(
        self, missing_modules: list[str], scanned: str | None,
    ) -> list[str]:
        """
        Thay nhãn CHUNG CHUNG "Tìm kiếm" (đến từ _STANDARD_CHECK_LABELS của
        Coverage Checker — chỉ là nhãn hiển thị cho báo cáo, KHÔNG phải tên
        chức năng thật) bằng tên chức năng CỤ THỂ lấy trực tiếp từ dòng OCR trong
        `scanned` (vd "Tìm kiếm theo mã hoặc tên chu kỳ") nếu tìm được, để
        prompt repair KHÔNG yêu cầu AI tạo chức năng tên chung chung "Tìm
        kiếm" — tên chung chung này không tương ứng UI element thật nào cụ
        thể, và là nguyên nhân sinh chức năng DƯ bên cạnh chức năng cụ thể-hơn.
        Nút "Tìm" và các nhãn khác (Thêm mới/Quay lại/Cập nhật/Xóa/Phân
        trang) giữ nguyên vì đó ĐÃ LÀ tên chức năng chuẩn của dự án.
        """
        if not missing_modules or not scanned:
            return missing_modules
        resolved = []
        for name in missing_modules:
            if name.strip().lower() == 'tìm kiếm':
                specific = self._extract_search_input_label(scanned)
                resolved.append(specific or name)
            else:
                resolved.append(name)
        return resolved
    def _extract_search_input_label(self, scanned: str) -> str | None:
        """
        Tìm dòng OCR mô tả Ô INPUT tìm kiếm CỤ THỂ (vd dòng dạng
        "- Tìm kiếm theo mã hoặc tên chu kỳ | input") để lấy đúng tên hiển
        thị trên UI, tránh dùng tên chung chung "Tìm kiếm" khi yêu cầu AI
        sinh chức năng mới. Trả về None nếu không tìm thấy dòng nào phù hợp.
        """
        if not scanned:
            return None
        for line in scanned.splitlines():
            m = re.search(r'-\s*(tìm kiếm[^|]*?)\s*\|', line, re.IGNORECASE)
            if m:
                label = m.group(1).strip()
                if label:
                    return label
        m2 = re.search(r'-\s*(tìm kiếm[^\n|]*)', scanned, re.IGNORECASE)
        if m2:
            label = m2.group(1).strip()
            return label or None
        return None
    def _build_combined_repair_prompt(
        self,
        proj: str,
        missing_modules: list[str],
        missing_counts: list[str],
        missing_types: list[str],
    ) -> str:
        """
        Gộp TẤT CẢ thiếu sót của Round 1 (chức năng thiếu hẳn + chức năng thiếu
        số lượng TC + chức năng thiếu loại kịch bản) vào ĐÚNG 1 prompt, thay
        cho 3 prompt riêng gọi API 3 lần trước đây — giảm số lần gọi AI bổ
        sung xuống tối đa 1 lần/request thay vì tối đa 3 lần/vòng x nhiều
        vòng. Prompt CHỈ chứa: tên project, danh sách thiếu theo từng loại,
        business rule tối thiểu liên quan, và JSON schema cần trả về —
        KHÔNG gửi lại ảnh dưới dạng mô tả thêm (đã có sẵn trong
        `image_blocks` truyền riêng cho _call_api), KHÔNG gửi lại RAG
        context, KHÔNG gửi lại toàn bộ TC cũ, KHÔNG lặp lại system prompt
        dài (system_prompt đã truyền riêng).
        """
        sections = [f'Màn hình: "{proj}"']
        if missing_modules:
            sections.append(
                "1) CHỨC NĂNG CÒN THIẾU HOÀN TOÀN — sinh CHỨC NĂNG MỚI HOÀN TOÀN cho "
                "TỪNG chức năng dưới đây. Mỗi chức năng chỉ có ĐÚNG 2 nhóm TC "
                "chính: Thành công / Không thành công (GỘP mọi biến thể lỗi "
                "— bỏ trống, sai định dạng, vượt giới hạn, dữ liệu trùng, "
                "không đủ quyền, dữ liệu đang tham chiếu... — vào 1 TC \"Không "
                "thành công\" DUY NHẤT, KHÔNG tách lẻ Positive/Negative/"
                "Validation/Boundary thành nhiều TC riêng). "
                "QUAN TRỌNG — phân biệt đúng UI element, giữ NGUYÊN VĂN tên "
                "chức năng được liệt kê dưới đây (KHÔNG đổi tên, KHÔNG dùng tên "
                "chung chung thay thế): nếu tên là \"Tìm\" thì đây là NÚT BẤM "
                "riêng biệt (KHÔNG PHẢI ô input), TUYỆT ĐỐI KHÔNG đổi thành "
                "\"Tìm kiếm\"/không gộp vào ô input tìm kiếm đã có; nếu tên có "
                "dạng \"Tìm kiếm theo [X]\" thì đây là Ô INPUT lọc dữ liệu, "
                "giữ nguyên tên cụ thể đó, KHÔNG rút gọn thành \"Tìm kiếm\" "
                "chung chung. Với chức năng \"Thêm mới\": KHÔNG ép cứng số lượng TC — "
                "tách theo NHÓM NGHIỆP VỤ hợp lý dựa trên field thực tế của form, "
                "mỗi TC CHỈ mô tả 1 nhóm lỗi liên quan (KHÔNG nhồi Required/Length/"
                "Whitespace/Validation/Boundary/Duplicate/XSS/SQL Injection vào 1 TC): "
                "(1) thành công — nhập đầy đủ dữ liệu hợp lệ, lưu OK; (2) nếu có "
                "trường định danh cho phép bỏ trống — TC riêng: để trống, hệ thống "
                "tự sinh hợp lệ, không trùng, lưu OK; (3) nếu có trường bắt buộc — "
                "TC riêng: bỏ trống/chỉ nhập khoảng trắng→lỗi bắt buộc, không lưu; "
                "(4) TC riêng: trùng/sai định dạng/vượt độ dài→lỗi phù hợp, không "
                "lưu; (5) TC riêng: XSS/SQL Injection→không thực thi mã, không lưu "
                "dữ liệu nguy hiểm. Với chức năng \"Cập nhật\" (icon bút vàng/Edit): ĐÚNG 2 "
                "TC — (1) Cập nhật thành công, (2) Cập nhật không thành công "
                "(GỘP mọi lỗi input vào 1 TC).\n- "
                + "\n- ".join(missing_modules)
            )
        if missing_counts:
            sections.append(
                "2) CHỨC NĂNG CHƯA ĐỦ SỐ TC TỐI THIỂU — đối chiếu phần \"cần đủ\" "
                "với TC đã có, CHỈ sinh thêm TC cho ĐÚNG kịch bản còn thiếu "
                "(KHÔNG lặp lại ý TC đã có, KHÔNG tạo thêm biến thể khác của "
                "kịch bản đã tồn tại):\n- "
                + "\n- ".join(missing_counts)
            )
        if missing_types:
            sections.append(
                "3) CHỨC NĂNG ĐỦ SỐ LƯỢNG NHƯNG THIẾU LOẠI KỊCH BẢN (TC hiện có "
                "đang lặp cùng 1 loại) — chỉ sinh THÊM TC cho ĐÚNG loại còn "
                "thiếu được liệt kê, KHÔNG lặp lại loại đã có, KHÔNG xoá/đổi "
                "TC cũ:\n- "
                + "\n- ".join(missing_types)
            )
        sections.append(
            "Trả về DUY NHẤT 1 JSON object {\"modules\": {...}} gồm CHỈ các "
            "chức năng/TC MỚI cần bổ sung theo đúng các mục trên — không lặp "
            "lại chức năng/TC đã có, không kèm giải thích ngoài JSON."
        )
        return "\n\n".join(sections)
    def _build_generic_module_template(self, module_name: str) -> list[dict]:
        """
        Template TC chuẩn (Thành công / Không thành công) sinh CỤC BỘ,
        KHÔNG cần gọi GPT — dùng làm fallback cuối trong _enforce_min_coverage
        khi chức năng vẫn thiếu sau max_rounds.
        """
        clean_name = module_name.strip()
        return [
            {
                "id": None,
                "chức năng": clean_name,
                "feature": clean_name,
                "title": f"{clean_name} thành công",
                "scenario": f"Thực hiện chức năng \"{clean_name}\" với dữ liệu/thao tác hợp lệ",
                "description": f"Thực hiện chức năng \"{clean_name}\" với dữ liệu/thao tác hợp lệ",
                "steps": (
                    f"1. Thực hiện thao tác \"{clean_name}\" với dữ liệu hợp lệ\n"
                    "2. Xác nhận/Lưu (nếu có)"
                ),
                "expected_result": (
                    "Thao tác thực hiện thành công, hệ thống phản hồi/cập nhật "
                    "đúng dữ liệu tương ứng"
                ),
                "priority": "Trung bình",
                "test_type": "Kiểm thử chức năng",
                "actual_result": "",
                "status": "Chưa chạy",
                "note": "",
            },
            {
                "id": None,
                "chức năng": clean_name,
                "feature": clean_name,
                "title": f"{clean_name} không thành công",
                "scenario": (
                    f"Thực hiện chức năng \"{clean_name}\" với dữ liệu không hợp lệ "
                    "hoặc bỏ trống trường bắt buộc"
                ),
                "description": (
                    f"Thực hiện chức năng \"{clean_name}\" với dữ liệu không hợp lệ "
                    "hoặc bỏ trống trường bắt buộc"
                ),
                "steps": (
                    f"1. Thực hiện thao tác \"{clean_name}\" với dữ liệu không hợp lệ/bỏ trống"
                ),
                "expected_result": (
                    "Xuất hiện thông báo lỗi phù hợp, không lưu/thực hiện thay đổi dữ liệu"
                ),
                "priority": "Trung bình",
                "test_type": "Kiểm thử chức năng",
                "actual_result": "",
                "status": "Chưa chạy",
                "note": "",
            },
        ]
    def _build_final_coverage_report(
        self,
        normalized: dict,
        scanned: str | None,
        image_blocks: list[dict] | None,
        targeted: bool,
        description: str | None,
    ) -> dict:
        """
        BƯỚC CUỐI CÙNG của pipeline (chạy SAU _enforce_min_coverage VÀ SAU
        _normalize_test_cases — tức SAU cả Round 2 lẫn normalize/dedupe):
        generate → coverage_round_1 → repair_missing_cases (nếu cần)
        → merge → coverage_round_2 → normalize → build_final_coverage_report
        → return
        Tính report CUỐI CÙNG hoàn toàn TỪ ĐẦU trên `normalized` (dữ liệu
        THẬT SỰ cuối cùng sẽ trả về cho người dùng) — KHÔNG tái sử dụng số
        liệu missing_modules/missing_counts/missing_types còn sót lại từ
        Round 1 hay Round 2 bên trong _enforce_min_coverage, vì normalize
        (dedupe/fold chức năng trùng, lọc chức năng tĩnh...) có thể đã thay đổi
        `modules` sau các vòng đó — dùng số liệu cũ sẽ gây MÂU THUẪN kiểu
        "Coverage: 100%" nhưng vẫn còn "remaining_missing_cases" (bug đã
        gặp thực tế).
        Đảm bảo BẤT BIẾN (invariant) bằng đúng 1 công thức duy nhất:
        coverage_percent == 100.0  <=>  remaining_missing_cases == []  <=>  warning == ""
        vì cả 3 giá trị đều được tính ra từ CÙNG 1 danh sách
        `remaining_missing_cases` trong CÙNG 1 lần gọi hàm này — không có
        đường nào để 1 trong 3 giá trị bị tính từ nguồn khác/cũ hơn.
        Bỏ qua (không tính, không gắn key) khi:
        - Không import được coverage_checker.build_coverage_report.
        - targeted=True VÀ có ảnh: user chỉ định cụ thể chức năng cần sinh,
        các chức năng khác trong ảnh là CỐ Ý không sinh (không phải thiếu
        sót) — % coverage ở mức chức năng không có ý nghĩa cho case này,
        hiển thị ra sẽ gây hiểu lầm là AI làm thiếu.
        """
        if not isinstance(normalized, dict):
            return normalized
        if self._current_screen_context:
            normalized['_screen_context'] = self._current_screen_context
        if self._current_workflow_relation:
            normalized['_workflow_relation'] = self._current_workflow_relation
        if self._current_workflow_context:
            normalized['_workflow_context'] = self._current_workflow_context
        modules = normalized.get('modules', {})
        form_structure = self._current_form_structure
        if form_structure and _cc_build_form_structure_coverage is not None:
            try:
                fs_report = _cc_build_form_structure_coverage(form_structure, modules)
                if fs_report:
                    normalized['form_structure_coverage'] = fs_report
                    print(
                        "=== FORM STRUCTURE COVERAGE (Field/Button/Business Rule/Workflow) ===\n"
                        f"Form: {fs_report.get('form_name')} | "
                        f"Field: {fs_report.get('field_coverage_percent')}% (thiếu: {fs_report.get('missing_fields') or '(không)'}) | "
                        f"Button: {fs_report.get('button_coverage_percent')}% (thiếu: {fs_report.get('missing_buttons') or '(không)'}) | "
                        f"Business Rule: {fs_report.get('business_rule_coverage_percent')}% (thiếu: {fs_report.get('missing_business_rules') or '(không)'}) | "
                        f"Workflow: {fs_report.get('workflow_coverage_percent')}% (thiếu: {fs_report.get('missing_workflow') or '(không)'})\n"
                        "======================================================================="
                    )
            except Exception as exc:
                print(f"[FormStructureCoverage] Lỗi khi tính coverage: {exc}")
        if _cc_build_coverage_report is None:
            return normalized
        if image_blocks and targeted:
            return normalized
        try:
            if image_blocks:
                module_report = _cc_build_coverage_report(modules, scanned=scanned or '', has_image=True)
            else:
                module_report = _cc_build_coverage_report(modules, description=description or '', has_image=False)
        except Exception as exc:
            print(f"[CoverageReport] Lỗi khi tính coverage report: {exc}")
            return normalized
        missing_counts = self._validate_testcase_count(normalized)
        missing_types = self._evaluate_coverage(normalized)
        remaining_missing_cases = (
            [f"[Thiếu chức năng] {m}" for m in module_report['missing_items']]
            + [f"[Thiếu số lượng TC] {m}" for m in missing_counts]
            + [f"[Thiếu loại kịch bản] {m}" for m in missing_types]
        )
        if not remaining_missing_cases:
            final_coverage_percent = 100.0
        else:
            incomplete_module_names = {
                entry.split(':', 1)[0].strip().lower() for entry in missing_counts
            } | {
                entry.split(':', 1)[0].strip().lower() for entry in missing_types
            }
            fully_covered = [
                item for item in module_report['covered_items']
                if item.strip().lower() not in incomplete_module_names
            ]
            total_expected = module_report['total_expected'] or (
                len(fully_covered) + len(remaining_missing_cases)
            ) or 1
            final_coverage_percent = round(len(fully_covered) / total_expected * 100, 1)
            if final_coverage_percent >= 100.0:
                final_coverage_percent = 99.9
        warning = (
            f"⚠ Còn {len(remaining_missing_cases)} mục chưa hoàn thiện sau "
            f"{self._last_coverage_rounds_used or 1} vòng Coverage Checker — "
            "xem 'remaining_missing_cases' để biết chi tiết."
        ) if remaining_missing_cases else ""
        final_report = {
            **module_report,
            "coverage_percent": final_coverage_percent,
            "remaining_missing_cases": remaining_missing_cases,
            "warning": warning,
            "coverage_rounds_used": self._last_coverage_rounds_used or 1,
        }
        print(
            "=== COVERAGE REPORT (FINAL — sau Round 2 + normalize) ===\n"
            f"Nguồn: {final_report['source']} | Áp dụng: {final_report['total_expected']} | "
            f"Đã có (module-level): {final_report['total_covered']} | "
            f"Coverage: {final_report['coverage_percent']}%\n"
            f"remaining_missing_cases: {remaining_missing_cases or '(rỗng)'}\n"
            f"warning: {warning or '(rỗng)'}\n"
            f"coverage_rounds_used: {final_report['coverage_rounds_used']}\n"
            "==========================================================="
        )
        normalized['coverage_report'] = final_report
        return normalized
    def _validate_testcase_count(self, data: dict):
        """
        LƯU Ý: các ngưỡng required bên dưới được thiết kế cho 1 chức năng
        gộp chung MỌI loại kịch bản (thành công + toàn bộ biến thể lỗi).
        Sau _finalize_success_failure_grouping, 1 chức năng gốc bị tách
        thành 2 module riêng ("<gốc> thành công" / "<gốc> không thành
        công") — nên PHẢI gộp lại theo chức năng gốc (_group_tcs_by_base_function)
        trước khi so ngưỡng, nếu không mỗi nửa sẽ bị so nhầm với ngưỡng
        của cả chức năng gộp và luôn báo "thiếu TC" dù dữ liệu thực tế đủ.
        """
        modules = data.get("modules", {})
        grouped = self._group_tcs_by_base_function(modules)
        missing = []
        for base_name, tcs in grouped.items():
            count = len(tcs)
            lower = base_name.lower()
            if "quay lại" in lower:
                required = 4
            elif "phân trang" in lower:
                required = 4
            elif self._is_popup_action_module_4tc(lower):
                required = 4
            elif self._is_them_moi_module(base_name) and "tiếp tục" not in lower:
                required = self._crud_required_count('create', 6)
            elif lower.startswith(('cập nhật', 'chỉnh sửa')) or lower in {'sửa', 'update', 'edit'}:
                required = self._crud_required_count('update', 7)
            elif lower.startswith(('xóa', 'xoá')) or lower in {'delete'}:
                required = self._crud_required_count('delete', 5)
            elif lower == 'tìm' or lower.startswith('tìm kiếm'):
                required = 2
            else:
                required = 4
            if count < required:
                hint = self._required_scenarios_hint(lower)
                suffix = f" — cần đủ: {hint}" if hint else ""
                missing.append(f"{base_name}: {count}/{required}{suffix}")
        return missing
    _COVERAGE_TYPE_RULES: list[tuple[list[str], list[tuple[str, list[str]]]]] = [
        (["thêm mới"], [
            ("thành công", ["thêm thành công", "lưu thành công", "được lưu"]),
            ("required/whitespace", ["bắt buộc", "bỏ trống", "để trống", "khoảng trắng"]),
            ("format/duplicate/boundary", ["sai định dạng", "trùng", "vượt độ dài", "giới hạn", "biên"]),
            ("security", ["xss", "sql injection", "mã độc", "không thực thi"]),
        ]),
        (["cập nhật", "chỉnh sửa"], [
            ("mở đúng bản ghi", ["đúng bản ghi", "đúng dữ liệu", "điền sẵn"]),
            ("cập nhật thành công", ["cập nhật thành công", "lưu hợp lệ", "được cập nhật"]),
            ("không thay đổi", ["không thay đổi", "không chỉnh sửa"]),
            ("required/format/boundary", ["bắt buộc", "sai định dạng", "trùng", "vượt độ dài", "khoảng trắng"]),
            ("security", ["xss", "sql injection", "mã độc", "không thực thi"]),
        ]),
        (["xóa", "xoá"], [
            ("mở xác nhận", ["xác nhận", "popup", "hộp thoại"]),
            ("xóa thành công", ["xóa thành công", "không còn trong danh sách"]),
            ("hủy/đóng", ["hủy", "đóng", "giữ nguyên"]),
            ("ràng buộc tham chiếu", ["tham chiếu", "đang sử dụng", "không thể xóa"]),
            ("permission/not-found/system", ["không có quyền", "không tồn tại", "lỗi hệ thống", "mất mạng"]),
        ]),
        (["phân trang"], [
            ("next/tiếp theo", ["next", "tiếp theo", "trang sau", "kế tiếp"]),
            ("prev/trước", ["prev", "trang trước", "trước đó"]),
            ("trang đầu/first", ["first", "trang đầu", "đầu tiên"]),
            ("trang cuối/last", ["last", "trang cuối", "cuối cùng"]),
        ]),
        (["quay lại"], [
            ("chưa thay đổi dữ liệu", ["chưa thay đổi", "chưa đổi", "không cảnh báo"]),
            ("đã thay đổi dữ liệu", ["đã thay đổi", "đã đổi", "thông báo thành công"]),
            ("từ màn hình chi tiết/sửa", ["chi tiết", "sửa"]),
            ("không có màn hình trước", ["không có màn hình trước", "mặc định"]),
        ]),
        (["sinh mã"], [
            ("sinh mã thành công", ["sinh mã thành công", "sinh thành công", "tạo mã thành công"]),
            ("mã không trùng/đúng định dạng", ["không trùng", "đúng định dạng", "duy nhất"]),
            ("nhấn nhiều lần", ["nhấn nhiều lần", "click nhiều lần", "bấm nhiều lần", "liên tục"]),
            ("lỗi sinh mã", ["lỗi sinh mã", "sinh mã thất bại", "không sinh được"]),
        ]),
        (["hủy bỏ", "huỷ bỏ"], [
            ("hủy khi chưa nhập dữ liệu", ["chưa nhập", "chưa nhập dữ liệu", "chưa thay đổi"]),
            ("hủy khi đã nhập dữ liệu", ["đã nhập", "đã nhập dữ liệu", "đã thay đổi"]),
            ("không lưu dữ liệu", ["không lưu"]),
            ("popup đóng đúng", ["popup đóng", "đóng popup", "đóng đúng"]),
        ]),
        (["đóng popup"], [
            ("đóng khi chưa thay đổi", ["chưa thay đổi", "chưa nhập"]),
            ("đóng khi đã nhập dữ liệu", ["đã nhập", "đã thay đổi"]),
            ("không lưu dữ liệu", ["không lưu"]),
            ("quay lại đúng màn hình trước", ["quay lại đúng", "màn hình trước", "về đúng màn hình"]),
        ]),
    ]
    _DEFAULT_COVERAGE_RULE: list[tuple[str, list[str]]] = [
        ("thành công", ["thành công"]),
        ("không thành công", ["không thành công", "thất bại", "không hợp lệ", "lỗi"]),
    ]
    def _evaluate_coverage(self, data: dict) -> list[str]:
        """
        BƯỚC CUỐI (chạy trong generate_test_cases, ngay trước khi trả kết
        quả) — khác với _validate_testcase_count (chỉ đếm SỐ LƯỢNG TC) và
        _coverage_checker (kiểm tra OCR/scan ảnh trước khi sinh TC). Hàm
        này kiểm tra LOẠI kịch bản: 1 chức năng có thể đủ số lượng TC tối
        thiểu nhưng vẫn THIẾU vì tất cả TC đều thuộc cùng 1 loại (vd 3 TC
        "Thêm mới" nhưng không có case lỗi/validation nào).

        Nhận diện loại kịch bản qua từ khóa xuất hiện trong scenario/
        description/title + expected_result/then của từng TC. Trả về danh
        sách mô tả loại còn thiếu theo từng chức năng, dùng để
        _enforce_min_coverage gọi lại API bổ sung (không đụng tới TC đã có).
        """
        modules = data.get('modules', {}) if isinstance(data, dict) else {}
        grouped = self._group_tcs_by_base_function(modules)
        missing: list[str] = []
        for base_name, tcs in grouped.items():
            if not tcs:
                continue
            lower = base_name.lower()
            requirements = None
            if self._is_them_moi_module(base_name) and 'tiếp tục' not in lower and self._crud_is_list_only('create'):
                requirements = self._DEFAULT_COVERAGE_RULE
            elif (lower.startswith(('cập nhật', 'chỉnh sửa')) or lower in {'sửa', 'update', 'edit'}) and self._crud_is_list_only('update'):
                requirements = self._DEFAULT_COVERAGE_RULE
            elif (lower.startswith(('xóa', 'xoá')) or lower == 'delete') and self._crud_is_list_only('delete'):
                requirements = self._DEFAULT_COVERAGE_RULE
            for keywords, reqs in self._COVERAGE_TYPE_RULES:
                if requirements is not None:
                    break
                if any(kw in lower for kw in keywords):
                    requirements = reqs
                    break
            if not requirements:
                requirements = self._DEFAULT_COVERAGE_RULE
            tc_texts = []
            for tc in tcs:
                if not isinstance(tc, dict):
                    continue
                scenario = tc.get('scenario') or tc.get('description') or tc.get('title') or ''
                expected = tc.get('expected_result') or tc.get('then') or ''
                tc_texts.append(f"{scenario} {expected}".lower())
            missing_labels = [
                label for label, kw_list in requirements
                if not any(any(kw in t for kw in kw_list) for t in tc_texts)
            ]
            if missing_labels:
                missing.append(f"{base_name}: thiếu loại kịch bản — {', '.join(missing_labels)}")
        return missing

    def _required_scenarios_hint(self, lower_module_name: str) -> str:
        """Gợi ý danh sách kịch bản bắt buộc cho 1 loại chức năng, dùng để prompt
        AI biết CHÍNH XÁC case nào còn thiếu khi bổ sung TC — tránh việc AI
        chỉ lặp lại thêm 1 biến thể của case đã có sẵn (VD thêm TC 'thành
        công' khác thay vì TC 'có dữ liệu chưa lưu' còn thiếu)."""
        lower = lower_module_name
        if "quay lại" in lower:
            return "(1) chưa thay đổi dữ liệu — về danh sách trước đó, không cảnh báo (2) đã thay đổi dữ liệu — về danh sách trước đó, hiện thông báo thành công (3) từ màn hình chi tiết/sửa — đúng về màn hình vừa click (4) không có màn hình trước — về trang mặc định/không hành động"
        if "phân trang" in lower:
            return "(1) Next (2) Prev (3) trang cuối/Last (4) trang đầu/First"
        if self._is_popup_action_module_4tc(lower):
            if "sinh mã" in lower:
                return "(1) sinh mã thành công (2) mã sinh không trùng và đúng định dạng (3) nhấn nhiều lần xử lý hợp lý (4) lỗi sinh mã"
            if "hủy bỏ" in lower or "huỷ bỏ" in lower:
                return "(1) hủy khi chưa nhập dữ liệu (2) hủy khi đã nhập dữ liệu (3) không lưu dữ liệu (4) popup đóng đúng"
            if "đóng popup" in lower:
                return "(1) đóng khi chưa thay đổi (2) đóng khi đã nhập dữ liệu (3) không lưu dữ liệu (4) quay lại đúng màn hình trước"
        if self._is_them_moi_module(lower) and "tiếp tục" not in lower:
            return "(1) thêm thành công (2) tự sinh mã nếu có (3) Required/Whitespace (4) Format (5) Duplicate/Length/Boundary (6) Dependency hoặc Security/System phù hợp; gộp biến thể cùng bản chất, mục tiêu 6–8 TC"
        if lower.startswith(("cập nhật", "chỉnh sửa")) or lower in {"sửa", "update", "edit"}:
            return "(1) mở đúng bản ghi (2) cập nhật thành công (3) không thay đổi (4) Required/Whitespace (5) Format (6) Duplicate/Length/Boundary (7) Dependency/Security/Permission/System phù hợp; mục tiêu 7–9 TC"
        if lower.startswith(("xóa", "xoá")) or lower == "delete":
            return "(1) mở xác nhận đúng đối tượng (2) xóa thành công (3) hủy/đóng (4) ràng buộc tham chiếu (5) không quyền/không tồn tại/lỗi hệ thống; mục tiêu 5–6 TC"
        return "(1) thành công (2) không thành công [gộp mọi biến thể lỗi]"
    def _get_max_tokens(self) -> int:
        """Trả về max_tokens phù hợp theo model đang dùng."""
        model = self.model.lower()
        if 'gpt-4o' in model:
            return 16000
        if 'gpt-4-turbo' in model or 'gpt-4-1106' in model:
            return 8000
        if 'gpt-3.5' in model:
            return 4000
        return 8000
    _OCR_PROMPT = (
        "Sao chép CHÍNH XÁC tất cả text trên ảnh giao diện, trái→phải, trên→dưới.\n"
        "Định dạng mỗi dòng: VÙNG | TEXT  (VÙNG: TITLE/NAV/TOOLBAR/FORM/TABLE/FOOTER/MODAL/MENU)\n\n"
        "Cần đọc: tiêu đề, breadcrumb, tab, label ô nhập (kể cả placeholder), "
        "text nút, tiêu đề cột, dấu * cạnh label, option dropdown đang mở, "
        "menu ngữ cảnh, text phân trang, badge/tag.\n\n"
        "ĐẶC BIỆT — Nút ở khu vực TOOLBAR (góc trên cùng màn hình, ngang hàng "
        "với tiêu đề trang):\n"
        "PHẢI liệt kê RIÊNG TỪNG nút thấy ở đây, dù chỉ có icon + chữ ngắn. "
        "Đây là các nút HAY BỊ BỎ SÓT nhất nên phải quét kỹ. CHỈ ghi phần TEXT "
        "của nút, KHÔNG ghi kèm ký tự icon/mũi tên (+, ←, →...) đứng trước:\n"
        "  TOOLBAR | BUTTON: Quay lại\n"
        "  TOOLBAR | BUTTON: Thêm mới\n\n"
        "ĐẶC BIỆT — Nút 'Tìm'/'Search' cạnh ô tìm kiếm:\n"
        "Nếu có ô input tìm kiếm VÀ một nút riêng (icon kính lúp + chữ 'Tìm' "
        "hoặc chỉ icon) nằm NGAY CẠNH ô đó để bấm thực hiện tìm kiếm, đây là "
        "2 element HOÀN TOÀN KHÁC NHAU — PHẢI ghi cả ô input và nút riêng:\n"
        "  TOOLBAR | INPUT: [placeholder ô tìm kiếm]\n"
        "  TOOLBAR | BUTTON: Tìm\n\n"
        "ĐẶC BIỆT — MODAL/POPUP/FORM:\nNếu ảnh là modal/popup/form, PHẢI chép RIÊNG đầy đủ mọi thành phần tương tác:\n  MODAL | ICON_BUTTON: Đóng popup (X)\n  FORM | LABEL: [tên trường và dấu * nếu có]\n  FORM | INPUT: [placeholder]\n  FORM | ICON_BUTTON: Sinh mã\n  FOOTER | BUTTON: Hủy bỏ\n  FOOTER | BUTTON: Thêm mới và tiếp tục\n  FOOTER | BUTTON: Thêm mới\nKhông được chỉ chép tiêu đề \"THÊM MỚI\" rồi bỏ sót các trường/nút phía dưới.\n\nĐẶC BIỆT — Cột 'Thao tác' trong bảng:\n"
        "Nếu thấy cột có tiêu đề 'Thao tác', 'Action', 'Hành động' hoặc cột cuối chứa các icon nhỏ, "
        "PHẢI liệt kê TỪNG icon THỰC SỰ NHÌN THẤY trên ảnh theo màu sắc, dùng "
        "định dạng: 'TABLE | THAO_TAC_ICON: [mô tả icon] màu [màu]'.\n"
        "Danh sách dưới đây CHỈ LÀ THAM KHẢO các loại icon hay gặp để bạn biết "
        "cách gọi tên, KHÔNG PHẢI checklist bắt buộc — chỉ ghi dòng nào icon đó "
        "THẬT SỰ XUẤT HIỆN trên ảnh, các loại không thấy thì KHÔNG được ghi:\n"
        " bút/pencil màu vàng → 'Cập nhật'\n"
        " thùng rác/trash màu đỏ → 'Xóa'\n"
        " mắt/eye màu xanh lá → 'Xem chi tiết'\n"
        " file Excel màu xanh lá → 'Xuất file Excel'\n"
        " file Word màu xanh dương → 'Xuất file Word'\n"
        " mũi tên xuống/download → 'Tải xuống'\n"
        " CẤM TUYỆT ĐỐI: tự suy diễn hoặc 'đoán' rằng còn icon nào khác có thể "
        "có nhưng không thấy rõ. Nếu một dòng đã đếm thấy 2 icon trên ảnh (vd chỉ "
        "bút vàng + thùng đỏ) thì CHỈ ghi đúng 2 dòng THAO_TAC_ICON cho cả bảng, "
        "không được thêm icon thứ 3, thứ 4 nào khác dù 'thường thấy' ở các bảng "
        "tương tự. Ví dụ ảnh chỉ có bút vàng + thùng đỏ thì CHỈ ghi:\n"
        "  TABLE | THAO_TAC_ICON: bút/pencil màu vàng\n"
        "  TABLE | THAO_TAC_ICON: thùng rác/trash màu đỏ\n"
        "(KHÔNG thêm dòng mắt/Excel/Word/download nếu không thấy trên ảnh.)\n"
        "Nếu icon không có text nhưng có màu đặc trưng → ghi màu và hình dạng icon.\n"
        " Breadcrumb (vd 'Trang chủ > Quản lý danh mục > ...') CHỈ ghi 1 dòng "
        "NAV để tham khảo ngữ cảnh, KHÔNG được dùng làm tên chức năng/chức năng cần "
        "test — đây là điều hướng, không phải hành động người dùng thao tác.\n"
        "Không cần: phân loại element, giải thích.\n"
        "Bắt đầu ngay, không preamble."
    )
    _ANALYZE_PROMPT_TEMPLATE = (
        "Chuyển text UI dưới đây thành danh sách element có cấu trúc.\n\n"
        "=== TEXT ===\n{raw_text}\n\n"
        "Dòng đầu (BẮT BUỘC): PROJECT_NAME: [tên trang/modal]\n"
        "Các dòng sau: - [Tên label] | [loại] | [chi tiết]\n\n"
        "Phân loại:\n"
        "  TITLE/NAV/breadcrumb          → title/breadcrumb\n"
        "  Label có *                    → input | required\n"
        "  Label không *                 → input | optional\n"
        "  Placeholder 'Nhập/Tìm kiếm'  → input | placeholder\n"
        "  Nút hành động                 → button\n"
        "  Nút nhỏ/icon (X, sinh mã...) → icon-button\n"
        "  Tiêu đề cột bảng             → table | columns=[...]\n"
        "  Select/dropdown               → dropdown | options=[...]\n"
        "  ««,«,»,»»                     → pagination | first/prev/next/last\n"
        "  'Hiển thị X trên tổng số Y'  → pagination-info | readonly\n"
        "  Số dòng/trang                 → page-size-dropdown | options=15,25,50\n"
        "  Menu item thường              → button | menu-item\n"
        "  Menu item đỏ (xóa)           → button | menu-item danger\n"
        "  Menu item vàng (cập nhật)    → button | menu-item warning\n"
        "  Menu item xanh lá (xem)      → button | menu-item view\n"
        "  Menu item xanh dương (xuất)  → button | menu-item export\n"
        "  Badge/tag                     → badge | [giá trị]\n\n"
        "Quy tắc đặc biệt:\n"
        "- Label có * → bỏ dấu * khỏi tên, ghi required. VD: 'Tên chu kỳ *' → 'Tên chu kỳ | input | required'\n"
        "- Với MODAL/POPUP/FORM: giữ RIÊNG từng field và từng nút footer; nút X phải thành '- Đóng popup | icon-button | action=close'.\n- Không gộp \"Mã kho\", \"Tên kho\", \"Sinh mã\", \"Thêm mới và tiếp tục\", \"Hủy bỏ\", \"Đóng popup\" vào chức năng \"Thêm mới\".\n- 'Sinh mã' gần trường → thêm: '- Sinh mã | icon-button | generate-code'\n"
        "- 'Tìm kiếm theo [X]' (ô lọc) ≠ 'Tìm' (nút) — 2 element riêng\n"
        "- Khi gặp dòng THAO_TAC_ICON từ OCR, chuyển thành element riêng:\n"
        "  bút/pencil vàng  → '- Cập nhật | icon-button | action=edit color=yellow'\n"
        "  thùng rác/trash đỏ → '- Xóa | icon-button | action=delete color=red'\n"
        "  mắt/eye xanh lá  → '- Xem chi tiết | icon-button | action=view color=green'\n"
        "  file Excel xanh lá → '- Xuất file Excel | icon-button | action=export_excel color=green'\n"
        "  file Word xanh dương → '- Xuất file Word | icon-button | action=export_word color=blue'\n"
        "  tải xuống/download → '- Tải xuống | icon-button | action=download'\n"
        "- Nếu thấy cột 'Thao tác' nhưng OCR không có dòng THAO_TAC_ICON thì KHÔNG tự đoán icon; chỉ giữ dòng table để tránh sinh chức năng không tồn tại.\n"
        "- Icon cột Thao tác: bút vàng → 'Cập nhật | icon-button | action=edit' ; "
        "thùng đỏ → 'Xóa | icon-button | action=delete' ; "
        "mắt → 'Xem chi tiết | icon-button | action=view' ; "
        "Excel xanh lá → 'Xuất file Excel | icon-button | action=export_excel' ; "
        "Word xanh dương → 'Xuất file Word | icon-button | action=export_word'\n"
        " TUYỆT ĐỐI KHÔNG liệt kê GIÁ TRỊ DỮ LIỆU của từng dòng trong bảng "
        "(vd mã cụ thể 'KHO-VTNN', tên cụ thể 'Kho đồng hồ nước'...) thành "
        "element riêng — đó là dữ liệu mẫu có sẵn, KHÔNG phải chức năng UI.\n"
        " TUYỆT ĐỐI KHÔNG tách từng TÊN CỘT tĩnh (STT, Mã, Tên, Ngày tạo...) "
        "thành các dòng '- [tên cột] | table | ...' riêng biệt — đây là lỗi "
        "hay gặp, PHẢI tránh. Toàn bộ tiêu đề cột bảng (kể cả cột 'Thao tác') "
        "CHỈ được ghi GỘP CHUNG trong DUY NHẤT 1 dòng 'table | columns=[...]'.\n"
        "  SAI (KHÔNG làm thế này):\n"
        "  - STT | table | columns=[STT, Mã, Tên, Thao tác]\n"
        "  - Mã | table | columns=[STT, Mã, Tên, Thao tác]\n"
        "  - Tên | table | columns=[STT, Mã, Tên, Thao tác]\n"
        "  - Thao tác | table | columns=[STT, Mã, Tên, Thao tác]\n"
        "  ĐÚNG (chỉ 1 dòng duy nhất):\n"
        "  - Bảng danh sách | table | columns=[STT, Mã, Tên, Thao tác]\n"
        "Mô tả màn hình: {description}\n\n"
        "Chỉ trả về PROJECT_NAME và danh sách '- ...' — không giải thích."
    )
    def _ocr_raw_text(self, image_blocks: list[dict]) -> str:
        """Lượt 1 — Vision API: chép text thô từ ảnh, không phân loại."""
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": [
                    {"type": "text", "text": self._OCR_PROMPT},
                    *image_blocks,
                ]}],
                temperature=0,
                max_tokens=3000,
            )
            return response.choices[0].message.content or ""
        except Exception:
            return ""
    def _collapse_duplicate_table_lines(self, scanned: str) -> str:
        """
        Lưới an toàn: nếu model vẫn lỡ tách mỗi tên cột tĩnh thành 1 dòng
        '- [tên cột] | table | columns=[...]' riêng biệt (lỗi hay gặp dù đã
        dặn trong prompt), gộp các dòng có CÙNG 'columns=[...]' lại thành 1
        dòng duy nhất, giữ lại dòng đầu tiên làm đại diện.
        """
        if not scanned:
            return scanned
        lines = scanned.splitlines()
        seen_columns: set[str] = set()
        result_lines: list[str] = []
        for line in lines:
            m = re.search(r'columns=\[([^\]]*)\]', line, re.IGNORECASE)
            if m and '| table' in line.lower():
                cols_key = m.group(1).strip().lower()
                if cols_key in seen_columns:
                    continue 
                seen_columns.add(cols_key)
            result_lines.append(line)
        return '\n'.join(result_lines)

    def _analyze_elements(self, raw_text: str, description: str) -> str:
        """Lượt 2 — Text-only: phân loại element từ raw text OCR."""
        prompt = self._ANALYZE_PROMPT_TEMPLATE.format(
            raw_text=raw_text, description=description
        )
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0,
                max_tokens=5000,
            )
            content = response.choices[0].message.content or ""
            return self._collapse_duplicate_table_lines(content)
        except Exception:
            return ""

    def _scan_image_elements(self, image_blocks: list[dict], description: str) -> str:
        """
        Chạy OCR pipeline 2 lượt.
        Lượt 1 lỗi → fallback 1-lượt (Vision + phân loại cùng lúc).
        Lượt 2 lỗi → trả raw text lượt 1 (vẫn parse được PROJECT_NAME).
        """
        raw = self._ocr_raw_text(image_blocks)
        if not raw:
            return self._scan_image_elements_legacy(image_blocks, description)
        return self._analyze_elements(raw, description) or raw

    def _scan_image_elements_legacy(self, image_blocks: list[dict], description: str) -> str:
        """Fallback 1-lượt: Vision + phân loại trong cùng 1 API call."""
        prompt = (
            "Nhìn ảnh giao diện, trả về:\n"
            "Dòng 1: PROJECT_NAME: [tên trang/modal]\n"
            "Các dòng sau: - [Tên label] | [loại] | [chi tiết]\n\n"
            "Loại element:\n"
            "  input (required/optional), textarea, dropdown (liệt kê options), datepicker,\n"
            "  button, icon-button, tab, table (liệt kê cột), breadcrumb, badge, notice, title\n\n"
            "Icon cột Thao tác:\n"
            "  bút vàng → Cập nhật | icon-button | action=edit\n"
            "  thùng đỏ → Xóa | icon-button | action=delete\n"
            "  mắt     → Xem chi tiết | icon-button | action=view\n"
            "  Excel   → Xuất file Excel | icon-button | action=export_excel\n"
            "  Word    → Xuất file Word | icon-button | action=export_word\n"
            "  tải xuống → Tải xuống | icon-button | action=download\n\n"
            "Menu ngữ cảnh: liệt kê từng mục — đỏ=danger, vàng=warning, xanh lá=view, xanh dương=export\n"
            "Phân trang: ««,«,»,»» → pagination | first/prev/next/last ; "
            "số dòng → page-size-dropdown | options=15,25,50 ; "
            "'Hiển thị X trên tổng số Y' → pagination-info | readonly\n"
            "Dấu * = required. 'Sinh mã' = icon-button riêng, KHÔNG gộp vào trường Mã.\n"
            "Nếu là modal/popup/form: PHẢI liệt kê riêng từng input, Sinh mã, từng nút footer, Hủy bỏ và nút X đóng popup.\n"
            "'Tìm kiếm theo [X]' ≠ 'Tìm' — 2 element riêng.\n"
            f"Mô tả bổ sung: {description}\n\n"
            "Dùng ĐÚNG text trên UI. Không giải thích."
        )
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": [
                    {"type": "text", "text": prompt},
                    *image_blocks,
                ]}],
                temperature=0,
                max_tokens=3000,
            )
            return response.choices[0].message.content or ""
        except Exception:
            return description or "Phân tích ảnh giao diện"
    _FORM_STRUCTURE_PROMPT_TEMPLATE = (
        "Phân tích danh sách UI ELEMENTS dưới đây (đã quét/phân loại từ ảnh giao "
        "diện) thành CẤU TRÚC FORM có nghĩa nghiệp vụ — hiểu ĐÂY LÀ FORM GÌ, có "
        "FIELD nào, có BUTTON nào, BUSINESS RULE gì, WORKFLOW thao tác ra sao — "
        "TRƯỚC KHI sinh testcase (không chỉ liệt kê lại input/button/icon thô).\n\n"
        "=== UI ELEMENTS ===\n{scanned}\n\n"
        "=== MÔ TẢ BỔ SUNG (nếu có) ===\n{description}\n\n"
        "Trả về DUY NHẤT 1 JSON object đúng cấu trúc sau (không markdown, không "
        "giải thích thêm):\n"
        "{{\n"
        '  "form_name": "<tên form/màn hình, vd \\"THÊM MỚI\\">",\n'
        '  "fields": [{{"name": "<tên field đúng theo UI>", "type": '
        '"input|textarea|dropdown|datepicker|checkbox|radio", "required": true|false}}],\n'
        '  "buttons": [{{"name": "<tên button/icon-button đúng theo UI>", "action": '
        '"submit|submit_continue|cancel|close|generate-code|<action khác phù hợp>"}}],\n'
        '  "business_rules": ["<mô tả ngắn 1 quy tắc nghiệp vụ, vd \\"Mã kho để trống '
        '→ hệ thống tự sinh\\">", "..."],\n'
        '  "workflow": ["<bước 1>", "<bước 2>", "..."]\n'
        "}}\n\n"
        "QUY TẮC BẮT BUỘC:\n"
        "- fields CHỈ liệt kê input/textarea/dropdown/datepicker/checkbox/radio "
        "THỰC SỰ có trong UI ELEMENTS — KHÔNG bịa field không tồn tại.\n"
        "- buttons CHỈ liệt kê button/icon-button THỰC SỰ có trong UI ELEMENTS "
        "(kể cả nút X đóng popup, Sinh mã).\n"
        "- business_rules: suy luận HỢP LÝ từ field required/optional và tên field "
        "(vd field tên có \"Mã\" thường có quy tắc \"để trống → tự sinh\"/\"không được "
        "trùng\"; field required luôn có quy tắc \"bắt buộc nhập\") — CHỈ áp dụng cho "
        "field/ngữ cảnh thực sự có, KHÔNG suy diễn quy tắc không liên quan.\n"
        "- workflow: liệt kê ĐÚNG thứ tự thao tác người dùng thực hiện trên form "
        "này, từ lúc mở đến lúc đóng (vd: Mở form → Nhập dữ liệu → Sinh mã (nếu "
        "cần) → Nhấn nút lưu → Lưu thành công → Đóng form).\n"
        "Chỉ trả JSON, không markdown, không giải thích thêm."
    )
    def _analyze_form_structure(self, scanned: str, description: str) -> dict:
        """
        FORM UNDERSTANDING — chạy TRƯỚC khi sinh testcase: phân tích `scanned`
        (đã OCR + phân loại) thành cấu trúc Form có nghĩa nghiệp vụ (form_name/
        fields/buttons/business_rules/workflow), để AI hiểu "đây là Form gì"
        trước khi quyết định chức năng nào cần sinh — thay vì chỉ OCR ra input/
        button/icon rời rạc rồi suy chức năng trực tiếp từ đó.
        CHỈ chạy khi `scanned` có dấu hiệu là FORM/MODAL thật sự (có ít nhất 1
        dòng input/textarea/dropdown/datepicker) — màn hình danh sách thuần
        (không field nhập liệu) không cần bước này, tránh tốn thêm 1 lượt gọi
        API không cần thiết cho các màn hình không phải Form.
        Trả về {} nếu không đủ điều kiện, lỗi gọi API, hoặc JSON không hợp lệ —
        KHÔNG raise, không làm gãy pipeline (cùng nguyên tắc với RAG/Rule Engine
        khác trong file này: lỗi ở bước bổ trợ không được làm hỏng luồng chính).
        """
        if not scanned:
            return {}
        has_field = bool(re.search(
            r'\|\s*(input|textarea|dropdown|datepicker)\b', scanned, re.IGNORECASE
        ))
        if not has_field:
            return {}
        prompt = self._FORM_STRUCTURE_PROMPT_TEMPLATE.format(
            scanned=scanned, description=description or ''
        )
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0,
                max_tokens=2000,
                response_format={"type": "json_object"},
            )
            content = response.choices[0].message.content or ""
            data = self._parse_json_response(content)
        except Exception as exc:
            print(f"[FormStructure] Bỏ qua bước Form Understanding vì lỗi: {exc}")
            return {}

        if not isinstance(data, dict):
            return {}
        data['form_name'] = data.get('form_name') or ''
        data['fields'] = data.get('fields') if isinstance(data.get('fields'), list) else []
        data['buttons'] = data.get('buttons') if isinstance(data.get('buttons'), list) else []
        data['business_rules'] = (
            data.get('business_rules') if isinstance(data.get('business_rules'), list) else []
        )
        data['workflow'] = data.get('workflow') if isinstance(data.get('workflow'), list) else []

        print(
            "=== FORM STRUCTURE (Form Understanding) ===\n"
            f"form_name: {data['form_name']}\n"
            f"fields: {[f.get('name') if isinstance(f, dict) else f for f in data['fields']]}\n"
            f"buttons: {[b.get('name') if isinstance(b, dict) else b for b in data['buttons']]}\n"
            f"business_rules: {data['business_rules']}\n"
            f"workflow: {data['workflow']}\n"
            "============================================"
        )
        return data

    def _build_form_structure_hint(self, form_structure: dict) -> str:
        """
        Chuyển Form Structure đã phân tích (form_name/fields/buttons/
        business_rules/workflow) thành block hint nhúng vào prompt sinh
        testcase — giúp AI hiểu "đây là Form gì/field nào/button nào/business
        rule gì/workflow ra sao" TRƯỚC KHI sinh testcase, thay vì chỉ dựa vào
        danh sách input/button/icon rời rạc.
        """
        if not isinstance(form_structure, dict) or not form_structure:
            return ""
        fields = form_structure.get('fields') or []
        buttons = form_structure.get('buttons') or []
        rules = form_structure.get('business_rules') or []
        workflow = form_structure.get('workflow') or []
        if not fields and not buttons:
            return ""
        lines = ["\n=== FORM STRUCTURE ĐÃ PHÂN TÍCH (Form Understanding) ==="]
        if form_structure.get('form_name'):
            lines.append(f'Form Name: {form_structure["form_name"]}')
        if fields:
            lines.append("Fields:")
            for f in fields:
                if not isinstance(f, dict):
                    continue
                name = f.get('name', '')
                ftype = f.get('type', 'input')
                req = 'required' if f.get('required') else 'optional'
                lines.append(f'  - {name} | type={ftype} | {req}')
        if buttons:
            lines.append("Buttons:")
            for b in buttons:
                if not isinstance(b, dict):
                    continue
                name = b.get('name', '')
                action = b.get('action', '')
                lines.append(f'  - {name} | action={action}')
        if rules:
            lines.append("Business Rules:")
            for r in rules:
                lines.append(f'  - {r}')
        if workflow:
            lines.append("Workflow: " + " → ".join(str(s) for s in workflow))
        lines.append(
            "QUY TẮC BẮT BUỘC KHI SINH CHỨC NĂNG:\n"
            "- CÁC FIELD liệt kê ở trên (Fields) TUYỆT ĐỐI KHÔNG được tạo chức năng "
            "riêng theo tên field — test case của từng field (Validation/Boundary/"
            "Business Rule/XSS/SQL Injection/Required/Length/Duplicate/Whitespace) "
            "PHẢI nằm trong chức năng hành động LƯU tương ứng (vd \"Thêm mới\"/\"Cập "
            "nhật\"), KHÔNG tách field thành chức năng con.\n"
            "- MỖI BUTTON liệt kê ở trên (Buttons) là 1 chức năng RIÊNG (vd \"Sinh mã\", "
            "\"Thêm mới\", \"Thêm mới và tiếp tục\", \"Hủy bỏ\", \"Đóng popup\").\n"
            "- Business Rules ở trên PHẢI được thể hiện trong test case của chức năng "
            "lưu tương ứng.\n"
            "- Workflow ở trên PHẢI được thể hiện đúng thứ tự trong scenario của TC "
            "\"thành công\" thuộc chức năng lưu chính."
        )
        return "\n".join(lines)
    def _split_elements(self, scanned: str) -> tuple[str, str]:
        """Chia danh sách elements thành 2 nhóm theo loại UI để batch xử lý."""
        lines = [l for l in scanned.splitlines() if l.strip().startswith('-')]
        G1 = ('input', 'button', 'breadcrumb', 'title', 'heading', 'datepicker', 'pagination', 'page-size')
        G2 = ('table', 'icon', 'tab', 'badge', 'dropdown', 'checkbox', 'radio')
        g1, g2, ungrouped = [], [], []
        for line in lines:
            low = line.lower()
            if any(k in low for k in G1):
                g1.append(line)
            elif any(k in low for k in G2):
                g2.append(line)
            else:
                ungrouped.append(line)

        for i, line in enumerate(ungrouped):
            (g1 if i % 2 == 0 else g2).append(line)

        if not g1 or not g2:
            mid = max(1, len(lines) // 2)
            g1, g2 = lines[:mid], lines[mid:]

        return '\n'.join(g1), '\n'.join(g2)
    def _call_api(self, prompt: str, image_blocks: list[dict] | None, system_prompt: str, is_full_scan: bool = False) -> dict | None:
        """Gọi API 1 lần với retry, trả về dict đã unwrap hoặc None nếu lỗi."""
        messages = self._build_messages(prompt, image_blocks=image_blocks, system_prompt=system_prompt, is_full_scan=is_full_scan)
        for attempt in range(self.max_retries):
            try:
                response = self.client.chat.completions.create(
                    model=self.model,
                    messages=messages,
                    temperature=0.1,
                    top_p=0.95,
                    max_tokens=self._get_max_tokens(),
                    response_format={"type": "json_object"},
                )
                if response.choices[0].finish_reason == 'length' and attempt < self.max_retries - 1:
                    time.sleep(2 ** attempt)
                    continue
                raw = self._parse_json_response(response.choices[0].message.content)
                result = self._unwrap_modules(raw)
                result = self._remove_static_modules(result)
                return result
            except Exception:
                if attempt < self.max_retries - 1:
                    time.sleep(2 ** attempt)
        return None
    def _generate_batch(
        self,
        description: str,
        elements_text: str,
        project_name: str,
        image_blocks: list[dict] | None,
        id_offset: int = 0,
        system_prompt: str | None = None,
        domain: str | None = None,
    ) -> dict:
        if system_prompt is None:
            system_prompt = self._build_system_prompt(SYSTEM_PROMPT_FULL)
        """Sinh TC cho 1 nhóm elements, kèm ảnh để AI đọc trực tiếp UI."""
        is_modal = any(
            k in elements_text.lower()
            for k in ('modal', 'popup', 'dialog', 'hủy bỏ', 'thêm mới và tiếp tục', 'sinh mã')
        )
        is_dropdown_menu = any(
            k in elements_text.lower()
            for k in ('menu-item', 'context menu', 'dropdown menu', 'xem & kiểm kê',
                      'sửa kế hoạch', 'xóa đợt kiểm kê', 'xuất file excel', 'xuất file word')
        )
        has_action_icons = any(
            k in elements_text.lower()
            for k in ('action=edit', 'action=delete', 'action=view',
                      'action=export_excel', 'action=export_word', 'action=download',
                      'icon-button | action')
        )
        modal_note = (
            "\nLƯU Ý: Đây là màn hình MODAL/POPUP. Áp dụng quy tắc:\n"
            "- Tạo chức năng riêng cho nút X đóng modal\n"
            "- Tạo chức năng riêng cho TỪNG nút (Hủy bỏ, Thêm mới, Thêm mới và tiếp tục)\n"
            "- KHÔNG tạo chức năng riêng theo tên field (vd \"Mã\", \"Tên\", \"Ghi chú\") — "
            "toàn bộ TC Required/Length/Whitespace/Validation/Boundary/Duplicate/XSS/"
            "SQL Injection của các field này PHẢI nằm trong chức năng \"Thêm mới\", TÁCH "
            "THEO NHÓM NGHIỆP VỤ (thành công / tự sinh mã nếu có / thiếu trường bắt "
            "buộc / trùng-sai định dạng-vượt độ dài / XSS-SQL Injection), mỗi TC chỉ "
            "mô tả 1 nhóm lỗi liên quan, KHÔNG nhồi tất cả vào 1 TC\n"
            "- Trường Sinh mã (icon riêng cạnh field): chức năng \"Sinh mã\" RIÊNG BIỆT, "
            "với TC tự sinh mã, TC nhập thủ công, TC để trống\n"
        ) if is_modal else ""

        dropdown_menu_note = (
            "\nLƯU Ý: Ảnh chứa DROPDOWN MENU / CONTEXT MENU. Áp dụng quy tắc:\n"
            "- Tạo chức năng riêng cho TỪNG mục trong menu (mỗi menu-item = 1 chức năng)\n"
            "- Đọc ĐÚNG text hiển thị trên mục menu, không tự đổi tên\n"
            "- Mục 'Xem & Kiểm kê' → chức năng 'Xem & Kiểm kê'\n"
            "- Mục 'Sửa kế hoạch' → chức năng 'Sửa kế hoạch'\n"
            "- Mục 'Xuất file Excel' → chức năng 'Xuất file Excel'\n"
            "- Mục 'Xuất file Word' → chức năng 'Xuất file Word'\n"
            "- Mục màu đỏ/destructive (vd 'Xóa đợt kiểm kê') → chức năng riêng, ưu tiên TC confirm xóa, cancel xóa, xóa thành công, quyền hạn\n"
            "- KHÔNG gộp các mục menu thành 1 chức năng\n"
        ) if is_dropdown_menu else ""
        action_icon_note = (
            "\nLƯU Ý: Danh sách elements chứa ICON-BUTTON trong cột THAO TÁC. Áp dụng quy tắc:\n"
            "- Tạo chức năng riêng cho TỪNG icon (KHÔNG gộp thành chức năng 'Thao tác' chung)\n"
            "- action=edit (bút vàng) → chức năng 'Cập nhật' ĐÚNG 2 TC (KHÔNG tách lẻ): (1) thành công — mở form đúng data dòng, sửa dữ liệu hợp lệ→lưu OK, hiển thị đúng tại danh sách | (2) không thành công — GỘP mọi biến thể lỗi (trống bắt buộc, sai định dạng, XSS/SQLi, >256 ký tự) vào 1 TC DUY NHẤT→lỗi phù hợp, không lưu\n"
            "- action=delete (thùng đỏ) → chức năng 'Xóa' với đủ TC: confirm xóa, hủy xóa, xóa bản ghi tham chiếu, phân quyền\n"
            "- action=view (mắt xanh lá) → chức năng 'Xem chi tiết' với đủ TC: mở đúng thông tin, đóng, phân quyền\n"
            "- action=export_excel (Excel xanh lá) → chức năng 'Xuất file Excel' với TC: xuất đúng file, phân quyền\n"
            "- action=export_word (Word xanh dương) → chức năng 'Xuất file Word' với TC: xuất đúng file, phân quyền\n"
            "Chức năng 'Thao tác' (không rõ icon) = LỖI. Phải tách ra từng chức năng theo action.\n"
        ) if has_action_icons else ""
        _allowed_module_names = []
        for _line in elements_text.splitlines():
            _line = _line.strip()
            if _line.startswith('-'):
                _label = _line.lstrip('- ').split('|')[0].strip()
                if _label:
                    _allowed_module_names.append(_label)
        _allowed_list_str = (
            '\n'.join(f'  • {n}' for n in _allowed_module_names)
            if _allowed_module_names else '  (xem danh sách elements bên trên)'
        )
        batch_desc = (
            f"Sinh testcase cho các UI elements sau của màn hình \"{project_name}\".\n"
            f"{modal_note}"
            f"{dropdown_menu_note}"
            f"{action_icon_note}\n"
            f"=== UI ELEMENTS CẦN TẠO MODULE (DANH SÁCH ĐẦY ĐỦ, KHÔNG ĐƯỢC THÊM) ===\n"
            f"{elements_text}\n\n"
            f" TUYỆT ĐỐI CẤM:\n"
            f"- KHÔNG tạo module cho bất kỳ chức năng nào NGOÀI danh sách trên.\n"
            f"- KHÔNG tự thêm module vì thấy trong ảnh hoặc suy diễn từ context.\n"
            f"- KHÔNG tạo module trùng lặp cho cùng 1 chức năng.\n"
            f"- 'Tìm kiếm', 'Tìm', 'Tìm kiếm theo mã' là các module KHÁC NHAU, KHÔNG gộp chung.\n"
            f"Danh sách tên module được phép tạo:\n{_allowed_list_str}\n\n"
            f"YÊU CẦU BẮT BUỘC:\n"
            f"- Chỉ tạo module cho chức năng được người dùng yêu cầu.\n"
            f"- Không tạo module dư từ ảnh giao diện.\n"
            f"- Mỗi module PHẢI có đủ: Positive + Negative + Validation + Boundary TC\n"
            f"- Module trường nhập liệu: tối thiểu 4 TC (hợp lệ, trống, nhập 256 ký tự → lỗi, XSS)\n"
            f"- Module nút hành động: tối thiểu 3 TC\n"
            f"- Module Cập nhật ('Cập nhật'/'Sửa'/'Chỉnh sửa'): ĐÚNG 2 TC (KHÔNG tách lẻ, KHÔNG tạo thêm TC thứ 3): "
            f"(1) thành công — sửa dữ liệu hợp lệ→lưu OK | (2) không thành công — GỘP mọi biến thể lỗi "
            f"(trống bắt buộc, sai định dạng, XSS/SQLi, >256 ký tự, khoảng trắng) vào 1 TC DUY NHẤT→lỗi, không lưu\n"
            f"- Module tìm kiếm ('Tìm kiếm'/'Tìm kiếm theo [X]'/'Tìm'): ĐÚNG 2 TC (thành công / không thành công — gộp \"không có dữ liệu\" và \"bỏ trống\" vào 1 TC \"không thành công\"), KHÔNG tạo thêm TC thứ 3\n"
            f"- Module Cập nhật và module Tìm kiếm là 2 chức năng KHÁC NHAU: KHÔNG được để nội dung 'scenario'/'expected_result' của 2 module này trùng hoặc gần giống nhau (mỗi module viết theo đúng ngữ cảnh thao tác của nó — Cập nhật là sửa dữ liệu dòng đã chọn, Tìm kiếm là lọc/tra cứu danh sách).\n"
            f"- TRƯỜNG BẮT BUỘC (có 'required'): BẮT BUỘC có TC để trống → báo lỗi\n"
            f"- TRƯỜNG KHÔNG BẮT BUỘC (có 'optional'): có TC để trống → lưu thành công\n"
            f"- Mỗi trường chỉ cần 1 câu ngắn, steps tối đa 3 bước\n"
            f"- id testcase bắt đầu từ TC_{id_offset+1:03d} và tăng dần liên tục\n"
            f"- project_name trong JSON = \"{project_name}\"\n"
            f"- test_data phải cụ thể, thực tế (vd: mã 'CK-001', tên 'Chu kỳ tháng', mô tả 'Test mô tả')\n"
            f"QUAN TRỌNG VỀ VĂN PHONG: 'scenario' và 'expected_result' PHẢI viết "
            f"ĐÚNG theo mẫu câu chữ đã quy định trong system prompt (mục === CÁCH VIẾT "
            f"TÌNH HUỐNG KIỂM ĐỊNH === và === CÁCH VIẾT KẾT QUẢ MONG ĐỢI ===) và bám sát "
            f"các ví dụ thực tế WEB2519 đã cho — KHÔNG tự ý diễn đạt lại ngắn gọn hay "
            f"khác đi. Ví dụ: TC thêm mới thành công bắt buộc có expected_result chứa "
            f"câu 'Xuất hiện thông báo \"Thêm thành công\"...', không được viết tắt thành "
            f"câu khác như 'Thông tin được lưu...' (thiếu phần thông báo).\n"
            f"Mô tả màn hình: {description}"
            f"{self._build_rule_engine_hints(elements_text)}"
            f"{self._get_domain_hints(domain, description, project_name, elements_text)}"
        )
        batch_rag_query = (
            f"Mô tả màn hình:\n{description}\n\n"
            f"Tên dự án/màn hình:\n{project_name}\n\n"
            f"UI Elements batch:\n{elements_text}\n\n"
            f"Rule Engine:\n{self._build_rule_engine_hints(elements_text)}\n"
        )
        batch_desc = self._append_rag_context(
            batch_desc,
            self._retrieve_rag_context(batch_rag_query, top_k=5)
        )
        messages = self._build_messages(
            batch_desc,
            previous_test_cases=None,
            image_blocks=image_blocks,
            system_prompt=system_prompt,
        )
        for attempt in range(self.max_retries):
            try:
                response = self.client.chat.completions.create(
                model=self.model,
                messages=messages,
                temperature=0.1,
                top_p=0.9,  
                max_tokens=self._get_max_tokens(),
                response_format={"type": "json_object"},
            )
                finish_reason = response.choices[0].finish_reason
                if finish_reason == 'length' and attempt < self.max_retries - 1:
                    time.sleep(2 ** attempt)
                    continue
                content = response.choices[0].message.content
                raw = self._parse_json_response(content)
                result = self._unwrap_modules(raw)
                result = self._remove_static_modules(result)
                if _allowed_module_names:
                    allowed_desc = ", ".join(_allowed_module_names)
                    result = self._filter_modules_by_description(result, allowed_desc)
                return result
            except Exception:
                if attempt < self.max_retries - 1:
                    time.sleep(2 ** attempt)
                else:
                    raise
    def _is_targeted_request(self, description: str) -> bool:
        """
        Phát hiện user chỉ muốn sinh TC cho một số chức năng cụ thể,
        không phải toàn bộ màn hình.
        Trả về True nếu description chứa từ khoá chỉ định chức năng rõ ràng
        (ví dụ: "tìm kiếm", "nút Tìm", "quay lại", "xóa"...)
        hoặc có dạng liệt kê chức năng ngắn gọn (không phải mô tả hệ thống dài).
        LƯU Ý: trước đây có bug — chỉ cần match 1 targeted_keyword là return
        True ngay (early return), nên mô tả CẢ HỆ THỐNG kiểu "Website ... với
        role Admin và Customer: đăng ký, đăng nhập, tìm kiếm sản phẩm, giỏ
        hàng, thanh toán, quản lý đơn hàng" bị nhận nhầm là targeted chỉ vì
        chứa cụm "tìm kiếm" (trong "tìm kiếm sản phẩm"). Sửa: nếu description
        có dấu hiệu mô tả TOÀN HỆ THỐNG (chứa "role"/"website"/"hệ thống"...)
        thì KHÔNG được coi là targeted dù có chứa targeted_keyword nào.
        """
        if not description or len(description.strip()) < 3:
            return False
        if ':' in description:
            prefix = description.split(':', 1)[0].strip().lower()
            system_prefixes = ('website', 'hệ thống', 'ứng dụng', 'phần mềm', 'trang web')
            if any(k in prefix for k in system_prefixes):
                return False
            after_colon = description.split(':', 1)[1].strip()
            items = [p.strip() for p in re.split(r'[,\n]', after_colon) if p.strip()]
            if items and 1 <= len(items) <= 10 and all(1 <= len(item.split()) <= 5 for item in items):
                return True
        targeted_keywords = [
            'tìm kiếm', 'tìm theo', 'tìm kiếm theo', 'ô tìm', 'nút tìm', 'tìm',
            'quay lại', 'nút quay lại',
            'thêm mới', 'nút thêm', 'thêm và tiếp tục', 'thêm',
            'xóa', 'nút xóa', 'xác nhận xóa',
            'cập nhật', 'chỉnh sửa', 'nút cập nhật', 'sửa',
            'xuất excel', 'xuất word', 'xuất file',
            'phân trang', 'checkbox', 'làm mới',
            'đóng popup', 'nút x', 'hủy bỏ', 'đóng',
            'sinh mã', 'tạo mã',
            'phân quyền', 'phan quyen',
            'chấm công', 'cham cong',
            'chỉ sinh', 'chỉ tạo', 'chỉ cần',
            'cho chức năng', 'cho chức năng', 'cho nút',
            'không cần', 'bỏ qua',
        ]
        system_desc_keywords = [
            'hệ thống', 'màn hình', 'trang', 'giao diện', 'phần mềm',
            'ứng dụng', 'website', 'role',
        ]

        desc_lower = description.lower()
        has_targeted_kw = any(kw in desc_lower for kw in targeted_keywords)
        has_system_keyword = any(kw in desc_lower for kw in system_desc_keywords)
        if not has_targeted_kw:
            return False
        if has_system_keyword:
            return False
        return True
    def _extract_targeted_elements(self, description: str, scanned: str) -> str:
        """
        Từ danh sách scanned elements, lọc ra chỉ những element
        mà user đề cập trong description.
        Trả về chuỗi element đã lọc để truyền vào _generate_batch.

        Thuật toán:
        1. Tách description thành từng "phrase" theo dấu phẩy / xuống dòng
           (vd: "tìm kiếm theo mã hoặc tên chu kỳ, tìm, quay lại"
                → ["tìm kiếm theo mã hoặc tên chu kỳ", "tìm", "quay lại"])
        2. Với mỗi phrase, thử khớp trực tiếp (substring) vào từng dòng element.
        3. Nếu không khớp trực tiếp, dùng keyword_map mở rộng các biến thể
           đồng nghĩa thường gặp (vd: "tìm" cũng khớp nút có chữ "tìm kiếm").
        4. Nếu một phrase hoàn toàn không khớp được element nào, BỎ QUA phrase đó
           (không suy diễn) — không bao giờ tự thêm các chức năng người dùng
           không yêu cầu.
        """
        element_lines = [l for l in scanned.splitlines() if l.strip().startswith('-')]
        phrases = [p.strip() for p in re.split(r'[,\n]', description) if p.strip()]
        if not phrases:
            phrases = [description.strip()]

        synonym_map = {
            'tìm kiếm theo mã': ['tìm theo mã', 'search by code'],
            'tìm kiếm theo tên': ['tìm theo tên', 'search by name'],
            'tìm kiếm': ['ô tìm kiếm', 'search box'],
            'tìm': ['nút tìm', 'button tìm'],
            'quay lại': ['back', 'nút quay lại'],
            'thêm mới': ['thêm', 'add', 'tạo mới'],
            'xóa': ['delete', 'xoá', 'thùng rác', 'delete'],
            'cập nhật': ['sửa', 'chỉnh sửa', 'edit', 'update'],
            'xuất excel': ['excel', 'xuất file'],
            'xuất word': ['word', 'xuất file'],
            'phân trang': ['pagination'],
            'sinh mã': ['tạo mã', 'generate code'],
        }
        canonical_groups = [
            {'tìm kiếm theo mã', 'tìm theo mã', 'tìm kiếm theo mã hoặc kho', 'tìm kiếm theo mã hoặc tên'},
            {'tìm kiếm theo tên', 'tìm theo tên'},
            {'tìm kiếm', 'ô tìm kiếm', 'search'},
            {'tìm', 'nút tìm'},
            {'quay lại', 'back', 'nút quay lại'},
        ]

        matched_lines: list[str] = []
        unmatched_phrases: list[str] = []
        seen_canonical_keys: set[str] = set()

        def _canonical_key(text: str) -> str | None:
            """Trả về key đại diện cho group nếu text thuộc 1 canonical group."""
            for group in canonical_groups:
                if any(g in text for g in group):
                    return '|'.join(sorted(group))
            return None
        for phrase in phrases:
            phrase_lower = phrase.lower()
            clean_phrase = re.sub(r'^(nút|ô|icon|chức năng)\s+', '', phrase_lower).strip()
            group_key = _canonical_key(clean_phrase)
            if group_key and group_key in seen_canonical_keys:
                continue
            candidates = {clean_phrase}
            for key, syns in synonym_map.items():
                if key in clean_phrase or clean_phrase in key:
                    candidates.update(syns)
                    candidates.add(key)
            found_for_phrase = False
            for line in element_lines:
                line_lower = line.lower()
                if any(c in line_lower for c in candidates if c):
                    if line not in matched_lines:
                        matched_lines.append(line)
                    found_for_phrase = True

            if found_for_phrase and group_key:
                seen_canonical_keys.add(group_key)

            if not found_for_phrase:
                unmatched_phrases.append(phrase)

        if matched_lines:
            result = '\n'.join(matched_lines)
            if unmatched_phrases:
                result += (
                    "\n\n(Lưu ý: các yêu cầu sau KHÔNG khớp được element nào trong ảnh, "
                    "vẫn sinh TC theo đúng mô tả người dùng, KHÔNG bỏ qua: "
                    + "; ".join(unmatched_phrases) + ")"
                )
            return result
        return (
            "(Không tìm thấy element khớp trong ảnh quét được. "
            "CHỈ sinh testcase đúng theo mô tả sau, KHÔNG suy diễn thêm "
            "chức năng khác xuất hiện trên ảnh)\n"
            f"- {description}"
        )

    def _should_expand_modal_request(self, description: str, scanned: str, project_name: str) -> bool:
        """Coi yêu cầu ngắn trùng tên modal (vd "Thêm mới") là FULL FORM."""
        if not scanned:
            return False
        desc = re.sub(r"\s+", " ", (description or "").strip().lower())
        project = re.sub(r"\s+", " ", (project_name or "").strip().lower())
        generic_titles = {"thêm mới", "cập nhật", "chỉnh sửa", "sửa", "tạo mới", "popup thêm mới", "form thêm mới", "modal thêm mới"}
        looks_like_screen_title = desc in generic_titles or (desc and project and desc == project)
        if not looks_like_screen_title:
            return False
        interactive = []
        modal_signal = False
        for line in scanned.splitlines():
            low = line.lower().strip()
            if any(k in low for k in ("modal", "popup", "dialog", "đóng popup", "action=close")):
                modal_signal = True
            if line.strip().startswith("-") and any(f"| {kind}" in low for kind in ("input", "textarea", "dropdown", "datepicker", "checkbox", "radio", "button", "icon-button")):
                interactive.append(line)
        if any(k in scanned.lower() for k in ("thêm mới và tiếp tục", "hủy bỏ", "sinh mã", "đóng popup")):
            modal_signal = True
        return modal_signal and len(interactive) >= 3
    _REGENERATE_TC_SCHEMA_FIELDS = (
        "title, scenario, description, given, when, then, precondition, "
        "steps, test_data, expected_result, priority, test_type, "
        "actual_result, status, note"
    )

    def regenerate_single_testcase(
        self, module_name: str, testcase: dict, project_name: str = ""
    ) -> dict:
        """
        Sinh lại ĐÚNG 1 testcase, giữ nguyên id/module/chức năng/feature —
        chỉ nội dung nghiệp vụ (title/scenario/expected_result/...) được
        AI viết lại. Raise RuntimeError nếu gọi AI thất bại hoặc AI trả
        sai định dạng — caller (app.py) chịu trách nhiệm giữ nguyên
        testcase cũ khi có lỗi (không được xoá dữ liệu đang có).
        """
        module_name = (module_name or "").strip()
        old_id = (testcase or {}).get("id") or ""
        prompt = (
            "Bạn chỉ được sinh lại ĐÚNG MỘT test case cho chức năng phần mềm sau.\n"
            f"Tên chức năng bắt buộc (không được đổi): {module_name}\n\n"
            "Test case hiện tại (làm cơ sở tham khảo — giữ đúng phạm vi nghiệp vụ "
            "của kịch bản này, có thể viết lại rõ ràng/chi tiết hơn):\n"
            f"{json.dumps(testcase or {}, ensure_ascii=False, indent=2)}\n\n"
            "YÊU CẦU BẮT BUỘC:\n"
            "- Không tạo chức năng khác, không đổi tên module.\n"
            "- Không tạo nhiều test case — chỉ trả đúng MỘT test case.\n"
            "- Giữ đúng phạm vi nghiệp vụ (thành công hay không thành công) của "
            "test case gốc.\n"
            f"- Trả về DUY NHẤT một object JSON với các field: {self._REGENERATE_TC_SCHEMA_FIELDS}.\n"
            "- Không kèm giải thích, không markdown, không bọc trong mảng hay key khác.\n"
        )
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.4,
                max_tokens=1500,
                response_format={"type": "json_object"},
            )
            content = response.choices[0].message.content or ""
            new_tc = self._parse_json_response(content)
        except Exception as exc:
            raise RuntimeError(f"Không sinh lại được test case: {exc}") from exc

        if not isinstance(new_tc, dict):
            raise RuntimeError("AI không trả về đúng định dạng test case")
        new_tc["id"] = old_id
        new_tc["module"] = module_name
        new_tc["chức năng"] = module_name
        new_tc["feature"] = module_name
        new_tc.setdefault("priority", (testcase or {}).get("priority") or "Trung bình")
        new_tc.setdefault("status", "Chưa chạy")
        new_tc.setdefault("test_type", (testcase or {}).get("test_type") or "Kiểm thử chức năng")
        new_tc.setdefault("actual_result", "")
        new_tc.setdefault("note", "")
        return new_tc

    def regenerate_entire_function(
        self, module_name: str, testcases: list, project_name: str = ""
    ) -> list:
        """
        Sinh lại TOÀN BỘ testcase của 1 chức năng — ép tên module về
        module_name cho mọi TC trả về, không đụng chức năng khác. Số
        lượng TC tối thiểu bám theo bộ hiện có (không bắt buộc bằng chính
        xác con số cũ, chỉ không được ít hơn). Raise RuntimeError nếu gọi
        AI thất bại hoặc AI trả sai định dạng.
        """
        module_name = (module_name or "").strip()
        testcases = [tc for tc in (testcases or []) if isinstance(tc, dict)]
        min_count = len(testcases) or 4
        prompt = (
            "Bạn chỉ được sinh lại TOÀN BỘ test case của chức năng phần mềm sau.\n"
            f"Tên chức năng bắt buộc (không được đổi): {module_name}\n\n"
            "Bộ test case hiện tại (tham khảo phạm vi nghiệp vụ, có thể cải thiện "
            f"nội dung nhưng KHÔNG được ít hơn {min_count} test case):\n"
            f"{json.dumps(testcases, ensure_ascii=False, indent=2)}\n\n"
            "YÊU CẦU BẮT BUỘC:\n"
            "- Không tạo chức năng khác, không đổi tên module.\n"
            "- Không sử dụng test case của module khác.\n"
            f"- Trả về DUY NHẤT một JSON object dạng "
            '{"test_cases": [ {...}, {...} ]}'
            f" — mảng test_cases chứa toàn bộ test case của đúng 1 chức năng "
            f"\"{module_name}\", mỗi test case theo đúng field: "
            f"{self._REGENERATE_TC_SCHEMA_FIELDS}.\n"
            "- Không kèm giải thích, không markdown.\n"
        )
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.4,
                max_tokens=4000,
                response_format={"type": "json_object"},
            )
            content = response.choices[0].message.content or ""
            raw = self._parse_json_response(content)
        except Exception as exc:
            raise RuntimeError(f"Không sinh lại được chức năng: {exc}") from exc

        new_list = raw.get("test_cases") if isinstance(raw, dict) else None
        if not isinstance(new_list, list) or not new_list:
            raise RuntimeError("AI không trả về đúng định dạng danh sách test case")

        result: list[dict] = []
        for tc in new_list:
            if not isinstance(tc, dict):
                continue
            tc["module"] = module_name
            tc["chức năng"] = module_name
            tc["feature"] = module_name
            tc.setdefault("priority", "Trung bình")
            tc.setdefault("status", "Chưa chạy")
            tc.setdefault("test_type", "Kiểm thử chức năng")
            tc.setdefault("actual_result", "")
            tc.setdefault("note", "")
            result.append(tc)
        if not result:
            raise RuntimeError("AI không trả về test case hợp lệ nào")
        return result

    @_log_generation_runtime
    def generate_test_cases(
        self,
        description: str,
        previous_test_cases: dict | None = None,
        image_blocks: list[dict] | None = None,
        domain: str | None = None,
        context_mode: str = "new",
    ) -> dict:
        """
        TH1 (targeted): user gửi ảnh + yêu cầu chức năng cụ thể
            → chỉ sinh đúng số chức năng user yêu cầu, không dư.
        TH2 (full): user gửi ảnh không kèm mô tả (hoặc mô tả hệ thống)
            → sinh TC cho TẤT CẢ chức năng trong ảnh.

        domain: tên domain do UI truyền vào (vd "bank", "hospital", "school",
            "recruitment") — dùng cho Domain Rule Engine (rule_engine.py) để
            lấy đúng BUSINESS_RULES/TEST_CASES mẫu của domain đó. None/""
            → Domain Rule Engine tự tắt (không ảnh hưởng các bước khác).

        Pipeline 4 bước:
          1. _scan_image_elements   — OCR quét toàn bộ UI elements trong ảnh
          2. _select_relevant_rules — chọn elements liên quan (lọc theo mô tả
                                       nếu targeted) + rule engine áp dụng
          3. _coverage_checker      — kiểm tra scan có bỏ sót chức năng được
                                       yêu cầu không, re-scan legacy nếu cần
          4. _split_elements        — chỉ áp dụng cho ảnh nhiều UI (>12
                                       elements) ở TH2, để tránh truncate
        """
        context_mode = (context_mode or "new").strip().lower()
        if context_mode not in {"new", "screen_only", "workflow"}:
            context_mode = "new"
        if context_mode == "new":
            previous_test_cases = None
        self._current_screen_context = None
        self._current_workflow_relation = None
        self._current_workflow_context = None

        targeted = self._is_targeted_request(description)
        print(
            f"[1/6] 🧭 Xác định chế độ: "
            f"{'TARGETED' if targeted else 'FULL/TEXT'} | context={context_mode}"
        )
        scanned = None
        proj = description or 'Dự án'
        image_hash = self._hash_image_blocks(image_blocks)

        if image_blocks:
            scan_started = time.perf_counter()
            print("[2/6] 🔍 Đang quét và phân tích UI...")
            cached_scanned = self._scan_cache.get(image_hash) if image_hash else None
            if self._cache_covers_request(cached_scanned, description, targeted):
                scanned = cached_scanned
                print("=== SCAN CACHE HIT: ảnh giống ảnh trước, cache đủ chức năng yêu cầu → dùng lại, KHÔNG quét lại ảnh ===")
            else:
                reason = "ảnh mới (chưa có trong cache)" if not cached_scanned else "cache thiếu chức năng được yêu cầu"
                print(f"=== SCAN CACHE MISS ({reason}) → quét lại ảnh ===")
                scanned = self._scan_image_elements(image_blocks, description)
                if image_hash and scanned:
                    self._scan_cache[image_hash] = scanned
            print("=== SCANNED ELEMENTS ===\n" + (scanned or "(rỗng)") + "\n=========================")
            if scanned:
                for line in scanned.splitlines():
                    if line.startswith('PROJECT_NAME:'):
                        proj = line.replace('PROJECT_NAME:', '').strip() or proj
                        break
            print(f"      ✅ Scan UI hoàn tất ({time.perf_counter() - scan_started:.2f}s)")
            self._current_screen_context = self.vision_service.analyze_scan(
                scanned=scanned,
                description=description,
            )
            previous_workflow_context = None
            if context_mode == 'workflow' and isinstance(previous_test_cases, dict):
                previous_workflow_context = previous_test_cases.get('_workflow_context')
            if context_mode == 'workflow':
                self._current_workflow_relation = self.workflow_service.find_best_parent(
                    previous_context=previous_workflow_context,
                    current_screen=self._current_screen_context,
                )
            else:
                self._current_workflow_relation = {
                    'linked': False,
                    'score': 0.0,
                    'reason': f'context_mode={context_mode}: không nối với ảnh trước.',
                }
            self._current_workflow_context = self.workflow_service.build_context(
                previous_context=previous_workflow_context,
                current_screen=self._current_screen_context,
                relation=self._current_workflow_relation,
            )
            print(
                "[Workflow] "
                f"screen={self._current_screen_context.get('screen_type')} | "
                f"entity={self._current_screen_context.get('business_entity') or '(không rõ)'} | "
                f"linked={self._current_workflow_relation.get('linked')} | "
                f"score={self._current_workflow_relation.get('score')}"
            )
        if image_blocks and targeted and self._should_expand_modal_request(description, scanned or "", proj):
            targeted = False
            print("=== MODE AUTO: mô tả trùng tên modal/form → chuyển TARGETED thành FULL FORM ===")
        print("[3/6] ✅ Đang kiểm tra coverage và Rule Engine...")
        coverage_scan_started = time.perf_counter()
        relevant_elements, rule_hints = '', ''
        if image_blocks and scanned is not None:
            scanned, relevant_elements, rule_hints = self._coverage_checker(
                description, targeted, scanned, image_blocks,
            )
            if image_hash and scanned:
                self._scan_cache[image_hash] = scanned
        print(f"      ✅ Coverage/Rule Engine hoàn tất ({time.perf_counter() - coverage_scan_started:.2f}s)")
        self._current_form_structure = None
        self._current_crud_context = None
        if image_blocks and scanned:
            form_structure = self._analyze_form_structure(scanned, description)
            if form_structure:
                self._current_form_structure = form_structure
            self._current_crud_context = self._detect_request_crud_context(
                scanned=scanned,
                description=description,
                previous_test_cases=previous_test_cases,
            )
            print(
                "[CRUDContext] "
                f"action={self._current_crud_context.get('action') or '(không)'} | "
                f"screen_type={self._current_crud_context.get('screen_type')} | "
                f"has_form_fields={self._current_crud_context.get('has_form_fields')} | "
                f"actions={self._current_crud_context.get('detected_actions', [])} | "
                f"strong_list={self._current_crud_context.get('strong_list_evidence', False)}"
            )
            if self._current_crud_context.get('screen_type') == 'list':
                self._current_form_structure = None
            elif self._current_form_structure:
                rule_hints = (rule_hints or '') + self._build_form_structure_hint(
                    self._current_form_structure
                )
        print("[4/6] 📚 Đang truy xuất RAG và domain rules...")
        rag_started = time.perf_counter()
        rag_query = (
            f"Mô tả người dùng:\n{description}\n\n"
            f"Tên dự án/màn hình:\n{proj}\n\n"
            f"UI đã scan:\n{relevant_elements or scanned or ''}\n\n"
            f"Rule Engine:\n{rule_hints or ''}\n"
        )
        rag_context = self._retrieve_rag_context(rag_query, top_k=5, targeted=targeted)
        domain_hints = self._get_domain_hints(domain, description, proj, relevant_elements or scanned or '')
        rule_hints = (rule_hints or '') + domain_hints
        crud_search_text = " ".join([description or '', proj or '', relevant_elements or scanned or ''])
        crud_hints = ''
        if _re_build_crud_rule_prompt is not None:
            try:
                previous_modules = (
                    previous_test_cases.get('modules', {})
                    if isinstance(previous_test_cases, dict) else None
                )
                crud_action = (
                    (self._current_crud_context or {}).get('action')
                    or (
                        _re_detect_crud_action(crud_search_text)
                        if _re_detect_crud_action is not None else None
                    )
                )
                crud_hints = _re_build_crud_rule_prompt(
                    action=crud_action,
                    search_text=crud_search_text,
                    previous_modules=previous_modules,
                )
            except Exception as exc:
                print(f"[RuleEngine] Bỏ qua CRUD hints do lỗi: {exc}")
        rule_hints = (rule_hints or '') + (crud_hints or '')
        if self._current_crud_context and self._current_crud_context.get('screen_type') == 'list':
            rule_hints += (
                "\n=== QUY TẮC ƯU TIÊN CAO NHẤT CHO CRUD TRÊN MÀN HÌNH DANH SÁCH ===\n"
                "- Mỗi chức năng Thêm mới, Cập nhật hoặc Xóa xuất hiện trong ảnh chỉ có ĐÚNG 2 testcase.\n"
                "- Thêm mới: mở form thành công / không mở được.\n"
                "- Cập nhật: mở đúng form và dữ liệu bản ghi / không mở được.\n"
                "- Xóa: mở popup xác nhận đúng đối tượng / không mở được.\n"
                "- KHÔNG sinh nhập liệu, lưu dữ liệu, Required, Format, Duplicate, Boundary, XSS/SQLi khi chưa thấy form/popup.\n"
                "- Ô tìm kiếm/lọc trên danh sách KHÔNG phải field của form CRUD.\n"
                "- Quay lại trên danh sách chỉ kiểm tra điều hướng, không giả định đang nhập hoặc sửa dữ liệu.\n"
                "- Quy tắc này ghi đè mọi checklist CRUD form tổng quát khác trong system prompt.\n"
            )
        if self._current_screen_context:
            rule_hints += self.vision_service.build_generation_hint(
                self._current_screen_context,
                self._current_workflow_relation,
            )
        if self._current_workflow_relation:
            rule_hints += self.workflow_service.build_generation_hint(
                self._current_workflow_relation
            )
        print(f"      ✅ RAG/domain rules hoàn tất ({time.perf_counter() - rag_started:.2f}s)")
        print("[5/6] 🤖 Đang gọi AI sinh testcase...")
        # TH1: TARGETED 
        if image_blocks and targeted and not previous_test_cases:
            prompt = (
                f"Màn hình: \"{proj}\"\n"
                f"Sinh testcase CHỈ cho các chức năng sau (ĐÚNG số module, KHÔNG hơn):\n"
                f"{description}\n\n"
                f"=== UI ELEMENTS HỖ TRỢ (chỉ dùng để lấy tên trường, data test) ===\n"
                f"{relevant_elements}\n"
                f"project_name = \"{proj}\""
                f"{rule_hints}"
            )
            prompt = self._append_rag_context(prompt, rag_context)
            result = self._call_api(prompt, image_blocks, self._build_system_prompt(SYSTEM_PROMPT_TARGETED))
            if result:
                result = self._filter_modules_by_description(result, description)
                result = self._enforce_min_coverage(
                    result, scanned, image_blocks,
                    self._build_system_prompt(SYSTEM_PROMPT_TARGETED), proj,
                    targeted=True,
                )
                result = self._filter_modules_by_description(result, description)
                print("[6/6] 🔧 Đang chuẩn hóa kết quả và lập coverage report...")
                normalized = self._normalize_test_cases(result, description=description)
                return self._build_final_coverage_report(
                    normalized, scanned, image_blocks, targeted=True, description=description,
                )
            raise Exception("Không thể sinh testcase (targeted) sau khi đã retry tối đa.")
        if image_blocks and not targeted and not previous_test_cases:
            element_lines = [l for l in (relevant_elements or '').splitlines() if l.strip().startswith('-')]
            n_elements = len(element_lines)
            if n_elements > 12:
                group1, group2 = self._split_elements(relevant_elements)
                batch1 = batch2 = None
                try:
                    batch1 = self._generate_batch('', group1, proj, image_blocks, id_offset=0, system_prompt=self._build_system_prompt(SYSTEM_PROMPT_FULL), domain=domain)
                except Exception:
                    pass
                tc_count = sum(len(v) for v in (batch1 or {}).get('modules', {}).values() if isinstance(v, list))
                try:
                    batch2 = self._generate_batch('', group2, proj, image_blocks, id_offset=tc_count, system_prompt=self._build_system_prompt(SYSTEM_PROMPT_FULL), domain=domain)
                except Exception:
                    pass
                if batch1 and batch2:
                    merged = self._merge_test_cases(batch1, batch2)
                    merged = self._enforce_min_coverage(
                        merged, scanned, image_blocks,
                        self._build_system_prompt(SYSTEM_PROMPT_FULL), proj,
                    )
                    print("[6/6] 🔧 Đang chuẩn hóa kết quả và lập coverage report...")
                    normalized = self._normalize_test_cases(merged, description=description)
                    return self._build_final_coverage_report(
                        normalized, scanned, image_blocks, targeted=False, description=description,
                    )
                elif batch1:
                    batch1 = self._enforce_min_coverage(
                        batch1, scanned, image_blocks,
                        self._build_system_prompt(SYSTEM_PROMPT_FULL), proj,
                    )
                    print("[6/6] 🔧 Đang chuẩn hóa kết quả và lập coverage report...")
                    normalized = self._normalize_test_cases(batch1, description=description)
                    return self._build_final_coverage_report(
                        normalized, scanned, image_blocks, targeted=False, description=description,
                    )
                else:
                    raise Exception("Không thể sinh testcase (full, split-batch) sau khi đã retry tối đa.")
            else:
                prompt = (
                    f"Phân tích ảnh giao diện màn hình \"{proj}\" và sinh testcase "
                    f"cho TẤT CẢ chức năng xuất hiện trong ảnh.\n\n"
                    f"=== UI ELEMENTS ===\n{relevant_elements}\n"
                    f"project_name = \"{proj}\""
                    f"{rule_hints}"
                )
                prompt = self._append_rag_context(prompt, rag_context)
                result = self._call_api(prompt, image_blocks, self._build_system_prompt(SYSTEM_PROMPT_FULL), is_full_scan=True)
                if result:
                    result = self._enforce_min_coverage(
                        result, scanned, image_blocks,
                        self._build_system_prompt(SYSTEM_PROMPT_FULL), proj,
                    )
                    print("[6/6] 🔧 Đang chuẩn hóa kết quả và lập coverage report...")
                    normalized = self._normalize_test_cases(result, description=description)
                    return self._build_final_coverage_report(
                        normalized, scanned, image_blocks, targeted=False, description=description,
                    )
                raise Exception("Không thể sinh testcase (full) sau khi đã retry tối đa.")
        if image_blocks:
            base_prompt = SYSTEM_PROMPT_TARGETED if targeted else SYSTEM_PROMPT_FULL
        else:
            base_prompt = SYSTEM_PROMPT_TARGETED if targeted else SYSTEM_PROMPT_TEXT_ONLY
        sys_prompt = self._build_system_prompt(base_prompt)
        sys_prompt += rule_hints or self._build_rule_engine_hints(scanned or '')
        description_with_rag = self._append_rag_context(description, rag_context)
        messages = self._build_messages(
            description_with_rag,
            previous_test_cases=previous_test_cases,
            image_blocks=image_blocks,
            system_prompt=sys_prompt,
            context_mode=context_mode,
        )
        last_error = None
        for attempt in range(self.max_retries):
            try:
                response = self.client.chat.completions.create(
                    model=self.model,
                    messages=messages,
                    temperature=0.1,
                    top_p=0.95,
                    max_tokens=self._get_max_tokens(),
                )
                finish_reason = response.choices[0].finish_reason
                content_str = response.choices[0].message.content
                if finish_reason == 'length':
                    last_error = "Response bị truncate."
                    if attempt < self.max_retries - 1:
                        retry_prefix = (
                            "QUAN TRỌNG: Response lần trước bị cắt. "
                            "Sinh JSON NGẮN GỌN hơn, mỗi trường 1 câu, steps 3 bước. "
                            "PHẢI đóng JSON đầy đủ.\n\n"
                        )
                        prev = messages[-1].get("content")
                        if isinstance(prev, str):
                            messages[-1]["content"] = retry_prefix + prev
                        elif isinstance(prev, list):
                            messages[-1]["content"] = [{"type": "text", "text": retry_prefix}, *prev]
                        time.sleep(2 ** attempt)
                        continue
                    raise Exception("Response bị truncate sau tất cả retry.")
                raw = self._parse_json_response(content_str)
                result = self._unwrap_modules(raw)

                if targeted:
                    result = self._filter_modules_by_description(result, description)
                if previous_test_cases and context_mode == "workflow":
                    new_result = self._enforce_min_coverage(
                        result, scanned, image_blocks, sys_prompt, proj,
                        targeted=targeted, description=description,
                    )
                    new_count = sum(
                        len(v) for v in (new_result or {}).get('modules', {}).values()
                        if isinstance(v, list)
                    )
                    old_count = sum(
                        len(v) for v in (previous_test_cases or {}).get('modules', {}).values()
                        if isinstance(v, list)
                    )
                    result = self._merge_test_cases(previous_test_cases, new_result or result)
                    merged_count = sum(
                        len(v) for v in (result or {}).get('modules', {}).values()
                        if isinstance(v, list)
                    )
                    print(
                        "[WorkflowMerge] Coverage phần mới đã chạy trước merge | "
                        f"old_tc={old_count} | new_tc={new_count} | merged_tc={merged_count}"
                    )
                else:
                    result = self._enforce_min_coverage(
                        result, scanned, image_blocks, sys_prompt, proj,
                        targeted=targeted, description=description,
                    )
                if targeted and (not previous_test_cases or context_mode == "screen_only"):
                    result = self._filter_modules_by_description(result, description)
                print("[6/6] 🔧 Đang chuẩn hóa kết quả và lập coverage report...")
                normalized = self._normalize_test_cases(
                    result, apply_static_filter=bool(image_blocks), description=description,
                )
                return self._build_final_coverage_report(
                    normalized, scanned, image_blocks, targeted=targeted, description=description,
                )

            except json.JSONDecodeError as e:
                last_error = f"JSON không hợp lệ: {str(e)}"
            except Exception as e:
                last_error = str(e)
            if attempt < self.max_retries - 1:
                time.sleep(2 ** attempt)
        raise Exception(f"Không thể sinh test case sau {self.max_retries} lần thử. Lỗi: {last_error}")